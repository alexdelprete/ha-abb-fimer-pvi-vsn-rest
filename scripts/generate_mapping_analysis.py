#!/usr/bin/env python3
"""Generate an XLSX analysis file from the updated VSN-to-SunSpec mapping.

This creates a spreadsheet showing all mapping entries with their metadata
for easy review and analysis.
"""

import json
from pathlib import Path
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def create_mapping_analysis_xlsx(mapping_path: Path, output_path: Path):
    """Create an XLSX file analyzing the mapping."""

    print(f"Reading mapping file: {mapping_path}")
    with open(mapping_path, encoding="utf-8") as f:
        mapping = json.load(f)

    print(f"Total entries in mapping: {len(mapping)}")

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Full mapping
    ws = wb.active
    ws.title = "Full Mapping"

    # Define headers
    headers = [
        "REST Name (VSN700)",
        "REST Name (VSN300)",
        "SunSpec Normalized Name",
        "HA Name",
        "HA Display Name",
        "Category",
        "HA Device Class",
        "HA Unit of Measurement",
        "HA State Class",
        "Suggested Display Precision",
        "In /livedata",
        "In /feeds",
        "Available in Modbus",
        "Description",
        "Models",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, entry in enumerate(mapping, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("REST Name (VSN700)", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("REST Name (VSN300)", ""))
        ws.cell(row=row_idx, column=3, value=entry.get("SunSpec Normalized Name", ""))
        ws.cell(row=row_idx, column=4, value=entry.get("HA Name", ""))
        ws.cell(row=row_idx, column=5, value=entry.get("HA Display Name", ""))
        ws.cell(row=row_idx, column=6, value=entry.get("Category", ""))
        ws.cell(row=row_idx, column=7, value=entry.get("HA Device Class", ""))
        ws.cell(row=row_idx, column=8, value=entry.get("HA Unit of Measurement", ""))
        ws.cell(row=row_idx, column=9, value=entry.get("HA State Class", ""))
        ws.cell(row=row_idx, column=10, value=entry.get("Suggested Display Precision", ""))
        ws.cell(row=row_idx, column=11, value=entry.get("In /livedata", ""))
        ws.cell(row=row_idx, column=12, value=entry.get("In /feeds", ""))
        ws.cell(row=row_idx, column=13, value=entry.get("Available in Modbus", ""))
        ws.cell(row=row_idx, column=14, value=entry.get("Description", ""))
        ws.cell(row=row_idx, column=15, value=", ".join(entry.get("models", [])))

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Sheet 2: Summary statistics
    ws2 = wb.create_sheet("Summary Statistics")

    # Count statistics
    total = len(mapping)
    with_device_class = sum(1 for e in mapping if e.get("HA Device Class", ""))
    with_unit = sum(1 for e in mapping if e.get("HA Unit of Measurement", ""))
    with_state_class = sum(1 for e in mapping if e.get("HA State Class", ""))
    with_precision = sum(1 for e in mapping if "Suggested Display Precision" in e)

    # Device class breakdown
    device_class_counts = {}
    for entry in mapping:
        dc = entry.get("HA Device Class", "") or "(none)"
        device_class_counts[dc] = device_class_counts.get(dc, 0) + 1

    # Category breakdown
    category_counts = {}
    for entry in mapping:
        cat = entry.get("Category", "") or "(none)"
        category_counts[cat] = category_counts.get(cat, 0) + 1

    # Write summary
    ws2.cell(row=1, column=1, value="Mapping Summary Statistics").font = Font(bold=True, size=14)

    row = 3
    ws2.cell(row=row, column=1, value="Total Entries").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=total)
    row += 1

    ws2.cell(row=row, column=1, value="With Device Class").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=with_device_class)
    ws2.cell(row=row, column=3, value=f"{with_device_class / total * 100:.1f}%")
    row += 1

    ws2.cell(row=row, column=1, value="With Unit").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=with_unit)
    ws2.cell(row=row, column=3, value=f"{with_unit / total * 100:.1f}%")
    row += 1

    ws2.cell(row=row, column=1, value="With State Class").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=with_state_class)
    ws2.cell(row=row, column=3, value=f"{with_state_class / total * 100:.1f}%")
    row += 1

    ws2.cell(row=row, column=1, value="With Precision").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=with_precision)
    ws2.cell(row=row, column=3, value=f"{with_precision / total * 100:.1f}%")
    row += 2

    # Device class breakdown
    ws2.cell(row=row, column=1, value="Device Class Breakdown").font = Font(bold=True, size=12)
    row += 1
    ws2.cell(row=row, column=1, value="Device Class").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1

    for dc, count in sorted(device_class_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=1, value=dc)
        ws2.cell(row=row, column=2, value=count)
        row += 1

    row += 1

    # Category breakdown
    ws2.cell(row=row, column=1, value="Category Breakdown").font = Font(bold=True, size=12)
    row += 1
    ws2.cell(row=row, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1

    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=count)
        row += 1

    # Auto-size columns in summary sheet
    for col_idx in range(1, 4):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 30

    # Save workbook
    print(f"\nSaving analysis to: {output_path}")
    wb.save(output_path)

    print("\n" + "=" * 60)
    print("MAPPING ANALYSIS SUMMARY")
    print("=" * 60)
    print(f"Total entries: {total}")
    print(f"With device_class: {with_device_class} ({with_device_class / total * 100:.1f}%)")
    print(f"With unit: {with_unit} ({with_unit / total * 100:.1f}%)")
    print(f"With state_class: {with_state_class} ({with_state_class / total * 100:.1f}%)")
    print(f"With precision: {with_precision} ({with_precision / total * 100:.1f}%)")
    print("=" * 60)


def main():
    """Main entry point."""
    # Get repository root
    repo_root = Path(__file__).parent.parent
    mapping_path = (
        repo_root
        / "custom_components"
        / "abb_fimer_pvi_vsn_rest"
        / "abb_fimer_vsn_rest_client"
        / "data"
        / "vsn-sunspec-point-mapping.json"
    )
    output_path = repo_root / "docs" / "vsn-data" / "vsn-sunspec-mapping-updated.xlsx"

    if not mapping_path.exists():
        print(f"Error: Mapping file not found at {mapping_path}")
        sys.exit(1)

    # Create output directory if needed
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Generate analysis
    create_mapping_analysis_xlsx(mapping_path, output_path)

    print(f"\n✓ Analysis XLSX file created: {output_path}")


if __name__ == "__main__":
    main()
