#!/usr/bin/env python3
"""Update HA Display Names in VSN-SunSpec point mapping Excel file."""

from pathlib import Path
import openpyxl


def update_ha_display_names():
    """Apply HA Display Name updates and fix data issues in the mapping Excel file."""
    excel_path = Path("docs/vsn-sunspec-point-mapping.xlsx")

    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["All Points"]

    # Find column indices
    headers = [cell.value for cell in ws[1]]
    ha_display_col = headers.index("HA Display Name") + 1
    label_col = headers.index("Label") + 1
    device_class_col = headers.index("HA Device Class") + 1
    unit_col = headers.index("HA Unit of Measurement") + 1
    entity_cat_col = headers.index("Entity Category") + 1

    # Note: HA Suggested Precision column doesn't exist in current Excel

    # Define display name updates based on current HA Display Name values
    display_name_updates = {
        # Datalogger entities
        "Serial number (datalogger)": "Serial Number",
        "WiFi local IP address": "IP address",
        "WiFi network name (SSID)": "WiFi SSID",
        "Wlan 0 Mode operating mode": "Wlan 0 mode",
        # Keep System Load as-is but add precision

        # Inverter entities
        "Current alarm status of the inverter": "Alarm status",
        "Booster stage temperature": "Booster temperature",
        "DC bulk capacitor voltage": "DC capacitor voltage",
        "DC bulk mid-point voltage": "DC mid-point voltage",
        "DC current measurement for string 1": "DC current #1",
        "DC current measurement for string 2": "DC current #2",
        "DC voltage measurement for string 1": "DC voltage #1",
        "DC voltage measurement for string 2": "DC voltage #2",
        "Phase A to neutral voltage measurement": "Phase Voltage AN",
        "Device Modbus address": "Modbus address",
        "Device model identifier from common model": "Model",
        "Device options and features from common model": "Options",
        "Device serial number from common model": "Serial Number",
        "Firmware version from device common model": "Firmware Version",
        "Manufacturer name from device common model": "Manufacturer",
        # System timestamp needs special handling
    }

    update_count = 0
    fix_count = 0

    # Iterate through all rows (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        current_display_name = row[ha_display_col - 1].value
        label = row[label_col - 1].value

        # Update display names
        if current_display_name in display_name_updates:
            new_display_name = display_name_updates[current_display_name]
            row[ha_display_col - 1].value = new_display_name
            print(f"Row {row_idx}: Updated '{current_display_name}' → '{new_display_name}'")
            update_count += 1

        # Special cases requiring additional fixes
        if label == "System Load":
            # Note: Would set precision to 2 but column doesn't exist in Excel
            print(f"Row {row_idx}: System Load found (precision setting skipped - column not available)")

        elif label == "Device Address" and current_display_name == "Device Modbus address":
            # Fix wrong device_class and unit for Modbus address
            if row[device_class_col - 1].value == "current":
                row[device_class_col - 1].value = None
                row[unit_col - 1].value = None
                print(f"Row {row_idx}: Fixed device_class and unit for Modbus address")
                fix_count += 1

        elif label == "Version" and "Firmware version" in str(current_display_name):
            # Fix wrong device_class and unit for Firmware Version
            if row[device_class_col - 1].value == "voltage":
                row[device_class_col - 1].value = None
                row[unit_col - 1].value = None
                print(f"Row {row_idx}: Fixed device_class and unit for Firmware Version")
                fix_count += 1

        elif label == "Sys Time":
            # Convert System timestamp to datetime sensor
            row[ha_display_col - 1].value = "System timestamp"
            row[device_class_col - 1].value = "timestamp"
            row[entity_cat_col - 1].value = "diagnostic"
            print(f"Row {row_idx}: Configured System timestamp as datetime sensor")
            fix_count += 1
            update_count += 1  # Also counts as display name update

    # Save the updated Excel file
    print(f"\nSaving updated Excel file...")
    wb.save(excel_path)

    print(f"\n✓ Successfully updated {update_count} display names")
    print(f"✓ Applied {fix_count} data fixes")
    print(f"Total changes: {update_count + fix_count}")

    return update_count, fix_count


if __name__ == "__main__":
    updates, fixes = update_ha_display_names()
    print(f"\nDone! Ready to regenerate JSON files with convert_excel_to_json.py")