#!/usr/bin/env python3
"""Generate comprehensive VSN-SunSpec point mapping Excel file."""

import json
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Complete mapping: VSN700 → SunSpec
VSN_TO_SUNSPEC_MAP = {
    # ===== Standard SunSpec Models (Available in BOTH Modbus + REST) =====
    # M103 - Inverter
    "Pgrid": {
        "sunspec": "W",
        "modbus": "m103_1_W",
        "model": "M103",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "Igrid": {
        "sunspec": "A",
        "modbus": "m103_1_A",
        "model": "M103",
        "category": "Inverter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "Vgrid": {
        "sunspec": "PhVphA",
        "modbus": "m103_1_PhVphA",
        "model": "M103",
        "category": "Inverter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "Fgrid": {
        "sunspec": "Hz",
        "modbus": "m103_1_Hz",
        "model": "M103",
        "category": "Inverter",
        "units": "Hz",
        "state_class": "measurement",
        "device_class": "frequency",
        "in_modbus": "YES",
    },
    "cosPhi": {
        "sunspec": "PF",
        "modbus": "m103_1_PF",
        "model": "M103",
        "category": "Inverter",
        "units": "%",
        "state_class": "measurement",
        "device_class": "power_factor",
        "in_modbus": "YES",
    },
    "Qgrid": {
        "sunspec": "VAR",
        "modbus": "m103_1_VAR",
        "model": "M103",
        "category": "Inverter",
        "units": "var",
        "state_class": "measurement",
        "device_class": "reactive_power",
        "in_modbus": "YES",
    },
    "Pin": {
        "sunspec": "DCW",
        "modbus": "m103_1_DCW",
        "model": "M103",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "ETotal": {
        "sunspec": "WH",
        "modbus": "m103_1_WH",
        "model": "M103",
        "category": "Inverter",
        "units": "Wh",
        "state_class": "total_increasing",
        "device_class": "energy",
        "in_modbus": "YES",
    },
    "Ein": {
        "sunspec": None,
        "modbus": None,
        "model": "VSN",
        "category": "Inverter",
        "units": "Wh",
        "state_class": "total_increasing",
        "device_class": "energy",
        "in_modbus": "NO",
    },
    "TempInv": {
        "sunspec": "TmpOt",
        "modbus": "m103_1_TmpOt",
        "model": "M103",
        "category": "Inverter",
        "units": "°C",
        "state_class": "measurement",
        "device_class": "temperature",
        "in_modbus": "YES",
    },
    "Temp1": {
        "sunspec": "TmpCab",
        "modbus": "m103_1_TmpCab",
        "model": "M103",
        "category": "Inverter",
        "units": "°C",
        "state_class": "measurement",
        "device_class": "temperature",
        "in_modbus": "YES",
    },
    # M120 - Nameplate
    "WRtg": {
        "sunspec": "WRtg",
        "modbus": "m120_1_WRtg",
        "model": "M120",
        "category": "Inverter",
        "units": "W",
        "state_class": None,
        "device_class": "power",
        "in_modbus": "YES",
    },
    # M160 - MPPT
    "Iin1": {
        "sunspec": "DCA_1",
        "modbus": "m160_1_DCA_1",
        "model": "M160",
        "category": "MPPT",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "Vin1": {
        "sunspec": "DCV_1",
        "modbus": "m160_1_DCV_1",
        "model": "M160",
        "category": "MPPT",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "Pin1": {
        "sunspec": "DCW_1",
        "modbus": "m160_1_DCW_1",
        "model": "M160",
        "category": "MPPT",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "Iin2": {
        "sunspec": "DCA_2",
        "modbus": "m160_1_DCA_2",
        "model": "M160",
        "category": "MPPT",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "Vin2": {
        "sunspec": "DCV_2",
        "modbus": "m160_1_DCV_2",
        "model": "M160",
        "category": "MPPT",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "Pin2": {
        "sunspec": "DCW_2",
        "modbus": "m160_1_DCW_2",
        "model": "M160",
        "category": "MPPT",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    # M124 - Storage
    "TSoc": {
        "sunspec": "ChaState",
        "modbus": "m124_1_ChaState",
        "model": "M124",
        "category": "Battery",
        "units": "%",
        "state_class": "measurement",
        "device_class": "battery",
        "in_modbus": "YES",
    },
    "PbaT": {
        "sunspec": "InWRte",
        "modbus": "m124_1_InWRte",
        "model": "M124",
        "category": "Battery",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    # M802 - Battery
    "Soc": {
        "sunspec": "SoC",
        "modbus": "m802_1_SoC",
        "model": "M802",
        "category": "Battery",
        "units": "%",
        "state_class": "measurement",
        "device_class": "battery",
        "in_modbus": "YES",
    },
    "Soh": {
        "sunspec": "SoH",
        "modbus": "m802_1_SoH",
        "model": "M802",
        "category": "Battery",
        "units": "%",
        "state_class": "measurement",
        "device_class": None,
        "in_modbus": "YES",
    },
    "Vba": {
        "sunspec": "V",
        "modbus": "m802_1_V",
        "model": "M802",
        "category": "Battery",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "Iba": {
        "sunspec": "A",
        "modbus": "m802_1_A",
        "model": "M802",
        "category": "Battery",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "Pba": {
        "sunspec": "W",
        "modbus": "m802_1_W",
        "model": "M802",
        "category": "Battery",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "Tba": {
        "sunspec": "Tmp",
        "modbus": "m802_1_Tmp",
        "model": "M802",
        "category": "Battery",
        "units": "°C",
        "state_class": "measurement",
        "device_class": "temperature",
        "in_modbus": "YES",
    },
    "CycleNum": {
        "sunspec": "NCyc",
        "modbus": "m802_1_NCyc",
        "model": "M802",
        "category": "Battery",
        "units": "",
        "state_class": "total",
        "device_class": None,
        "in_modbus": "YES",
    },
    "VcMax": {
        "sunspec": "CellVMax",
        "modbus": "m802_1_CellVMax",
        "model": "M802",
        "category": "Battery",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "VcMin": {
        "sunspec": "CellVMin",
        "modbus": "m802_1_CellVMin",
        "model": "M802",
        "category": "Battery",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "TcMax": {
        "sunspec": "CellTmpMax",
        "modbus": "m802_1_CellTmpMax",
        "model": "M802",
        "category": "Battery",
        "units": "°C",
        "state_class": "measurement",
        "device_class": "temperature",
        "in_modbus": "YES",
    },
    "TcMin": {
        "sunspec": "CellTmpMin",
        "modbus": "m802_1_CellTmpMin",
        "model": "M802",
        "category": "Battery",
        "units": "°C",
        "state_class": "measurement",
        "device_class": "temperature",
        "in_modbus": "YES",
    },
    "PCh": {
        "sunspec": "AChaMax",
        "modbus": "m802_1_AChaMax",
        "model": "M802",
        "category": "Battery",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "PDh": {
        "sunspec": "ADisChaMax",
        "modbus": "m802_1_ADisChaMax",
        "model": "M802",
        "category": "Battery",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "Fcc": {
        "sunspec": "AHRtg",
        "modbus": "m802_1_AHRtg",
        "model": "M802",
        "category": "Battery",
        "units": "Ah",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
    },
    # M101/102 - Split Phase
    "VgridL1_N": {
        "sunspec": "PhVphA",
        "modbus": "m101_1_PhVphA",
        "model": "M101",
        "category": "Inverter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "VgridL2_N": {
        "sunspec": "PhVphB",
        "modbus": "m101_1_PhVphB",
        "model": "M101",
        "category": "Inverter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "SplitPhase": {
        "sunspec": None,
        "modbus": None,
        "model": "M101",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
    },
    # M203 - Meter
    "MeterPgrid_Tot": {
        "sunspec": "W",
        "modbus": "m203_1_W",
        "model": "M203",
        "category": "Meter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "MeterPgrid_L1": {
        "sunspec": "WphA",
        "modbus": "m203_1_WphA",
        "model": "M203",
        "category": "Meter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "MeterPgrid_L2": {
        "sunspec": "WphB",
        "modbus": "m203_1_WphB",
        "model": "M203",
        "category": "Meter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "MeterPgrid_L3": {
        "sunspec": "WphC",
        "modbus": "m203_1_WphC",
        "model": "M203",
        "category": "Meter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
    },
    "MeterVgrid_L1": {
        "sunspec": "PhVphA",
        "modbus": "m203_1_PhVphA",
        "model": "M203",
        "category": "Meter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "MeterVgrid_L2": {
        "sunspec": "PhVphB",
        "modbus": "m203_1_PhVphB",
        "model": "M203",
        "category": "Meter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "MeterVgrid_L3": {
        "sunspec": "PhVphC",
        "modbus": "m203_1_PhVphC",
        "model": "M203",
        "category": "Meter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
    },
    "MeterIgrid_L1": {
        "sunspec": "AphA",
        "modbus": "m203_1_AphA",
        "model": "M203",
        "category": "Meter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "MeterIgrid_L2": {
        "sunspec": "AphB",
        "modbus": "m203_1_AphB",
        "model": "M203",
        "category": "Meter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "MeterIgrid_L3": {
        "sunspec": "AphC",
        "modbus": "m203_1_AphC",
        "model": "M203",
        "category": "Meter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
    },
    "MeterQgrid_Tot": {
        "sunspec": "VAR",
        "modbus": "m203_1_VAR",
        "model": "M203",
        "category": "Meter",
        "units": "var",
        "state_class": "measurement",
        "device_class": "reactive_power",
        "in_modbus": "YES",
    },
    "MeterQgrid_L1": {
        "sunspec": "VARphA",
        "modbus": "m203_1_VARphA",
        "model": "M203",
        "category": "Meter",
        "units": "var",
        "state_class": "measurement",
        "device_class": "reactive_power",
        "in_modbus": "YES",
    },
    "MeterQgrid_L2": {
        "sunspec": "VARphB",
        "modbus": "m203_1_VARphB",
        "model": "M203",
        "category": "Meter",
        "units": "var",
        "state_class": "measurement",
        "device_class": "reactive_power",
        "in_modbus": "YES",
    },
    "MeterQgrid_L3": {
        "sunspec": "VARphC",
        "modbus": "m203_1_VARphC",
        "model": "M203",
        "category": "Meter",
        "units": "var",
        "state_class": "measurement",
        "device_class": "reactive_power",
        "in_modbus": "YES",
    },
    "MeterFgrid": {
        "sunspec": "Hz",
        "modbus": "m203_1_Hz",
        "model": "M203",
        "category": "Meter",
        "units": "Hz",
        "state_class": "measurement",
        "device_class": "frequency",
        "in_modbus": "YES",
    },
    "EGridImport": {
        "sunspec": "TotWhImp",
        "modbus": "m203_1_TotWhImp",
        "model": "M203",
        "category": "Meter",
        "units": "Wh",
        "state_class": "total_increasing",
        "device_class": "energy",
        "in_modbus": "YES",
    },
    "EGridExport": {
        "sunspec": "TotWhExp",
        "modbus": "m203_1_TotWhExp",
        "model": "M203",
        "category": "Meter",
        "units": "Wh",
        "state_class": "total_increasing",
        "device_class": "energy",
        "in_modbus": "YES",
    },
}

# M64061 points - will be added with abb_ prefix
M64061_POINTS = {
    "AlarmState",
    "DC1State",
    "DC2State",
    "GlobState",
    "InvState",
    "IleakInv",
    "IleakDC",
    "Riso",
    "VBulk",
    "VBulkMid",
    "Vgnd",
    "TempBst",
    "SysTime",
    "Ppeak",
    "CountryStd",
}

# ABB Proprietary (non-M64061)
ABB_PROPRIETARY = {
    "BattExtCtrlEna",
    "BattExtCtrlState",
    "BattNum",
    "BatteryMode",
    "Chc",
    "ClockState",
    "CommissioningFreeze",
    "DI0_mode",
    "DI1_mode",
    "DI_status",
    "DetectedBatteryNumber",
    "Dhc",
    "DynamicFeedInCtrl",
    "EBackup",
    "ECharge",
    "ECt",
    "EDischarge",
    "EDt",
    "ETotalAbsorbed",
    "ETotalApparent",
    "Ein1",
    "Ein2",
    "EnergyPolicy",
    "FRT_state",
    "Fan1rpm",
    "Fan2rpm",
    "HouseIgrid_L1",
    "HouseIgrid_L2",
    "HouseIgrid_L3",
    "HousePInverter",
    "HousePgrid_L1",
    "HousePgrid_L2",
    "HousePgrid_L3",
    "HousePgrid_Tot",
    "InputMode",
    "NumOfMPPT",
    "PACDeratingFlags",
    "PacStandAlone",
    "PacTogrid",
    "QACDeratingFlags",
    "SACDeratingFlags",
    "ShU",
    "WarningFlags",
    "gridExtCtrlEna",
    "gridExtCtrlState",
    "m126Mod_Ena",
    "m132Mod_Ena",
}


# Helper functions for title classification and data loading
def is_title_a_description(title):
    """Determine if a title field is a description (True) or a point name reference (False)."""
    if not title or title.strip() == "":
        return False

    # Contains ' - ' separator (e.g., "E3 - Energy to the Grid - 30D")
    if " - " in title:
        return True

    # Multiple words (more than 2) suggests description
    words = title.split()
    if len(words) > 2:
        return True

    # System monitoring descriptions
    if title in [
        "flash free",
        "ram free",
        "Version",
        "flash used",
        "system load",
        "uptime",
    ]:
        return True

    # Short capitalized phrases that are point names (IGrid, Pin, Pout, VoutR, etc.)
    # If it's short and alphanumeric (with spaces), likely a point name reference
    if len(title) <= 15 and " - " not in title:
        return False

    # Default: if it has spaces and is longer, it's likely a description
    return " " in title and len(title) > 10


def load_feeds_titles(vsn300_feeds_path, vsn700_feeds_path):
    """Load title and unit data from VSN feeds files and categorize them."""
    print("Loading feeds title data...")

    with open(vsn300_feeds_path) as f:
        vsn300_feeds = json.load(f)
    with open(vsn700_feeds_path) as f:
        vsn700_feeds = json.load(f)

    titles = {}

    # Process VSN300 feeds
    for feed_data in vsn300_feeds["feeds"].values():
        if "datastreams" in feed_data:
            for point_name, point_data in feed_data["datastreams"].items():
                title = point_data.get("title", "").strip()
                units = point_data.get("units", "").strip()
                if title:
                    titles[point_name] = {
                        "title": title,
                        "units": units,
                        "is_description": is_title_a_description(title),
                        "source": "VSN300",
                    }

    # Process VSN700 feeds
    for feed_data in vsn700_feeds["feeds"].values():
        if "datastreams" in feed_data:
            for point_name, point_data in feed_data["datastreams"].items():
                title = point_data.get("title", "").strip()
                units = point_data.get("units", "").strip()
                if title and point_name not in titles:
                    titles[point_name] = {
                        "title": title,
                        "units": units,
                        "is_description": is_title_a_description(title),
                        "source": "VSN700",
                    }
                elif title and point_name in titles:
                    titles[point_name]["source"] = "Both"
                    # Update units if not already set
                    if units and not titles[point_name].get("units"):
                        titles[point_name]["units"] = units

    desc_count = sum(1 for t in titles.values() if t["is_description"])
    name_count = sum(1 for t in titles.values() if not t["is_description"])

    print(f"  Found {len(titles)} points with titles")
    print(f"  {desc_count} are descriptions, {name_count} are point name references")

    return titles


# Helper functions for label/description generation
def load_sunspec_models_metadata(workbook_path):
    """Load label and description data from SunSpec models workbook."""
    print(f"Loading SunSpec models metadata from {workbook_path}...")

    models_data = {}
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    # Models we care about
    model_sheets = [
        "1",
        "101",
        "103",
        "120",
        "124",
        "160",
        "201",
        "203",
        "204",
        "802",
        "803",
        "804",
    ]

    for sheet_name in model_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        model_points = {}

        # Parse point data (skip header row)
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=3).value  # Column C: Name
            label = ws.cell(row=row_idx, column=13).value  # Column M: Label
            description = ws.cell(row=row_idx, column=14).value  # Column N: Description

            if name and label:
                model_points[name] = {
                    "label": label,
                    "description": description if description else "",
                }

        if model_points:
            models_data[f"M{sheet_name}"] = model_points
            print(f"  Loaded {len(model_points)} points from M{sheet_name}")

    wb.close()
    return models_data


def generate_label_from_name(point_name):
    """Generate human-readable label from point name using intelligent parsing."""

    # VSN300-specific SunSpec point patterns (handle BEFORE abbrev_map)
    # Phase voltages (line-to-line)
    if point_name in ["PhVphAB", "PhVphBC", "PhVphCA"]:
        return f"Phase Voltage {point_name[-2:]}"

    # Per-phase frequencies
    if point_name in ["HzA", "HzB", "HzC"]:
        return f"Frequency Phase {point_name[-1]}"

    # Peak power
    if point_name == "PowerPeakAbs":
        return "Peak Power (Absolute)"
    if point_name == "PowerPeakToday":
        return "Peak Power Today"

    # M64061 state codes
    state_map = {
        "AlarmSt": "Alarm State",
        "GlobalSt": "Global State",
        "InverterSt": "Inverter State",
        "DcSt1": "DC Input 1 State",
        "DcSt2": "DC Input 2 State",
    }
    if point_name in state_map:
        return state_map[point_name]

    # Leakage currents
    if point_name == "ILeakDcAc":
        return "Leakage Current DC-AC"
    if point_name == "ILeakDcDc":
        return "Leakage Current DC-DC"

    # Isolation
    if point_name == "Isolation_Ohm1":
        return "Isolation Resistance"

    # Temperature
    if point_name == "Booster_Tmp":
        return "Boost Converter Temperature"

    # Periodic energy counters (M64061 style)
    energy_map = {
        "DayWH": "Daily Energy",
        "WeekWH": "Weekly Energy",
        "MonthWH": "Monthly Energy",
        "YearWH": "Yearly Energy",
    }
    if point_name in energy_map:
        return energy_map[point_name]

    # Known abbreviations
    abbrev_map = {
        "Soc": "State of Charge",
        "Soh": "State of Health",
        "Pgrid": "Grid Power",
        "Vgrid": "Grid Voltage",
        "Igrid": "Grid Current",
        "Fgrid": "Grid Frequency",
        "Qgrid": "Grid Reactive Power",
        "Pba": "Battery Power",
        "Vba": "Battery Voltage",
        "Iba": "Battery Current",
        "Tba": "Battery Temperature",
        "Iin": "Input Current",
        "Vin": "Input Voltage",
        "Pin": "Input Power",
        "Ein": "Input Energy",
        "Riso": "Isolation Resistance",
        "TempInv": "Inverter Temperature",
        "TempBst": "Boost Temperature",
        "WRtg": "Power Rating",
        "Fcc": "Full Charge Capacity",
        "Chc": "Charge Capacity",
        "Dhc": "Discharge Capacity",
        "ShU": "Shutdown",
    }

    if point_name in abbrev_map:
        return abbrev_map[point_name]

    # Handle periodic energy counters (runtime/7D/30D/1Y style)
    if (
        "_runtime" in point_name
        or "_7D" in point_name
        or "_30D" in point_name
        or "_1Y" in point_name
    ):
        base = point_name.split("_")[0]
        period = point_name.split("_")[1] if "_" in point_name else ""
        period_label = {
            "runtime": "Lifetime",
            "7D": "7 Day",
            "30D": "30 Day",
            "1Y": "Yearly",
        }.get(period, period)
        return f"{base} Energy {period_label}"

    # Handle common patterns
    # Split camelCase and underscores
    words = re.findall(r"[A-Z]?[a-z]+|[A-Z]+(?=[A-Z][a-z]|\d|\W|$)|\d+", point_name)
    words = [w for w in words if w]  # Remove empty

    # Join and capitalize
    if not words:
        words = point_name.split("_")

    label = " ".join(word.capitalize() for word in words)

    # Common replacements
    label = label.replace("Pgrid", "Power Grid")
    label = label.replace("Igrid", "Current Grid")
    label = label.replace("Vgrid", "Voltage Grid")
    label = label.replace("Rpm", "RPM")
    label = label.replace("Mppt", "MPPT")
    label = label.replace("Dc", "DC")
    return label.replace("Ac", "AC")


def generate_entity_id_from_label(label, model=None):
    """Generate Home Assistant entity ID from human-readable label.

    Converts labels like "Phase Voltage BC" to "abb_m103_phase_voltage_bc"
    For points without model: "abb_vsn_{label_id}"
    """
    # Convert to lowercase and replace spaces/special chars with underscores
    entity_id = label.lower()

    # Replace common separators and special characters
    entity_id = entity_id.replace(" - ", "_")
    entity_id = entity_id.replace("-", "_")
    entity_id = entity_id.replace(" ", "_")
    entity_id = entity_id.replace("(", "")
    entity_id = entity_id.replace(")", "")
    entity_id = entity_id.replace("/", "_")
    entity_id = entity_id.replace(".", "")
    entity_id = entity_id.replace(",", "")
    entity_id = entity_id.replace(":", "")

    # Remove duplicate underscores
    while "__" in entity_id:
        entity_id = entity_id.replace("__", "_")

    # Strip leading/trailing underscores
    entity_id = entity_id.strip("_")

    # Add prefix with model if available, otherwise use vsn prefix
    if model:
        return f"abb_{model.lower()}_{entity_id}"
    return f"abb_vsn_{entity_id}"


def generate_description_from_name(point_name, category=None):
    """Generate description from point name and category."""

    # Known descriptions
    desc_map = {
        "AlarmState": "Current alarm status of the inverter",
        "GlobState": "Global operating state",
        "InvState": "Inverter operating state",
        "DC1State": "DC input 1 status",
        "DC2State": "DC input 2 status",
        "VBulk": "DC bulk capacitor voltage",
        "VBulkMid": "DC bulk mid-point voltage",
        "Vgnd": "Ground voltage",
        "IleakInv": "Inverter leakage current",
        "IleakDC": "DC side leakage current",
        "Riso": "Isolation resistance",
        "TempBst": "Boost converter temperature",
        "CountryStd": "Grid standard/country code",
        "SysTime": "System timestamp",
        "Ppeak": "Peak power",
        "NumOfMPPT": "Number of MPPT channels",
        "BatteryMode": "Battery operating mode",
        "EBackup": "Energy provided in backup mode",
        "ECharge": "Total battery charging energy",
        "EDischarge": "Total battery discharging energy",
        "Fan1rpm": "Cooling fan 1 rotation speed",
        "Fan2rpm": "Cooling fan 2 rotation speed",
        "DynamicFeedInCtrl": "Dynamic feed-in control status",
        "EnergyPolicy": "Energy management policy setting",
    }

    if point_name in desc_map:
        return desc_map[point_name]

    # Generate based on patterns
    if point_name.startswith("House"):
        if "Pgrid" in point_name:
            return "Household power consumption from grid"
        if "Igrid" in point_name:
            return "Household current consumption from grid"
        if "PInverter" in point_name:
            return "Household power consumption from inverter"

    if point_name.startswith("Meter"):
        if "Pgrid" in point_name:
            return "Grid power measured by meter"
        if "Vgrid" in point_name:
            return "Grid voltage measured by meter"
        if "Igrid" in point_name:
            return "Grid current measured by meter"
        if "Qgrid" in point_name:
            return "Grid reactive power measured by meter"

    if point_name.startswith("E") and category == "Energy Counter":
        return f"{generate_label_from_name(point_name)} counter"

    if "Ctrl" in point_name:
        return f"{generate_label_from_name(point_name)} control setting"

    if "Flags" in point_name:
        return f"{generate_label_from_name(point_name)} status flags"

    if "_mode" in point_name.lower():
        return f"{generate_label_from_name(point_name)} operating mode"

    # Default generic description
    return f"{generate_label_from_name(point_name)}"


def lookup_label_description(models_data, model, sunspec_point):
    """Lookup label and description from models metadata."""
    if not sunspec_point:
        return None, None

    # Try exact match first
    if model in models_data and sunspec_point in models_data[model]:
        return (
            models_data[model][sunspec_point]["label"],
            models_data[model][sunspec_point]["description"],
        )

    # For repeating groups (M160, M203, etc.), strip _N suffix and try base name
    # Example: DCA_1 -> DCA, WphA -> WphA (no change)
    if "_" in sunspec_point and sunspec_point.split("_")[-1].isdigit():
        base_point = "_".join(sunspec_point.split("_")[:-1])
        if model in models_data and base_point in models_data[model]:
            # Append instance number to label
            instance = sunspec_point.split("_")[-1]
            base_label = models_data[model][base_point]["label"]
            base_desc = models_data[model][base_point]["description"]
            return (f"{base_label} #{instance}", base_desc)

    return None, None


def get_description_with_priority(
    point_name, feeds_titles, label, workbook_description, model=None
):
    """Get description using 4-tier priority system with data source tracking.

    Priority:
    1. Description from SunSpec models workbook (most authoritative)
       - Exception: M1 Common Model prefers label over generic descriptions
    2. Title from feeds.json (when it's a description, not a point name)
    3. Enhanced description from generate_description_from_name()
    4. Label from SunSpec models workbook / generated label

    Args:
        point_name: The point name to look up
        feeds_titles: Dict of title data from feeds
        label: The label (from workbook or generated)
        workbook_description: Description from workbook (may be None)
        model: The SunSpec model (e.g., "M103") for data source tracking

    Returns:
        Tuple of (description, data_source)

    """
    # Priority 1: Description from SunSpec workbook (most authoritative)
    if (
        workbook_description
        and str(workbook_description).strip()
        and str(workbook_description).lower() != "none"
    ):
        # Special case: M1 Common Model has generic boilerplate descriptions
        # Prefer user-friendly labels instead
        if model == "M1":
            desc_lower = str(workbook_description).lower()
            if (
                "manufacturer specific value" in desc_lower
                or "well known value registered with sunspec" in desc_lower
                or "modbus device address" in desc_lower
            ):
                # Use label instead for M1 generic descriptions
                if label:
                    return label, f"SunSpec Label ({model}) - generic description"

        # Use description for all other cases
        source = f"SunSpec Description ({model})" if model else "SunSpec Description"
        return workbook_description, source

    # Priority 2: Title from feeds (if it's a description, not a point name)
    if point_name in feeds_titles:
        title_data = feeds_titles[point_name]
        if title_data["is_description"]:
            feed_source = title_data["source"]
            return title_data["title"], f"{feed_source} Feeds"

    # Priority 3: Try enhanced description generation
    enhanced_desc = generate_description_from_name(point_name)
    if enhanced_desc and enhanced_desc != label:
        return enhanced_desc, "Enhanced from point name"

    # Priority 4: Use label as fallback
    if label:
        source = f"SunSpec Label ({model})" if model else "Generated from point name"
        return label, source

    return point_name, "Point name (no label found)"


def create_row_data(
    vsn700_name,
    vsn300_name,
    sunspec_name,
    ha_name,
    in_livedata,
    in_feeds,
    label,
    description,
    ha_display_name,
    model,
    category,
    units,
    state_class,
    device_class,
    entity_category,
    available_in_modbus,
    data_source,
):
    """Create a complete row with all 20 columns.

    This helper ensures all rows have the same structure with HA-specific columns.
    """
    return [
        vsn700_name,  # 1. REST Name (VSN700)
        vsn300_name,  # 2. REST Name (VSN300)
        sunspec_name,  # 3. SunSpec Normalized Name
        ha_name,  # 4. HA Name
        in_livedata,  # 5. In /livedata
        in_feeds,  # 6. In /feeds
        label,  # 7. Label
        description,  # 8. Description
        ha_display_name,  # 9. HA Display Name (what users see)
        model,  # 10. SunSpec Model
        category,  # 11. Category
        units,  # 12. Units
        units,  # 13. HA Unit of Measurement (same as Units for now)
        state_class,  # 14. State Class
        state_class,  # 15. HA State Class (same as State Class)
        device_class,  # 16. Device Class
        device_class,  # 17. HA Device Class (same as Device Class)
        entity_category,  # 18. Entity Category
        available_in_modbus,  # 19. Available in Modbus
        data_source,  # 20. Data Source
    ]


print("Loading SunSpec models metadata...")
sunspec_metadata = load_sunspec_models_metadata(
    "docs/sunspec-maps/models_workbook.xlsx"
)

print()
print("Loading feeds title data...")
feeds_titles = load_feeds_titles(
    "docs/vsn-data/vsn300-data/feeds.json", "docs/vsn-data/vsn700-data/feeds.json"
)

print()
print("Loading VSN data...")

# Load VSN700 data
with open("docs/vsn-data/vsn700-data/livedata.json") as f:
    vsn700_livedata = json.load(f)

with open("docs/vsn-data/vsn700-data/feeds.json") as f:
    vsn700_feeds = json.load(f)

# Load VSN300 data
with open("docs/vsn-data/vsn300-data/livedata.json") as f:
    vsn300_livedata = json.load(f)

# Extract VSN700 point presence
vsn700_livedata_points = set()
for device_data in vsn700_livedata.values():
    if "points" in device_data:
        for point in device_data["points"]:
            vsn700_livedata_points.add(point["name"])

vsn700_feeds_points = set()
for feed_data in vsn700_feeds["feeds"].values():
    if "datastreams" in feed_data:
        for point_name in feed_data["datastreams"]:
            vsn700_feeds_points.add(point_name)

# Extract VSN300 point presence
vsn300_livedata_points = set()
for device_data in vsn300_livedata.values():
    if "points" in device_data:
        for point in device_data["points"]:
            vsn300_livedata_points.add(point["name"])

# Combine for overall presence check
livedata_points = vsn700_livedata_points | vsn300_livedata_points
feeds_points = vsn700_feeds_points

print(f"Found {len(vsn700_livedata_points)} VSN700 points in /livedata")
print(f"Found {len(vsn700_feeds_points)} VSN700 points in /feeds")
print(f"Found {len(vsn300_livedata_points)} VSN300 points in /livedata")

# Get periodic energy counters
periodic_suffixes = ["_runtime", "_7D", "_30D", "_1Y"]
periodic_points = {
    p for p in livedata_points if any(suffix in p for suffix in periodic_suffixes)
}

print(f"Found {len(periodic_points)} periodic energy counters")

# Create workbook
print("Creating Excel workbook...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Points"

# Define headers
headers = [
    "REST Name (VSN700)",
    "REST Name (VSN300)",
    "SunSpec Normalized Name",
    "HA Name",
    "In /livedata",
    "In /feeds",
    "Label",
    "Description",
    "HA Display Name",
    "SunSpec Model",
    "Category",
    "Units",
    "HA Unit of Measurement",
    "State Class",
    "HA State Class",
    "Device Class",
    "HA Device Class",
    "Entity Category",
    "Available in Modbus",
    "Data Source",
]

# Write headers with formatting
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Prepare all data rows
rows = []

# Standard SunSpec points
for vsn_name, mapping in VSN_TO_SUNSPEC_MAP.items():
    if vsn_name in livedata_points or vsn_name in feeds_points:
        sunspec_name = mapping["sunspec"] if mapping["sunspec"] else vsn_name
        model = mapping["model"]

        # Generate HA entity name using new convention: abb_{model}_{point}
        if sunspec_name:
            ha_name = f"abb_{model.lower()}_{sunspec_name.lower()}"
        else:
            ha_name = f"abb_{model.lower()}_{vsn_name.lower()}"

        # Lookup label and description from SunSpec metadata
        label, workbook_description = lookup_label_description(
            sunspec_metadata, model, sunspec_name
        )

        # Fallback to generated label if not found
        if not label:
            label = generate_label_from_name(vsn_name)

        # Get description using 4-tier priority system with data source tracking
        description, data_source = get_description_with_priority(
            vsn_name, feeds_titles, label, workbook_description, model
        )

        # HA Display Name is the description (what users will see in HA)
        ha_display_name = description

        # Find VSN300 name if it exists
        vsn300_name = "N/A"
        if mapping["modbus"]:
            # Check if this modbus-formatted point exists in VSN300 data
            if mapping["modbus"] in vsn300_livedata_points:
                vsn300_name = mapping["modbus"]

        rows.append(
            create_row_data(
                vsn700_name=vsn_name,
                vsn300_name=vsn300_name,
                sunspec_name=sunspec_name if sunspec_name else "N/A",
                ha_name=ha_name,
                in_livedata="✓"
                if vsn_name in vsn700_livedata_points
                or vsn_name in vsn300_livedata_points
                else "",
                in_feeds="✓" if vsn_name in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=ha_display_name,
                model=model,
                category=mapping["category"],
                units=mapping["units"],
                state_class=mapping["state_class"] if mapping["state_class"] else "",
                device_class=mapping["device_class"] if mapping["device_class"] else "",
                entity_category="",  # empty for normal sensor points
                available_in_modbus=mapping["in_modbus"],
                data_source=data_source,
            )
        )

# M64061 periodic energy counters
for p in sorted(periodic_points):
    if p.startswith("E"):
        # Check VSN300 presence
        vsn300_name = (
            f"m64061_1_{p}"
            if ("_runtime" not in p and f"m64061_1_{p}" in vsn300_livedata_points)
            else "N/A"
        )

        # Generate label
        label = generate_label_from_name(p)

        # Generate HA entity name from human-readable label
        ha_name = generate_entity_id_from_label(label, "M64061")

        # Get description using 4-tier priority with data source tracking
        workbook_desc = None  # M64061 periodic counters not in workbook
        description, data_source = get_description_with_priority(
            p, feeds_titles, label, workbook_desc, "M64061"
        )

        # HA Display Name is the description
        ha_display_name = description

        rows.append(
            create_row_data(
                vsn700_name=p,
                vsn300_name=vsn300_name,
                sunspec_name=p,  # SunSpec normalized name is same as VSN700 name
                ha_name=ha_name,
                in_livedata="✓"
                if p in vsn700_livedata_points or p in vsn300_livedata_points
                else "",
                in_feeds="✓" if p in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=ha_display_name,
                model="M64061",
                category="Energy Counter",
                units="Wh",
                state_class="total",
                device_class="energy",
                entity_category="",  # empty for normal sensor points
                available_in_modbus="MAYBE",
                data_source=data_source,
            )
        )

# M64061 non-periodic
for p in sorted(M64061_POINTS):
    if p in livedata_points or p in feeds_points:
        category = "Diagnostic"
        if "State" in p:
            category = "Status"
        elif "leak" in p or p in ["Riso", "VBulk", "VBulkMid", "Vgnd"]:
            category = "Isolation"

        # Check VSN300 presence (case-insensitive search)
        vsn300_name = "N/A"
        expected_vsn300 = f"m64061_1_{p}"
        # Try exact match first
        if expected_vsn300 in vsn300_livedata_points:
            vsn300_name = expected_vsn300
        else:
            # Try case-insensitive match
            for vsn300_point in vsn300_livedata_points:
                if vsn300_point.lower() == expected_vsn300.lower():
                    vsn300_name = vsn300_point
                    break

        # Generate label
        label = generate_label_from_name(p)

        # Generate HA entity name from human-readable label
        ha_name = generate_entity_id_from_label(label, "M64061")

        # Get description using 4-tier priority with data source tracking
        workbook_desc = None  # M64061 points not in standard workbook
        description, data_source = get_description_with_priority(
            p, feeds_titles, label, workbook_desc, "M64061"
        )

        # HA Display Name is the description
        ha_display_name = description

        # Determine units for M64061 points
        units = ""
        device_class = ""
        state_class = "measurement"

        if "V" in p:
            units = "V"
            device_class = "voltage"
        elif "leak" in p:
            units = "A"
            device_class = "current"
        elif p == "Riso":
            units = "Ω"
        elif p == "TempBst" or "Temp" in p:
            units = "°C"
            device_class = "temperature"
        elif p == "Ppeak":
            units = "W"
            device_class = "power"

        rows.append(
            create_row_data(
                vsn700_name=p,
                vsn300_name=vsn300_name,
                sunspec_name=p,  # SunSpec normalized name is same as VSN700 name
                ha_name=ha_name,
                in_livedata="✓"
                if p in vsn700_livedata_points or p in vsn300_livedata_points
                else "",
                in_feeds="✓" if p in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=ha_display_name,
                model="M64061",
                category=category,
                units=units,
                state_class=state_class,
                device_class=device_class,
                entity_category="",  # empty for normal sensor points
                available_in_modbus="MAYBE",
                data_source=data_source,
            )
        )

# ABB Proprietary
for p in sorted(ABB_PROPRIETARY):
    if p in livedata_points or p in feeds_points:
        category = "Other"
        if p.startswith("House"):
            category = "House Meter"
        elif p.startswith("Batt") or p in ["Chc", "Dhc", "ShU", "EnergyPolicy"]:
            category = "Battery Ext"
        elif p.startswith("E"):
            category = "Energy Counter"
        elif "Ctrl" in p or "mode" in p or "Flags" in p or "Mod" in p:
            category = "Control"
        elif p in ["Fan1rpm", "Fan2rpm", "NumOfMPPT", "InputMode"]:
            category = "System Info"
        elif p in ["PacTogrid", "PacStandAlone", "EBackup"]:
            category = "Hybrid"

        # VSN300 doesn't have proprietary points, they're VSN700-only
        vsn300_name = "N/A"

        # Generate HA entity name using abb_vsn_ prefix for proprietary
        ha_name = f"abb_vsn_{p.lower()}"

        # Generate label
        # Special handling for model enable flags (m126Mod_Ena, m132Mod_Ena)
        if p.startswith("m") and "Mod_Ena" in p:
            model_num = p.split("Mod_Ena")[0][1:]  # Extract model number
            label = f"Model {model_num} Enable"
        else:
            label = generate_label_from_name(p)

        # Get description using 4-tier priority with data source tracking
        workbook_desc = None  # Proprietary points not in workbook
        description, data_source = get_description_with_priority(
            p, feeds_titles, label, workbook_desc, "ABB Proprietary"
        )

        # If description is poorly formatted, improve it
        if p.startswith("m") and "Mod_Ena" in p:
            if description == label or "M " in description:
                model_num = p.split("Mod_Ena")[0][1:]
                description = f"SunSpec Model {model_num} Enable Flag"
                data_source = "ABB Documentation"

        # HA Display Name is the description
        ha_display_name = description

        # Determine units for ABB Proprietary points
        # Priority 1: Get from feeds if available
        units = ""
        device_class = ""
        state_class = "measurement"

        if p in feeds_titles and feeds_titles[p].get("units"):
            feeds_units = feeds_titles[p]["units"]
            # Use same conversion logic as VSN300-only points
            if feeds_units == "W":
                units = "W"
                device_class = "power"
            elif feeds_units == "kW":
                units = "kW"
                device_class = "power"
            elif feeds_units == "Wh":
                units = "Wh"
                device_class = "energy"
                state_class = "total_increasing"
            elif feeds_units == "kWh":
                units = "kWh"
                device_class = "energy"
                state_class = "total_increasing"
            elif feeds_units in ["A", "uA"]:
                units = feeds_units
                device_class = "current"
            elif feeds_units == "V":
                units = "V"
                device_class = "voltage"
            elif feeds_units in ["var", "kvar"]:
                units = feeds_units
                device_class = "reactive_power"
            elif feeds_units == "VAh":
                units = "VAh"
                device_class = "apparent_energy"
                state_class = "total_increasing"
            elif feeds_units in ["degC", "°C"]:
                units = "°C"
                device_class = "temperature"
            elif feeds_units in {"none", ""}:
                units = ""
            else:
                units = feeds_units

        # Priority 2: Pattern matching fallback if no feeds data
        if not units:
            if p in ["Chc", "Dhc"]:
                units = "Ah"
            elif p.startswith("E") and p not in ["EnergyPolicy"]:
                units = "Wh"
                device_class = "energy"
                state_class = "total_increasing"
            elif "P" in p and ("grid" in p.lower() or "Inverter" in p or "StandAlone" in p):
                units = "W"
                device_class = "power"
            elif "Igrid" in p:
                units = "A"
                device_class = "current"

        rows.append(
            create_row_data(
                vsn700_name=p,
                vsn300_name=vsn300_name,
                sunspec_name=p,  # No SunSpec mapping, use original name
                ha_name=ha_name,
                in_livedata="✓" if p in vsn700_livedata_points else "",
                in_feeds="✓" if p in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=ha_display_name,
                model="ABB Proprietary",
                category=category,
                units=units,
                state_class=state_class,
                device_class=device_class,
                entity_category="",  # empty for normal sensor points
                available_in_modbus="NO",
                data_source=data_source,
            )
        )

# Add VSN300-only points (points that exist in VSN300 but not mapped yet)
print("Processing VSN300-only points...")
all_vsn300 = vsn300_livedata_points | set(feeds_titles.keys())

# Determine which VSN300 points are already mapped
mapped_vsn300 = set()
for row in rows:
    vsn300_name = row[1]  # Column 2 = REST Name (VSN300)
    if vsn300_name and vsn300_name != "N/A":
        mapped_vsn300.add(vsn300_name)

missing_vsn300 = all_vsn300 - mapped_vsn300

for vsn300_point in sorted(missing_vsn300):
    # Parse SunSpec model and point from VSN300 name
    model = None
    sunspec_point = None
    vsn700_equivalent = None

    # Extract model number from modbus format (e.g., m103_1_AphB -> M103, AphB)
    match = re.match(r"m(\d+)_\d+_(.+)", vsn300_point)
    if match:
        model_num = match.group(1)
        sunspec_point = match.group(2)
        model = f"M{model_num}"

        # Check if feeds title gives us VSN700 equivalent (point name reference)
        if vsn300_point in feeds_titles:
            title_data = feeds_titles[vsn300_point]
            if not title_data["is_description"]:
                # Title is a VSN700 point name reference
                vsn700_equivalent = title_data["title"]

    # Early detection of M1 Common Model points (C_ prefix)
    # Must happen BEFORE entity ID generation so model is set correctly
    if vsn300_point.startswith("C_"):
        model = "M1"
        # Strip C_ prefix for SunSpec lookup (e.g., C_Mn -> Mn)
        sunspec_point = vsn300_point[2:]  # Remove "C_" prefix

    # Get label from workbook or generate
    label, workbook_description = lookup_label_description(
        sunspec_metadata, model, sunspec_point
    )
    if not label:
        # Generate human-readable label using enhanced function
        if model and sunspec_point:
            label = generate_label_from_name(sunspec_point)
        else:
            label = generate_label_from_name(vsn300_point)

    # Generate HA entity name from human-readable label
    ha_name = generate_entity_id_from_label(label, model)

    # Get description with priority and data source tracking
    description, data_source = get_description_with_priority(
        vsn300_point, feeds_titles, label, workbook_description, model
    )

    # HA Display Name is the description
    ha_display_name = description

    # Determine category
    category = "Unknown"
    if model in ["M103", "M101"]:
        category = "Inverter"
    elif model == "M160":
        category = "MPPT"
    elif model == "M64061":
        category = "Diagnostic"
    elif model in ["M802", "M803", "M804"]:
        category = "Battery"
    elif model in ["M201", "M203", "M204"]:
        category = "Meter"

    # Units and device class
    # Priority 1: Get from feeds JSON (most reliable source)
    units = ""
    device_class = ""
    state_class = "measurement"

    if vsn300_point in feeds_titles and feeds_titles[vsn300_point].get("units"):
        feeds_units = feeds_titles[vsn300_point]["units"]
        # Map VSN feeds units to HA units (keep original scale - HA supports kW, kWh, etc.)
        if feeds_units == "kW":
            units = "kW"
            device_class = "power"
            state_class = "measurement"
        elif feeds_units == "kWh":
            units = "kWh"
            device_class = "energy"
            state_class = "total_increasing"
        elif feeds_units == "kvar":
            units = "kvar"
            device_class = "reactive_power"
            state_class = "measurement"
        elif feeds_units in ["A", "uA"]:
            units = feeds_units
            device_class = "current"
        elif feeds_units == "V":
            units = "V"
            device_class = "voltage"
        elif feeds_units == "Hz":
            units = "Hz"
            device_class = "frequency"
        elif feeds_units in ["degC", "°C"]:
            units = "°C"
            device_class = "temperature"
        elif feeds_units == "%":
            units = "%"
            # Device class depends on context - leave empty for now
        elif feeds_units == "s":
            units = "s"
            device_class = "duration"
        elif feeds_units == "bytes":
            units = "B"
            # No standard device class for bytes
        elif feeds_units in ["MOhm", "Ohm"]:
            units = feeds_units
            # No standard device class for resistance
        elif feeds_units == "Ah":
            units = "Ah"
            # No standard device class for Ah
        elif feeds_units == "W":
            units = "W"
            device_class = "power"
        elif feeds_units == "Wh":
            units = "Wh"
            device_class = "energy"
            state_class = "total_increasing"
        elif feeds_units == "var":
            units = "var"
            device_class = "reactive_power"
        elif feeds_units in {"none", ""}:
            # Feeds says no units
            units = ""
        else:
            # Unknown unit from feeds, keep it as-is
            units = feeds_units

    # Priority 2: Common abbreviations and known patterns
    if not units:
        point_check = sunspec_point if sunspec_point else vsn300_point
        point_lower = point_check.lower() if point_check else ""

        # Temperature (must check before Tmp to avoid false positives)
        if ("temp" in point_lower or "tmp" in point_lower) and "bst" in point_lower:
            units = "°C"
            device_class = "temperature"
        # Isolation resistance
        elif "riso" in point_lower or "isolation" in point_lower:
            units = "Ω"  # Ohm symbol
            # No standard device class for resistance
        # Capacities (Ah)
        elif point_check in ["Chc", "Dhc"]:  # Charge/Discharge capacity
            units = "Ah"
            # No standard device class for Ah
        # Peak power
        elif "ppeak" in point_lower or "powerpeak" in point_lower:
            units = "W"
            device_class = "power"
        # Energy points (E prefix typically means energy)
        elif point_check and point_check.startswith("E") and len(point_check) <= 10 and "_" not in point_check:
            # EBackup, ECharge, EDischarge, etc.
            units = "Wh"
            device_class = "energy"
            state_class = "total_increasing"

    # Priority 3: Pattern matching on SunSpec point name (original logic)
    if not units and sunspec_point:
        if sunspec_point.startswith("Aph"):
            # Phase current points (AphA, AphB, AphC)
            units = "A"
            device_class = "current"
        elif "A" in sunspec_point and "phA" not in sunspec_point:
            units = "A"
            device_class = "current"
        elif "V" in sunspec_point:
            units = "V"
            device_class = "voltage"
        elif "PowerPeak" in sunspec_point:
            # Peak power points (PowerPeakToday, PowerPeakAbs)
            units = "W"
            device_class = "power"
        elif "W" in sunspec_point and "WH" not in sunspec_point:
            units = "W"
            device_class = "power"
        elif "WH" in sunspec_point:
            units = "Wh"
            device_class = "energy"
            state_class = "total_increasing"
        elif "Hz" in sunspec_point:
            units = "Hz"
            device_class = "frequency"
        elif "Tmp" in sunspec_point or "Temp" in vsn300_point:
            units = "°C"
            device_class = "temperature"

    # Determine entity_category
    # M1 Common Model points and VSN system monitoring -> diagnostic
    entity_category = ""
    if vsn300_point.startswith("C_") or vsn300_point in [
        "flash_free",
        "free_ram",
        "fw_ver",
        "store_size",
        "sys_load",
        "uptime",
    ]:
        entity_category = "diagnostic"
        # Override category for these diagnostic points
        if vsn300_point.startswith("C_"):
            category = "Device Info"
            # Model M1 already set earlier at line 1420-1422
        else:
            category = "System Monitoring"

    rows.append(
        create_row_data(
            vsn700_name=vsn700_equivalent if vsn700_equivalent else "N/A",
            vsn300_name=vsn300_point,
            sunspec_name=sunspec_point if sunspec_point else vsn300_point,
            ha_name=ha_name,
            in_livedata="✓" if vsn300_point in vsn300_livedata_points else "",
            in_feeds="✓" if vsn300_point in feeds_titles else "",
            label=label,
            description=description,
            ha_display_name=ha_display_name,
            model=model if model else "VSN300-only",
            category=category,
            units=units,
            state_class=state_class,
            device_class=device_class,
            entity_category=entity_category,  # diagnostic for M1/system points, empty otherwise
            available_in_modbus="YES"
            if model in ["M103", "M101", "M160", "M120", "M124", "M802", "M203"]
            or model == "M1"
            else "NO",
            data_source=data_source,
        )
    )

print(f"Added {len(missing_vsn300)} VSN300-only points")

# Add VSN700-only points (points in feeds but not in livedata)
print("Processing VSN700-only points...")
vsn700_feeds_only = set()
for point_name, title_data in feeds_titles.items():
    if (
        title_data["source"] in ["VSN700", "Both"]
        and point_name not in vsn700_livedata_points
    ):
        vsn700_feeds_only.add(point_name)

# Check which are already mapped
mapped_vsn700 = set()
for row in rows:
    vsn700_name = row[0]  # Column 1 = REST Name (VSN700)
    if vsn700_name and vsn700_name != "N/A":
        mapped_vsn700.add(vsn700_name)

missing_vsn700 = vsn700_feeds_only - mapped_vsn700

for vsn700_point in sorted(missing_vsn700):
    # These are VSN700-specific points, mostly totals
    # Get label
    label = generate_label_from_name(vsn700_point)

    # Generate HA entity name from human-readable label
    ha_name = generate_entity_id_from_label(label, None)

    # Get description with data source tracking (feeds title has priority)
    description, data_source = get_description_with_priority(
        vsn700_point, feeds_titles, label, None, "VSN700-only"
    )

    # HA Display Name is the description
    ha_display_name = description

    # Determine category and entity_category
    category = "Energy Counter" if vsn700_point.startswith("E") else "Other"
    entity_category = ""

    # VSN700 system monitoring points -> diagnostic
    if vsn700_point in [
        "flash_free",
        "free_ram",
        "fw_ver",
        "store_size",
        "sys_load",
        "uptime",
    ]:
        entity_category = "diagnostic"
        category = "System Monitoring"

    rows.append(
        create_row_data(
            vsn700_name=vsn700_point,
            vsn300_name="N/A",  # doesn't exist
            sunspec_name=vsn700_point,  # SunSpec normalized
            ha_name=ha_name,
            in_livedata="",  # Not in livedata
            in_feeds="✓",  # In feeds only
            label=label,
            description=description,
            ha_display_name=ha_display_name,
            model="VSN700-only",
            category=category,
            units="Wh" if vsn700_point.startswith("E") else "",
            state_class="total_increasing" if vsn700_point.startswith("E") else "measurement",
            device_class="energy" if vsn700_point.startswith("E") else "",
            entity_category=entity_category,  # diagnostic for system monitoring, empty otherwise
            available_in_modbus="NO",
            data_source=data_source,
        )
    )

print(f"Added {len(missing_vsn700)} VSN700-only points")

# Write all rows
for row_num, row_data in enumerate(rows, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border

        # Color code by model (column index 9 = SunSpec Model, 0-indexed)
        if row_data[9] == "M64061":
            cell.fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )
        elif row_data[9] == "ABB Proprietary":
            cell.fill = PatternFill(
                start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
            )

# Auto-size columns
for col_num in range(1, len(headers) + 1):
    column_letter = get_column_letter(col_num)
    max_length = len(headers[col_num - 1])
    for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
        if row[0].value:
            max_length = max(max_length, len(str(row[0].value)))
    ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

# Freeze header row
ws.freeze_panes = "A2"

# Add filter
ws.auto_filter.ref = ws.dimensions

# Add summary statistics
summary_ws = wb.create_sheet("Summary")
summary_ws.append(["Category", "Count"])
summary_ws.append(["Total Points", len(rows)])
summary_ws.append(
    ["Standard SunSpec (Both protocols)", len([r for r in rows if r[18] == "YES"])]
)  # Column 19 (0-indexed: 18) = Available in Modbus
summary_ws.append(
    ["M64061 (Maybe in Modbus)", len([r for r in rows if r[9] == "M64061"])]
)  # Column 10 (0-indexed: 9) = SunSpec Model
summary_ws.append(
    ["ABB Proprietary (REST only)", len([r for r in rows if r[9] == "ABB Proprietary"])]
)  # Column 10 (0-indexed: 9) = SunSpec Model
summary_ws.append(
    ["Diagnostic entities", len([r for r in rows if r[17] == "diagnostic"])]
)  # Column 18 (0-indexed: 17) = Entity Category
summary_ws.append(
    ["In /livedata", len([r for r in rows if r[4] == "✓"])]
)  # Column 5 (0-indexed: 4) = In /livedata
summary_ws.append(
    ["In /feeds", len([r for r in rows if r[5] == "✓"])]
)  # Column 6 (0-indexed: 5) = In /feeds
summary_ws.append(
    [
        "Points with labels from SunSpec",
        len([r for r in rows if r[6] and r[6] != "N/A"]),
    ]
)  # Column 7 (0-indexed: 6) = Label
summary_ws.append(
    [
        "Points with generated labels",
        len([r for r in rows if not r[6] or r[6] == "N/A"]),
    ]
)

# Format summary
for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row):
    row[0].font = Font(bold=True)

# Save workbook
output_path = "docs/vsn-sunspec-point-mapping.xlsx"
wb.save(output_path)
print(f"\n✓ Excel file created: {output_path}")
print(f"  Total rows: {len(rows)}")
print(
    f"  Standard SunSpec: {len([r for r in rows if r[18] == 'YES'])}"
)  # Column 19 (0-indexed: 18) = Available in Modbus
print(
    f"  M64061: {len([r for r in rows if r[9] == 'M64061'])}"
)  # Column 10 (0-indexed: 9) = SunSpec Model
print(f"  ABB Proprietary: {len([r for r in rows if r[9] == 'ABB Proprietary'])}")
print(
    f"  Diagnostic entities: {len([r for r in rows if r[17] == 'diagnostic'])}"
)  # Column 18 (0-indexed: 17) = Entity Category
print(
    f"  Points with SunSpec labels: {len([r for r in rows if r[6] and r[6] != 'N/A'])}"
)  # Column 7 (0-indexed: 6) = Label
print(
    f"  Points with generated labels: {len([r for r in rows if not r[6] or r[6] == 'N/A'])}"
)
