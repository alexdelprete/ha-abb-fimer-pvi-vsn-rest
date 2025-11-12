#!/usr/bin/env python3
"""Extract manual changes from edited Excel file and generate Python code for backporting.

This script compares the original generated Excel file with a manually edited version
and extracts all differences in Label, Description, and HA Display Name fields.
It then generates Python code that can be copied into the DESCRIPTION_IMPROVEMENTS
and DISPLAY_NAME_CORRECTIONS dictionaries in generate_mapping.py.

Usage:
    1. Run generate_mapping.py to create vsn-sunspec-point-mapping.xlsx
    2. Save the original as vsn-sunspec-point-mapping-original.xlsx
    3. Edit vsn-sunspec-point-mapping.xlsx manually
    4. Run this script: python extract_manual_changes.py
    5. Copy the generated code into generate_mapping.py

Author: Claude
Version: 1.0.0
Date: 2024-11-03
"""

from pathlib import Path

import openpyxl

# File paths
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"
ORIGINAL_FILE = OUTPUT_DIR / "vsn-sunspec-point-mapping-original.xlsx"
EDITED_FILE = OUTPUT_DIR / "vsn-sunspec-point-mapping.xlsx"


def load_excel_data(file_path):
    """Load data from Excel file into a dictionary keyed by SunSpec Normalized Name."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb["All Points"]

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    # Find column indices
    col_indices = {header: idx for idx, header in enumerate(headers)}

    # Load data
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sunspec_name = row[col_indices["SunSpec Normalized Name"]]
        if sunspec_name and sunspec_name != "N/A":
            data[sunspec_name] = {
                "Label": row[col_indices["Label"]],
                "Description": row[col_indices["Description"]],
                "HA Display Name": row[col_indices["HA Display Name"]],
                "REST Name (VSN700)": row[col_indices["REST Name (VSN700)"]],
                "REST Name (VSN300)": row[col_indices["REST Name (VSN300)"]],
            }

    wb.close()
    return data


def compare_data(original, edited):
    """Compare original and edited data, return differences."""
    changes = {
        "label": [],
        "description": [],
        "display_name": []
    }

    for sunspec_name in edited:
        if sunspec_name not in original:
            continue  # New point in edited file, skip

        orig = original[sunspec_name]
        edit = edited[sunspec_name]

        # Check Label changes
        if orig["Label"] != edit["Label"]:
            changes["label"].append({
                "sunspec_name": sunspec_name,
                "vsn700": edit["REST Name (VSN700)"],
                "vsn300": edit["REST Name (VSN300)"],
                "old": orig["Label"],
                "new": edit["Label"]
            })

        # Check Description changes
        if orig["Description"] != edit["Description"]:
            changes["description"].append({
                "sunspec_name": sunspec_name,
                "vsn700": edit["REST Name (VSN700)"],
                "vsn300": edit["REST Name (VSN300)"],
                "old": orig["Description"],
                "new": edit["Description"]
            })

        # Check HA Display Name changes
        if orig["HA Display Name"] != edit["HA Display Name"]:
            changes["display_name"].append({
                "sunspec_name": sunspec_name,
                "vsn700": edit["REST Name (VSN700)"],
                "vsn300": edit["REST Name (VSN300)"],
                "old": orig["HA Display Name"],
                "new": edit["HA Display Name"]
            })

    return changes


def generate_python_code(changes):
    """Generate Python code for DESCRIPTION_IMPROVEMENTS and DISPLAY_NAME_CORRECTIONS."""
    output = []

    output.append("=" * 100)
    output.append("MANUAL CHANGES EXTRACTED - COPY THIS TO generate_mapping.py")
    output.append("=" * 100)
    output.append("")

    # Generate DESCRIPTION_IMPROVEMENTS code
    if changes["description"]:
        output.append("# Add to DESCRIPTION_IMPROVEMENTS dictionary:")
        output.append("# ==========================================")
        output.append(f"# Manual Improvements from Excel Edit ({len(changes['description'])} changes)")
        output.append("# ==========================================")
        output.append("")

        for change in changes["description"]:
            output.append(f"# {change['sunspec_name']} (VSN700: {change['vsn700']}, VSN300: {change['vsn300']})")
            output.append(f'"{change["sunspec_name"]}": "{change["new"]}",')
            output.append("")
    else:
        output.append("# No description changes found")
        output.append("")

    # Generate DISPLAY_NAME_CORRECTIONS code
    if changes["display_name"]:
        output.append("")
        output.append("=" * 100)
        output.append("# Add to DISPLAY_NAME_CORRECTIONS dictionary:")
        output.append("# ==========================================")
        output.append(f"# Manual Display Name Corrections from Excel Edit ({len(changes['display_name'])} changes)")
        output.append("# ==========================================")
        output.append("")

        for change in changes["display_name"]:
            # Use the OLD description as the key (since that's how DISPLAY_NAME_CORRECTIONS works)
            output.append(f"# {change['sunspec_name']} (VSN700: {change['vsn700']}, VSN300: {change['vsn300']})")
            output.append(f'"{change["old"]}": "{change["new"]}",')
            output.append("")
    else:
        output.append("# No display name changes found")
        output.append("")

    # Generate LABEL_CORRECTIONS code if needed (labels are usually auto-generated)
    if changes["label"]:
        output.append("")
        output.append("=" * 100)
        output.append("# Add to LABEL_CORRECTIONS dictionary:")
        output.append("# ==========================================")
        output.append(f"# Manual Label Corrections from Excel Edit ({len(changes['label'])} changes)")
        output.append("# ==========================================")
        output.append("")

        for change in changes["label"]:
            output.append(f"# {change['sunspec_name']} (VSN700: {change['vsn700']}, VSN300: {change['vsn300']})")
            output.append(f'"{change["old"]}": "{change["new"]}",')
            output.append("")
    else:
        output.append("# No label changes found")

    output.append("")
    output.append("=" * 100)
    output.append("SUMMARY")
    output.append("=" * 100)
    output.append(f"Label changes: {len(changes['label'])}")
    output.append(f"Description changes: {len(changes['description'])}")
    output.append(f"Display name changes: {len(changes['display_name'])}")
    output.append(f"Total changes: {len(changes['label']) + len(changes['description']) + len(changes['display_name'])}")

    return "\n".join(output)


def main():
    """Main function."""
    print("=" * 100)
    print("EXTRACT MANUAL CHANGES FROM EXCEL")
    print("=" * 100)
    print()

    # Check files exist
    if not ORIGINAL_FILE.exists():
        print(f"❌ ERROR: Original file not found: {ORIGINAL_FILE}")
        print(f"   Please save the original generated file as: {ORIGINAL_FILE.name}")
        return

    if not EDITED_FILE.exists():
        print(f"❌ ERROR: Edited file not found: {EDITED_FILE}")
        print(f"   Please edit the file and save it as: {EDITED_FILE.name}")
        return

    print(f"✓ Found original file: {ORIGINAL_FILE.name}")
    print(f"✓ Found edited file: {EDITED_FILE.name}")
    print()

    # Load data
    print("Loading original data...")
    original_data = load_excel_data(ORIGINAL_FILE)
    print(f"  Loaded {len(original_data)} points")

    print("Loading edited data...")
    edited_data = load_excel_data(EDITED_FILE)
    print(f"  Loaded {len(edited_data)} points")
    print()

    # Compare
    print("Comparing files...")
    changes = compare_data(original_data, edited_data)
    print(f"  Found {len(changes['label'])} label changes")
    print(f"  Found {len(changes['description'])} description changes")
    print(f"  Found {len(changes['display_name'])} display name changes")
    print()

    # Generate code
    if not any(changes.values()):
        print("✓ No changes found! Files are identical.")
        return

    print("Generating Python code...")
    code = generate_python_code(changes)

    # Save to file
    output_file = OUTPUT_DIR / "manual_changes_backport.txt"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(code)

    print(f"✓ Code saved to: {output_file}")
    print()
    print("=" * 100)
    print("NEXT STEPS:")
    print("=" * 100)
    print(f"1. Open {output_file}")
    print("2. Copy the generated code sections")
    print("3. Paste into the appropriate dictionaries in generate_mapping.py:")
    print("   - DESCRIPTION_IMPROVEMENTS (around line 579)")
    print("   - DISPLAY_NAME_CORRECTIONS (around line 139)")
    print("   - LABEL_CORRECTIONS (around line 422)")
    print("4. Run generate_mapping.py to regenerate with your changes")
    print("5. Verify the changes were applied correctly")
    print()


if __name__ == "__main__":
    main()
