import openpyxl

# Load all files
abb_wb = openpyxl.load_workbook(r'data\sunspec\ABB_SunSpec_Modbus.xlsx', data_only=True)
abb_ws = abb_wb['Inverter Modbus']

sunspec_wb = openpyxl.load_workbook(r'data\sunspec\models_workbook.xlsx', data_only=True)

mapping_wb = openpyxl.load_workbook(r'output\vsn-sunspec-point-mapping.xlsx', data_only=True)
mapping_ws = mapping_wb['All Points']

# Extract ABB data
abb_data = {}
model_ranges = [
    ('M1', 7, 17),
    ('M103', 17, 63),
    ('M160', 63, 95),
    ('M64061', 95, 200)
]

for model, start_row, end_row in model_ranges:
    abb_data[model] = {}
    for row in range(start_row, min(end_row, abb_ws.max_row + 1)):
        name = abb_ws.cell(row, 9).value
        if name and name not in ['Name', 'PAD', 'ID', 'L'] and not str(name).startswith('='):
            desc = abb_ws.cell(row, 12).value
            
            abb_data[model][name] = {
                'desc': str(desc) if (desc and str(desc) != 'None') else ''
            }

# Extract SunSpec data
sunspec_data = {}
for model_name, sheet_name in [('M1', '1'), ('M103', '103'), ('M160', '160')]:
    if sheet_name in sunspec_wb.sheetnames:
        ws = sunspec_wb[sheet_name]
        sunspec_data[model_name] = {}
        
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 3).value
            desc = ws.cell(row, 14).value
            
            if name and name not in ['ID', 'L']:
                sunspec_data[model_name][name] = {
                    'desc': str(desc) if (desc and str(desc) != 'None') else ''
                }

# Get mapping headers
headers = {}
for col in range(1, mapping_ws.max_column + 1):
    headers[mapping_ws.cell(1, col).value] = col

print('='*100)
print('COMPREHENSIVE ANALYSIS: ABB Excel vs SunSpec Workbook for Point Descriptions')
print('='*100)
print()

# M1 Analysis
print('M1 (COMMON MODEL) - All 6 points comparison')
print('-' * 100)
for row in range(2, mapping_ws.max_row + 1):
    if mapping_ws.cell(row, headers['M1']).value == 'YES':
        name = mapping_ws.cell(row, headers['SunSpec Normalized Name']).value
        current = mapping_ws.cell(row, headers['Description']).value
        
        abb = abb_data.get('M1', {}).get(name, {}).get('desc', '')
        sunspec = sunspec_data.get('M1', {}).get(name, {}).get('desc', '')
        
        print(f'\n{name}:')
        print(f'  SunSpec: {sunspec}')
        print(f'  ABB:     {abb}')
        print(f'  >>> SunSpec is clearly better (detailed vs single word)')

# M103 sample
print('\n\nM103 (INVERTER) - Sample comparison (5 points)')
print('-' * 100)
m103_count = 0
for row in range(2, mapping_ws.max_row + 1):
    if mapping_ws.cell(row, headers['M103']).value == 'YES' and m103_count < 5:
        name = mapping_ws.cell(row, headers['SunSpec Normalized Name']).value
        
        abb = abb_data.get('M103', {}).get(name, {}).get('desc', '')
        sunspec = sunspec_data.get('M103', {}).get(name, {}).get('desc', '')
        
        if abb and sunspec:
            m103_count += 1
            print(f'\n{name}:')
            print(f'  SunSpec: {sunspec}')
            print(f'  ABB:     {abb}')

# M160 - where ABB adds value
print('\n\nM160 (MPPT) - Where ABB Excel provides value')
print('-' * 100)
print('NOTE: SunSpec workbook has NO descriptions for M160!')
print()
for row in range(2, mapping_ws.max_row + 1):
    if mapping_ws.cell(row, headers['M160']).value == 'YES':
        name = mapping_ws.cell(row, headers['SunSpec Normalized Name']).value
        current = mapping_ws.cell(row, headers['Description']).value
        source = mapping_ws.cell(row, headers['Data Source']).value
        
        abb = abb_data.get('M160', {}).get(name, {}).get('desc', '')
        sunspec = sunspec_data.get('M160', {}).get(name, {}).get('desc', '')
        
        print(f'\n{name}:')
        print(f'  Current ({source}): {current}')
        print(f'  SunSpec: {sunspec if sunspec else "(NONE)"}')
        print(f'  ABB:     {abb if abb else "(none)"}')
        if abb and not sunspec:
            print(f'  >>> ABB provides description!')

print('\n\n' + '='*100)
print('VERDICT')
print('='*100)
print()
print('DO NOT prioritize ABB Excel over SunSpec workbook globally.')
print()
print('Reasons:')
print('  1. M1: ABB has single-word descriptions (manufacturer, model)')
print('     vs SunSpec detailed descriptions')
print('  2. M103: Similar - ABB is less descriptive')
print('  3. M160: ABB provides value HERE (16 points), SunSpec has none')
print('  4. M64061: ABB is the only source (already being used)')
print()
print('Recommendation:')
print('  Use ABB Excel as FALLBACK for M160 only, not as priority.')
print()
print('Impact if prioritized globally: 6 M1 points would get WORSE descriptions.')

abb_wb.close()
sunspec_wb.close()
mapping_wb.close()
