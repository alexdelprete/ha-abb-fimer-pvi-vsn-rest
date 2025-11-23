#!/usr/bin/env python3
"""Generate complete VSN-SunSpec point mapping Excel with all customizations built-in.

This comprehensive script replaces the multi-script workflow with a single source of truth
that generates the complete 26-column Excel with model flags, all fixes applied, and a
changelog tab tracking all updates.

This version is self-contained with all data files in the data/ subdirectory and
processes status.json endpoints along with livedata and feeds.

Author: Claude/Alex
Version: 2.0.0
Date: 2024-11-02
"""

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Script metadata
SCRIPT_VERSION = "2.0.9"
SCRIPT_DATE = datetime.now().strftime("%Y-%m-%d")

# Get script directory for relative paths
SCRIPT_DIR = Path(__file__).parent

# ==============================================================================
# PART 1: CONSTANTS AND MAPPINGS
# ==============================================================================

# Model flag columns
MODEL_FLAGS = [
    "M1",
    "M103",
    "M160",
    "M203",
    "M802",
    "M64061",
    "VSN300_Only",
    "VSN700_Only",
    "ABB_Proprietary",
]

# Full Excel headers (28 columns total - added "Suggested Display Precision" in v1.1.7+)
EXCEL_HEADERS = [
    *MODEL_FLAGS,
    "REST Name (VSN700)",
    "REST Name (VSN300)",
    "SunSpec Normalized Name",
    "HA Name",
    "In /livedata",
    "In /feeds",
    "Label",
    "Description",
    "HA Display Name",
    "Category",
    "HA Unit of Measurement",
    "HA State Class",
    "HA Device Class",
    "HA Icon",
    "Entity Category",
    "Suggested Display Precision",
    "Available in Modbus",
    "Data Source",
    "Model_Notes",
]

# VSN to SunSpec mappings (from original script)
VSN_TO_SUNSPEC_MAP = {
    # Inverter measurements (M103)
    "Pgrid": {
        "sunspec": "W",
        "model": "M103",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": "m103_1_W",
    },
    "Qgrid": {
        "sunspec": "VAr",
        "model": "M103",
        "category": "Inverter",
        "units": "var",
        "state_class": "measurement",
        "device_class": "reactive_power",
        "in_modbus": "YES",
        "modbus": "m103_1_VAr",
    },
    "Sgrid": {
        "sunspec": "VA",
        "model": "M103",
        "category": "Inverter",
        "units": "VA",
        "state_class": "measurement",
        "device_class": "apparent_power",
        "in_modbus": "YES",
        "modbus": "m103_1_VA",
    },
    "Fgrid": {
        "sunspec": "Hz",
        "model": "M103",
        "category": "Inverter",
        "units": "Hz",
        "state_class": "measurement",
        "device_class": "frequency",
        "in_modbus": "YES",
        "modbus": "m103_1_Hz",
    },
    "Igrid": {
        "sunspec": "A",
        "model": "M103",
        "category": "Inverter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
        "modbus": "m103_1_A",
    },
    "Vgrid": {
        "sunspec": "PhVphA",
        "model": "M103",
        "category": "Inverter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
        "modbus": "m103_1_PhVphA",
    },
    "Etotal": {
        "sunspec": "TotWhExp",
        "model": "M103",
        "category": "Energy Counter",
        "units": "kWh",
        "state_class": "total_increasing",
        "device_class": "energy",
        "in_modbus": "YES",
        "modbus": "m103_1_TotWhExp",
    },
    "PF": {
        "sunspec": "PF",
        "model": "M103",
        "category": "Inverter",
        "units": "",
        "state_class": "measurement",
        "device_class": "power_factor",
        "in_modbus": "YES",
        "modbus": "m103_1_PF",
    },
    # Three-phase measurements (M103)
    "Pgrid_L1": {
        "sunspec": "WphA",
        "model": "M103",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": None,
    },
    "Pgrid_L2": {
        "sunspec": "WphB",
        "model": "M103",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": None,
    },
    "Pgrid_L3": {
        "sunspec": "WphC",
        "model": "M103",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": None,
    },
    # DC measurements (M160)
    "Ppv": {
        "sunspec": "DCW",
        "model": "M160",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": "m160_1_DCW",
    },
    "Vpv_1": {
        "sunspec": "DCV_1",
        "model": "M160",
        "category": "Inverter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
        "modbus": "m160_1_DCV_1",
    },
    "Ipv_1": {
        "sunspec": "DCA_1",
        "model": "M160",
        "category": "Inverter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
        "modbus": "m160_1_DCA_1",
    },
    "Ppv_1": {
        "sunspec": "DCW_1",
        "model": "M160",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": "m160_1_DCW_1",
    },
    "Vpv_2": {
        "sunspec": "DCV_2",
        "model": "M160",
        "category": "Inverter",
        "units": "V",
        "state_class": "measurement",
        "device_class": "voltage",
        "in_modbus": "YES",
        "modbus": "m160_1_DCV_2",
    },
    "Ipv_2": {
        "sunspec": "DCA_2",
        "model": "M160",
        "category": "Inverter",
        "units": "A",
        "state_class": "measurement",
        "device_class": "current",
        "in_modbus": "YES",
        "modbus": "m160_1_DCA_2",
    },
    "Ppv_2": {
        "sunspec": "DCW_2",
        "model": "M160",
        "category": "Inverter",
        "units": "W",
        "state_class": "measurement",
        "device_class": "power",
        "in_modbus": "YES",
        "modbus": "m160_1_DCW_2",
    },
    # M1 Common Model
    "C_Mn": {
        "sunspec": "Mn",
        "model": "M1",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
        "modbus": "m1_Mn",
    },
    "C_Md": {
        "sunspec": "Md",
        "model": "M1",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
        "modbus": "m1_Md",
    },
    "C_Opt": {
        "sunspec": "Opt",
        "model": "M1",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
        "modbus": "m1_Opt",
    },
    "C_Vr": {
        "sunspec": "Vr",
        "model": "M1",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
        "modbus": "m1_Vr",
    },
    "C_SN": {
        "sunspec": "SN",
        "model": "M1",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
        "modbus": "m1_SN",
    },
    "C_DA": {
        "sunspec": "DA",
        "model": "M1",
        "category": "Inverter",
        "units": "",
        "state_class": None,
        "device_class": None,
        "in_modbus": "YES",
        "modbus": "m1_DA",
    },
}

# M64061 vendor-specific points
M64061_POINTS = {
    "AlarmState",
    "Alm1",
    "Alm2",
    "Alm3",
    "BatteryMode",
    "BatteryStatus",
    "FaultStatus",
    "IsolResist",
    "RemoteControlRequest",
    "RemoteControlStatus",
    "Mode",
    "CmdTimeout",
    "CtrlTimeout",
    "CtrlOpt",
    "AUX1",
    "AUX2",
    "BatteryVoltage",
    "BatteryTemperature",
    "BatterySoC",
    "BatterySoH",
    "BatteryChargeCurrent",
    "BatteryDischargeCurrent",
    "BatteryPower",
    "BatteryChargePower",
    "E0",
    "E1",
    "E2",
    "E3",
    "E4",
    "E5",
    "E6",
    "E7",
    "E8",
    "ECharge",
    "EDischarge",
    "ETotCharge",
    "ETotDischarge",
}

# VSN name normalization for duplicate detection
# PRIORITY: VSN300 names are SunSpec-based and are canonical
# Maps VSN700 proprietary names → VSN300 SunSpec names for merging
# VSN300 names don't need normalization (they're already canonical)
VSN_NAME_NORMALIZATION = {
    # Insulation Resistance (CRITICAL - user-reported icon issue)
    # VSN300: m64061_1_Isolation_Ohm1 → Isolation_Ohm1 (SunSpec-based, canonical)
    # VSN700: Riso → Isolation_Ohm1 (normalize to SunSpec name)
    # Result: Single "isolation_ohm1" sensor, both REST names, unit=MOhm, icon=mdi:omega
    "Riso": "Isolation_Ohm1",  # VSN700 → VSN300 (SunSpec) name

    # Leakage Current - Inverter (CRITICAL - unit mismatch A vs mA)
    # VSN300: m64061_1_ILeakDcAc → ILeakDcAc (mA, describes DC-to-AC path, canonical)
    # VSN700: IleakInv → ILeakDcAc (A, less specific, normalize to VSN300)
    # Result: Single "ileakdcac" sensor, unit=mA, A→mA conversion in normalizer
    "IleakInv": "ILeakDcAc",  # VSN700 → VSN300 (SunSpec) name

    # Leakage Current - DC (CRITICAL - unit mismatch A vs mA)
    # VSN300: m64061_1_ILeakDcDc → ILeakDcDc (mA, describes DC-to-DC path, canonical)
    # VSN700: IleakDC → ILeakDcDc (A, less specific, normalize to VSN300)
    # Result: Single "ileakdcdc" sensor, unit=mA, A→mA conversion in normalizer
    "IleakDC": "ILeakDcDc",  # VSN700 → VSN300 (SunSpec) name

    # Voltage - Ground Reference (capitalization only)
    # VSN300: VGnd (V, PascalCase - SunSpec convention, canonical)
    # VSN700: Vgnd (V, lowercase g, normalize to VSN300)
    # Result: Single "vgnd" sensor with both REST names
    "Vgnd": "VGnd",  # VSN700 → VSN300 (SunSpec) name

    # Battery State of Charge (naming convention)
    # VSN300: Soc (%, standard SunSpec name, canonical)
    # VSN700: TSoc (%, proprietary prefix, normalize to VSN300)
    # Result: Single "soc" sensor with both REST names
    "TSoc": "Soc",  # VSN700 → VSN300 (SunSpec) name
}

# ABB Proprietary points (REST API only)
ABB_PROPRIETARY = {
    # Battery measurements
    "Vbat",
    "Ibat",
    "Pbat",
    "Soc",
    "Soh",
    "Wbat_charge",
    "Wbat_discharge",
    "Tmp",
    "BattNum",
    "Chc",
    "Dhc",
    # System configuration
    "UploadAddr",
    "UploadPort",
    "UploadAddr2",
    "UploadPort2",
    "ApplVer",
    "SerNum",
    # Time/Schedule
    "HWVer",
    "Year",
    "Month",
    "Day",
    "Weekday",
    "Hour",
    "Minute",
    "Second",
    # Grid parameters
    "SplitPhase",
    "CountryStd",
    # House meter measurements
    "HousePgrid",
    "HousePgrid_L1",
    "HousePgrid_L2",
    "HousePgrid_L3",
    "HouseIgrid",
    "HouseIgrid_L1",
    "HouseIgrid_L2",
    "HouseIgrid_L3",
    # Other proprietary
    "SysTime",
    "MemFree",
    "FlashFree",
    "HDD1Size",
    "HDD1Used",
    "SysLoad",
}

# ==============================================================================
# DISPLAY NAME CORRECTIONS (from update_ha_display_names.py)
# ==============================================================================

DISPLAY_NAME_CORRECTIONS = {
    # Datalogger entities
    "Serial number (datalogger)": "Serial Number",
    "WiFi network name (SSID)": "WiFi SSID",
    "Wlan 0 Mode operating mode": "Wlan 0 mode",
    # Inverter entities
    "Current alarm status of the inverter": "Alarm status",
    "Booster stage temperature": "Booster temperature",
    "DC bulk capacitor voltage": "DC capacitor voltage",
    "DC bulk mid-point voltage": "DC mid-point voltage",
    "DC current measurement for string 1": "DC Current #1",
    "DC current measurement for string 2": "DC Current #2",
    "DC Current (String 1)": "DC Current #1",
    "DC Current (String 2)": "DC Current #2",
    "DC voltage measurement for string 1": "DC Voltage #1",
    "DC voltage measurement for string 2": "DC Voltage #2",
    "DC Voltage (String 1)": "DC Voltage #1",
    "DC Voltage (String 2)": "DC Voltage #2",
    "DC Power (String 1)": "DC Power #1",
    "DC Power (String 2)": "DC Power #2",
    "Phase A to neutral voltage measurement": "Phase Voltage AN",
    "Device Modbus address": "Modbus address",
    "Device model identifier from common model": "Model",
    "Device options and features from common model": "Options",
    "Device serial number from common model": "Serial Number",
    "Firmware version from device common model": "Firmware Version",
    "Manufacturer name from device common model": "Manufacturer",
    # New temperature monitoring corrections
    "Maximum cabinet temperature recorded": "Cabinet Temperature Max",
    "Minimum cabinet temperature recorded": "Cabinet Temperature Min",
    # Digital input corrections
    "Operating mode configuration for digital input 0": "Digital Input 0 Mode",
    "Operating mode configuration for digital input 1": "Digital Input 1 Mode",
    # Energy counter corrections
    "Total energy supplied from backup source": "Backup Energy",
    "Energy measured by current transformer": "Energy CT",
    "Energy measured by direct transducer": "Energy DT",
    "Total DC energy from MPPT string 1": "String 1 Energy",
    "Total DC energy from MPPT string 2": "String 2 Energy",
    "Total DC Energy (String 1)": "String 1 Energy",
    "Total DC Energy (String 2)": "String 2 Energy",
    # String measurement corrections
    "DC current from MPPT string 1": "String 1 Current",
    "DC current from MPPT string 2": "String 2 Current",
    "Combined DC input power from all strings": "Total DC Input Power",
    "DC input power from MPPT string 1": "String 1 Power",
    "DC input power from MPPT string 2": "String 2 Power",
    "DC Input Power (String 1)": "String 1 Power",
    "DC Input Power (String 2)": "String 2 Power",
    # Grid monitoring corrections
    "Grid support Fault Ride Through feature status": "FRT Status",
    "AC frequency measurement for phase A": "Frequency Phase A",
    "AC frequency measurement for phase B": "Frequency Phase B",
    "AC frequency measurement for phase C": "Frequency Phase C",
    "AC Frequency Phase A": "Frequency Phase A",
    "AC Frequency Phase B": "Frequency Phase B",
    "AC Frequency Phase C": "Frequency Phase C",
    "Frequency measured by central controller": "CC Frequency",
    # Leakage current corrections
    "DC side ground fault leakage current": "DC Leakage Current",
    "Inverter ground fault leakage current": "Inverter Leakage Current",
    "Battery DC current flow": "Battery Current",
    # Status and system corrections
    "Current operational state of inverter": "Inverter State",
    "Maximum instantaneous power": "Peak Power",
    "Inverter rated power capacity": "Power Rating",
    "Inverter system timestamp": "System Time",
    "Voltage to ground reference": "Ground Voltage",
    "Power factor (cosine phi)": "Power Factor",
    # M1 Common Model fixes
    "Inverter model number and designation": "Device Model",
    "Inverter configuration and installed options": "Device Options",
    "Inverter firmware version": "Firmware Version",
    "Inverter unique serial number": "Serial Number",
    # Alarm and status display names
    "Current alarm status code": "Alarm Status",
    "Overall inverter operating state": "Global State",
    "Digital input status register": "Digital Input Status",
    "Active warning condition flags": "Warning Flags",
    "AC power derating active condition flags": "AC Power Derating Flags",
    "DC string 1 connection status": "DC Input 1 Status",
    "DC string 2 connection status": "DC Input 2 Status",
    "DC string 1 operating state": "DC Input 1 State",
    "DC string 2 operating state": "DC Input 2 State",
    # Temperature display names
    "Inverter internal temperature": "Inverter Temperature",
    "DC-DC boost converter temperature": "Booster Temperature",
    "Additional temperature sensor 1": "Temperature Sensor 1",
    "Battery pack temperature": "Battery Temperature",
    # Voltage display names
    "DC bus bulk capacitor voltage": "Bulk DC Voltage",
    "Three-phase AC line-to-line voltage between L1 and L2": "AC Voltage A-B",
    "Three-phase AC line-to-line voltage between L2 and L3": "AC Voltage B-C",
    "Three-phase AC line-to-line voltage between L3 and L1": "AC Voltage C-A",
    "AC Voltage Phase A-B": "AC Voltage A-B",
    "AC Voltage Phase B-C": "AC Voltage B-C",
    "AC Voltage Phase C-A": "AC Voltage C-A",
    "Grid line-to-neutral voltage phase L1": "AC Voltage A-N",
    "Grid line-to-neutral voltage phase L2": "AC Voltage B-N",
    "AC Voltage Phase A-N": "AC Voltage A-N",
    "AC Voltage Phase B-N": "AC Voltage B-N",
    "AC Voltage Phase C-N": "AC Voltage C-N",
    "DC string 1 input voltage": "DC Input 1 Voltage",
    "DC string 2 input voltage": "DC Input 2 Voltage",
    "Battery pack voltage": "Battery Voltage",
    "Maximum voltage across battery cells": "Battery Cell Max Voltage",
    "Minimum voltage across battery cells": "Battery Cell Min Voltage",
    # Power display names
    "Battery charge/discharge power": "Battery Power",
    "AC power in stand-alone (off-grid) mode": "AC Power Stand Alone",
    "AC power exported to grid": "AC Power To Grid",
    "Inverter nominal rated power": "Nominal Power",
    "Inverter nominal apparent power rating": "Nominal Apparent Power",
    "AC power factor (cosine phi)": "Power Factor",
    # Energy display names
    "Cumulative energy absorbed from grid": "Total Energy Absorbed",
    "Cumulative apparent energy (kVAh)": "Total Apparent Energy",
    # Battery display name
    "Battery state of charge percentage": "Battery State of Charge",
    # Grid control display names
    "External grid control enabled status": "Grid External Control Enabled",
    "External grid control current state": "Grid External Control State",
    # Other measurement display names
    "Current measurement shunt voltage": "Shunt Voltage",
    "Inverter device type identifier": "Device Type",
    # ==========================================
    # v2.0.7 MASSIVE Display Name Improvements (58 points)
    # ==========================================
    # WiFi/Network Display Names (11)
    "WiFi broadcast address": "WiFi Broadcast",
    "WiFi DHCP state (acquired/pending)": "WiFi DHCP State",
    "WiFi DNS server address": "WiFi DNS Server",
    "WiFi gateway IP address": "WiFi Gateway",
    "WiFi IPv4 address": "WiFi IP Address",
    "WiFi network subnet mask": "WiFi Netmask",
    "WiFi connection status": "WiFi Status",
    "WiFi access point mode status": "WiFi AP Status",
    "WiFi finite state machine status": "WiFi FSM Status",
    "WiFi local IP address": "WiFi Local IP",
    "WiFi operating mode (client/AP)": "WiFi Mode",
    # Meter Display Names (NEW - Phase A/B/C notation)
    "Meter Power Phase A": "Meter Power A",
    "Meter Power Phase B": "Meter Power B",
    "Meter Power Phase C": "Meter Power C",
    "Meter Reactive Power Phase A": "Meter Reactive Power A",
    "Meter Reactive Power Phase B": "Meter Reactive Power B",
    "Meter Reactive Power Phase C": "Meter Reactive Power C",
    "Meter reactive power total": "Meter Reactive Power Total",
    "Meter Voltage Phase A": "Meter Voltage A",
    "Meter Voltage Phase B": "Meter Voltage B",
    "Meter Voltage Phase C": "Meter Voltage C",
    "Meter Current Phase A": "Meter Current A",
    "Meter Current Phase B": "Meter Current B",
    "Meter Current Phase C": "Meter Current C",
    "Meter AC frequency": "Meter Frequency",
    # Device/Status Display Names (16)
    "Device 0 component name": "Device 0 Name",
    "Device 1 component name": "Device 1 Name",
    "Device 2 component name": "Device 2 Name",
    "Device 0 component type": "Device 0 Type",
    "Device 1 component type": "Device 1 Type",
    "Device 2 component type": "Device 2 Type",
    "Device 0 firmware version": "Device 0 Firmware",
    "Device 1 firmware version": "Device 1 Firmware",
    "Device 2 firmware version": "Device 2 Firmware",
    "Firmware build date and time": "Firmware Build Date",
    "Firmware build commit hash": "Firmware Build Hash",
    "Firmware release version number": "Firmware Version",
    "Firmware revision identifier": "Firmware Revision",
    "Manufacturing part number": "Part Number",
    "Manufacturing serial number": "Serial Number",
    "Manufacturing week and year (WWYY)": "Manufacturing Week/Year",
    # Battery/Power Display Names (17)
    "Battery external control enabled flag": "Battery Control Enabled",
    "Battery external control state": "Battery Control State",
    "Battery charge/discharge cycle count": "Battery Cycle Count",
    "Battery charging power": "Battery Charge Power",
    "Battery discharging power": "Battery Discharge Power",
    "Cooling fan 1 speed in RPM": "Fan 1 Speed",
    "Cooling fan 2 speed in RPM": "Fan 2 Speed",
    "Number of MPPT DC input channels": "MPPT Count",
    "Reactive power AC derating flags": "Reactive Power Derating",
    "Apparent power AC derating flags": "Apparent Power Derating",
    "DC bus bulk capacitor midpoint voltage": "DC Bus Midpoint Voltage",
    "DC-to-AC ground leakage current": "DC-AC Leakage Current",
    "DC-to-DC ground leakage current": "DC-DC Leakage Current",
    "DC insulation resistance to ground": "Insulation Resistance",
    "SunSpec Model 126 module enabled": "Model 126 Enabled",
    "SunSpec Model 132 module enabled": "Model 132 Enabled",
    # VSN-specific display name fixes (user-requested)
    "Inverter St": "Inverter Status",
    "Inverter status": "Inverter Status",
    "ram free": "Free RAM",
    "flash free": "Free Flash Storage",
    "Logger Sn": "Logger Serial Number",
    "Logger Logger Id": "Logger ID",
    "Logger ID": "Logger ID",
    # Additional measurement word removals
    "Current shunt voltage": "Shunt Voltage",
    "Grid frequency": "Grid Frequency",
    "Global inverter status": "Global Status",
    # Phase current corrections
    "AC Current Phase A": "Current Phase A",
    "AC Current Phase B": "Current Phase B",
    "AC Current Phase C": "Current Phase C",
    # House consumption corrections
    "House Power Phase A": "House Power A",
    "House Power Phase B": "House Power B",
    "House Power Phase C": "House Power C",
    "House Current Phase A": "House Current A",
    "House Current Phase B": "House Current B",
    "House Current Phase C": "House Current C",
    # Datalogger device entity fixes (remove redundant prefixes)
    "Datalogger product number": "Product Number",
    "Datalogger serial number": "Serial Number",
    "Device 2 Firmware": "Firmware Version",
    "WiFi network name (SSID) connected to": "WiFi network name (SSID)",
    # DC voltage display names - add string context
    "DC Voltage #1": "DC Voltage (String 1)",
    "DC Voltage #2": "DC Voltage (String 2)",
}

# ==============================================================================
# DISPLAY NAME STANDARDIZATION (v1.0.0-beta.22)
# ==============================================================================
# Standardizes display names to pattern: [Type] [AC/DC] - [Details] - [Time Period]
# This enables better grouping in Home Assistant's entity lists and UI
# Format: {current_display_name: standardized_display_name}

DISPLAY_NAME_STANDARDIZATION = {
    # ========== POWER ENTITIES (24 changes) ==========
    "AC Power": "Power AC",
    "DC Power": "Power DC",
    "DC Power #1": "Power DC - String 1",
    "DC Power #2": "Power DC - String 2",
    "House Power A": "Power AC - House Phase A",
    "House Power B": "Power AC - House Phase B",
    "House Power C": "Power AC - House Phase C",
    "House load active power - Total": "Power AC - House Load Total",
    "Battery Discharge Power": "Power - Battery Discharge",
    "Battery Charge Power": "Power - Battery Charge",
    "Battery Power": "Power - Battery Total",
    "Battery power - Total": "Power - Battery Total",
    "Total DC Input Power": "Power DC - Input Total",
    "String 1 Power": "Power DC - String 1",
    "String 2 Power": "Power DC - String 2",
    "Peak Power": "Power - Peak",
    "AC Power To Grid": "Power AC - Export to Grid",
    "AC Power Stand Alone": "Power AC - Stand Alone Mode",
    "Home PV Power": "Power - Home PV",
    "Meter Power A": "Power AC - Meter Phase A",
    "Meter Power B": "Power AC - Meter Phase B",
    "Meter Power C": "Power AC - Meter Phase C",
    "Meter active power - Total": "Power AC - Meter Total",
    "Power Rating": "Power - Rating",
    # ========== CURRENT ENTITIES (19 changes) ==========
    "AC Current": "Current AC",
    "Phase A Current": "Current AC - Phase A",
    "Phase B Current": "Current AC - Phase B",
    "Phase C Current": "Current AC - Phase C",
    "DC Current #1": "Current DC - String 1",
    "DC Current #2": "Current DC - String 2",
    "House Current A": "Current AC - House Phase A",
    "House Current B": "Current AC - House Phase B",
    "House Current C": "Current AC - House Phase C",
    "DC-DC Leakage Current": "Current DC - Leakage DC-DC",
    "DC-AC Leakage Current": "Current DC - Leakage DC-AC",
    "Inverter Leakage Current": "Current - Inverter Leakage",
    "DC Leakage Current": "Current DC - Leakage",
    "Battery Current": "Current - Battery",
    "String 1 Current": "Current DC - String 1",
    "String 2 Current": "Current DC - String 2",
    "Meter Current A": "Current AC - Meter Phase A",
    "Meter Current B": "Current AC - Meter Phase B",
    "Meter Current C": "Current AC - Meter Phase C",
    # ========== VOLTAGE ENTITIES (27 changes) ==========
    "AC Voltage A-N": "Voltage AC - Phase A-N",
    "Phase Voltage BN": "Voltage AC - Phase B-N",
    "Phase Voltage CN": "Voltage AC - Phase C-N",
    "AC Voltage A-B": "Voltage AC - Phase A-B",
    "AC Voltage B-C": "Voltage AC - Phase B-C",
    "AC Voltage C-A": "Voltage AC - Phase C-A",
    "DC Voltage #1": "Voltage DC - String 1",
    "DC Voltage #2": "Voltage DC - String 2",
    "DC Voltage (Bulk Capacitor)": "Voltage DC - Bulk Capacitor",
    "DC Bus Midpoint Voltage": "Voltage DC - Bus Midpoint",
    "Ground Voltage": "Voltage - Ground",
    "Battery Voltage": "Voltage - Battery",
    "Battery Cell Max Voltage": "Voltage - Battery Cell Max",
    "Battery Cell Min Voltage": "Voltage - Battery Cell Min",
    "DC Input 1 Voltage": "Voltage DC - Input 1",
    "DC Input 2 Voltage": "Voltage DC - Input 2",
    "Shunt Voltage": "Voltage - Shunt",
    "Meter Voltage A": "Voltage AC - Meter Phase A",
    "Meter Voltage B": "Voltage AC - Meter Phase B",
    "Meter Voltage C": "Voltage AC - Meter Phase C",
    "DC Voltage (String 1)": "Voltage DC - String 1",
    "DC Voltage (String 2)": "Voltage DC - String 2",
    "Bulk DC Voltage": "Voltage DC - Bulk Capacitor",
    "Phase Voltage AN": "Voltage AC - Phase A-N",
    "AC Voltage B-N": "Voltage AC - Phase B-N",
    "AC Voltage C-N": "Voltage AC - Phase C-N",
    "DC capacitor voltage": "Voltage DC - Bulk Capacitor",
    # ========== FREQUENCY ENTITIES (6 changes) ==========
    "Grid Frequency": "Frequency AC - Grid",
    "Frequency Phase A": "Frequency AC - Phase A",
    "Frequency Phase B": "Frequency AC - Phase B",
    "Frequency Phase C": "Frequency AC - Phase C",
    "CC Frequency": "Frequency - Central Controller",
    "Meter Frequency": "Frequency AC - Meter",
    # ========== REACTIVE POWER ENTITIES (5 changes) ==========
    "Reactive power at grid connection point": "Reactive Power - Grid Connection",
    "Meter Reactive Power A": "Reactive Power AC - Meter Phase A",
    "Meter Reactive Power B": "Reactive Power AC - Meter Phase B",
    "Meter Reactive Power C": "Reactive Power AC - Meter Phase C",
    "Meter reactive power - Total": "Reactive Power AC - Meter Total",
    # ========== ENERGY ENTITIES (75 changes) ==========
    # Base energy entities
    "AC Energy": "Energy AC - Lifetime",
    "Energy - Today": "Energy AC - Produced (Today)",
    "Energy - Year": "Energy - (Current Year)",
    "Energy - Week": "Energy - (Current Week)",
    "Energy - Month": "Energy - (Current Month)",
    "Battery discharge energy - Lifetime": "Energy - Battery Discharge Lifetime",
    "Battery charge energy - Lifetime": "Energy - Battery Charge Lifetime",
    "Inverter PV energy (DC input) - Lifetime": "Energy DC - PV Input Lifetime",
    "Inverter AC energy produced - Lifetime": "Energy AC - Produced Lifetime",
    "String 1 Energy": "Energy DC - String 1 Lifetime",
    "String 2 Energy": "Energy DC - String 2 Lifetime",
    "Total Energy Absorbed": "Energy AC - Absorbed Lifetime",
    "Total Apparent Energy": "Energy AC Apparent - Lifetime",
    "Energy imported from grid - Lifetime": "Energy AC - Grid Import Lifetime",
    "Energy exported to grid - Lifetime": "Energy AC - Grid Export Lifetime",
    "Inverter backup output energy - Lifetime": "Energy AC - Backup Output Lifetime",
    # E0 series (Inverter AC energy produced)
    "Inverter AC energy produced - Since Restart": "Energy AC - Produced (Since Restart)",
    "Inverter AC energy produced - Last 7 Days": "Energy AC - Produced (Current Week)",
    "Inverter AC energy produced - Last 30 Days": "Energy AC - Produced (Current Month)",
    "Inverter AC energy produced - Last Year": "Energy AC - Produced (Current Year)",
    # E1 series (Inverter absorbed AC energy)
    "Inverter absorbed AC energy - Since Restart": "Energy AC - Absorbed (Since Restart)",
    "Inverter absorbed AC energy - Last 7 Days": "Energy AC - Absorbed (Current Week)",
    "Inverter absorbed AC energy - Last 30 Days": "Energy AC - Absorbed (Current Month)",
    "Inverter absorbed AC energy - Lifetime": "Energy AC - Absorbed Lifetime",
    # E2 series (Inverter apparent AC energy)
    "Inverter apparent AC energy - Since Restart": "Energy AC Apparent - (Since Restart)",
    "Inverter apparent AC energy - Last 7 Days": "Energy AC Apparent - (Current Week)",
    "Inverter apparent AC energy - Last 30 Days": "Energy AC Apparent - (Current Month)",
    "Inverter apparent AC energy - Lifetime": "Energy AC Apparent - Lifetime",
    # E3 series (Energy exported to grid)
    "Energy exported to grid - Since Restart": "Energy AC - Grid Export (Since Restart)",
    "Energy exported to grid - Last 7 Days": "Energy AC - Grid Export (Current Week)",
    "Energy exported to grid - Last 30 Days": "Energy AC - Grid Export (Current Month)",
    # E4 series (Self-consumed energy)
    "Self-consumed energy - Since Restart": "Energy - Self-Consumed (Since Restart)",
    "Self-consumed energy - Last 7 Days": "Energy - Self-Consumed (Current Week)",
    "Self-consumed energy - Last 30 Days": "Energy - Self-Consumed (Current Month)",
    "Self-consumed energy - Last Year": "Energy - Self-Consumed (Current Year)",
    # E6 series (House consumption from inverter)
    "House consumption from inverter - Lifetime": "Energy - House from Inverter Lifetime",
    "House consumption from inverter - Since Restart": "Energy - House from Inverter (Since Restart)",
    "House consumption from inverter - Last 7 Days": "Energy - House from Inverter (Current Week)",
    "House consumption from inverter - Last 30 Days": "Energy - House from Inverter (Current Month)",
    # E7 series (Total house consumption)
    "Total house consumption - Lifetime": "Energy - House Consumption Lifetime",
    "Total house consumption - Since Restart": "Energy - House Consumption (Since Restart)",
    "Total house consumption - Last 7 Days": "Energy - House Consumption (Current Week)",
    "Total house consumption - Last 30 Days": "Energy - House Consumption (Current Month)",
    # E8 series (Energy imported from grid)
    "Energy imported from grid - Since Restart": "Energy AC - Grid Import (Since Restart)",
    "Energy imported from grid - Last 7 Days": "Energy AC - Grid Import (Current Week)",
    "Energy imported from grid - Last 30 Days": "Energy AC - Grid Import (Current Month)",
    # E15 series (Inverter backup output energy)
    "Inverter backup output energy - Since Restart": "Energy AC - Backup Output (Since Restart)",
    "Inverter backup output energy - Last 7 Days": "Energy AC - Backup Output (Current Week)",
    "Inverter backup output energy - Last 30 Days": "Energy AC - Backup Output (Current Month)",
    # Ein series (Inverter PV energy DC input)
    "Inverter PV energy (DC input) - Since Restart": "Energy DC - PV Input (Since Restart)",
    "Inverter PV energy (DC input) - Last 7 Days": "Energy DC - PV Input (Current Week)",
    "Inverter PV energy (DC input) - Last 30 Days": "Energy DC - PV Input (Current Month)",
    # Total battery charge/discharge energy
    "Total battery charge energy - Lifetime": "Energy - Battery Charge Total Lifetime",
    "Total battery charge energy - Since Restart": "Energy - Battery Charge Total (Since Restart)",
    "Total battery charge energy - Last 7 Days": "Energy - Battery Charge Total (Current Week)",
    "Total battery charge energy - Last 30 Days": "Energy - Battery Charge Total (Current Month)",
    "Total battery discharge energy - Lifetime": "Energy - Battery Discharge Total Lifetime",
    "Total battery discharge energy - Since Restart": "Energy - Battery Discharge Total (Since Restart)",
    "Total battery discharge energy - Last 7 Days": "Energy - Battery Discharge Total (Current Week)",
    "Total battery discharge energy - Last 30 Days": "Energy - Battery Discharge Total (Current Month)",
    # Individual battery charge/discharge energy
    "Battery charge energy - Since Restart": "Energy - Battery Charge (Since Restart)",
    "Battery charge energy - Last 7 Days": "Energy - Battery Charge (Current Week)",
    "Battery charge energy - Last 30 Days": "Energy - Battery Charge (Current Month)",
    "Battery discharge energy - Since Restart": "Energy - Battery Discharge (Since Restart)",
    "Battery discharge energy - Last 7 Days": "Energy - Battery Discharge (Current Week)",
    "Battery discharge energy - Last 30 Days": "Energy - Battery Discharge (Current Month)",
    # ========== TEMPERATURE ENTITIES (9 changes) ==========
    "Cabinet Temperature": "Temperature - Cabinet",
    "Other Temperature": "Temperature - Other",
    "Booster temperature": "Temperature - Booster",
    "Inverter Temperature": "Temperature - Inverter",
    "Booster Temperature": "Temperature - Booster",
    "Cabinet Temperature Max": "Temperature - Cabinet Max",
    "Cabinet Temperature Min": "Temperature - Cabinet Min",
    "Battery Temperature": "Temperature - Battery",
    "Temperature Sensor 1": "Temperature - Sensor 1",
    # ========== RESISTANCE ENTITIES (2 changes) ==========
    "Insulation Resistance": "Resistance - Insulation",
    # ========== MISCELLANEOUS MEASUREMENT ENTITIES (20 changes) ==========
    "Power factor": "Power Factor",
    "Power Factor": "Power Factor",
    "Power Factor (Cos Phi)": "Power Factor - Cos Phi",
    "MPPT Count": "Count - MPPT Channels",
    "Battery Cycle Count": "Count - Battery Cycles",
    "Fan 1 Speed": "Speed - Fan 1",
    "Fan 2 Speed": "Speed - Fan 2",
    "Peak Output Power - Today": "Power - Peak Today",
    "Peak Output Power - Lifetime": "Power - Peak Lifetime",
    "Nominal Power": "Power - Nominal Rating",
    "Nominal Apparent Power": "Power - Nominal Apparent Rating",
}

# Device class/unit fixes (keyed by label, not SunSpec name)
DEVICE_CLASS_FIXES = {
    "Device Address": {
        "device_class": None,
        "unit": None,
        "entity_category": "diagnostic",
    },
    "Version": {"device_class": None, "unit": None, "entity_category": "diagnostic"},
    # System time: Remove timestamp device_class to show formatted date/time instead of relative time
    "Sys Time": {"device_class": None, "entity_category": "diagnostic"},
    # Device information fields should be diagnostic (use label names)
    "Manufacturer": {"entity_category": "diagnostic"},
    "Model": {"entity_category": "diagnostic"},
    "Serial Number": {
        "entity_category": "diagnostic"
    },  # Applies to both inverter SN and datalogger sn
    "Firmware Version": {"entity_category": "diagnostic"},
    # Datalogger identification fields
    "Product Number": {
        "device_class": None,
        "unit": None,
        "entity_category": "diagnostic",
    },
    # Inverter status entities should be diagnostic
    "Alarm St": {"entity_category": "diagnostic"},
    "Dc St1": {"entity_category": "diagnostic"},
    "Dc St2": {"entity_category": "diagnostic"},
    "Options": {"entity_category": "diagnostic"},
    "Global St": {"entity_category": "diagnostic"},
    "Inverter St": {"entity_category": "diagnostic"},
    # Configuration settings should be diagnostic
    "Country Std": {"entity_category": "diagnostic"},
}

# ==============================================================================
# v2.0.9 LABEL CORRECTIONS
# ==============================================================================

LABEL_CORRECTIONS = {
    # DC spacing fixes (v2.0.9)
    "D C1 State": "DC1 State",
    "D C2 State": "DC2 State",
    # Datalogger identification fields (not power!)
    "Sn": "Serial Number",
    "Pn": "Product Number",
}

# ==============================================================================
# v2.0.9 UNIT CORRECTIONS
# ==============================================================================

UNIT_CORRECTIONS = {
    # Wh vs kWh fixes (v2.0.9) - these points have "Wh" in name but were set to kWh
    "DayWH": "Wh",
    "MonthWH": "Wh",
    "WeekWH": "Wh",
    "YearWH": "Wh",
    "WH": "Wh",  # AC Energy should be Wh not kWh
    # uA to mA fixes - HA doesn't support uA for current device class
    "ILeakDcAc": "mA",
    "ILeakDcDc": "mA",
    # System monitoring unit fixes
    "sys_load": "",  # Fix "none" unit to empty string
    # DC power unit fixes - feeds data has kW but should be W
    "DCW": "W",
    "DCW_1": "W",
    "DCW_2": "W",
}

# Temperature unit normalization - degC → °C
TEMPERATURE_UNIT_FIX = {
    "degC": "°C",
    "Deg C": "°C",
    "deg C": "°C",
}

# ==============================================================================
# v2.0.9 COMPREHENSIVE HA METADATA DICTIONARY
# ==============================================================================

SUNSPEC_TO_HA_METADATA = {
    # Energy measurements (kWh, total_increasing)
    "WH": {"device_class": "energy", "state_class": "total_increasing", "unit": "kWh"},
    "TotWhExp": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "TotWhImp": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "ETotal": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "Ein": {"device_class": "energy", "state_class": "total_increasing", "unit": "kWh"},
    "ECharge": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "EDischarge": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "ETotCharge": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "ETotDischarge": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "EGridImport": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "EGridExport": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "ETotalAbsorbed": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "kWh",
    },
    "ETotalApparent": {
        "device_class": "energy",
        "state_class": "total",
        "unit": "kVAh",
    },
    # Energy counters by period (total)
    # Note: Values come as Wh from API, normalizer converts to kWh, so unit must be kWh
    "DayWH": {"device_class": "energy", "state_class": "total", "unit": "kWh"},
    "WeekWH": {"device_class": "energy", "state_class": "total", "unit": "kWh"},
    "MonthWH": {"device_class": "energy", "state_class": "total", "unit": "kWh"},
    "YearWH": {"device_class": "energy", "state_class": "total", "unit": "kWh"},
    # Energy counters - E series (Wh, total)
    "E0_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E0_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E0_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E0_1Y": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E1_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E1_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E1_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E3_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E3_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E3_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E6_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E6_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E6_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E6_total": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E7_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E7_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E7_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E7_total": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E8_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E8_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E8_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E15_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "E15_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "E15_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "Ein_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "Ein_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "Ein_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "ECharge_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "ECharge_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "ECharge_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "EDischarge_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "EDischarge_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "EDischarge_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "ETotCharge_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "ETotCharge_7D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "ETotCharge_30D": {"device_class": "energy", "state_class": "total", "unit": "Wh"},
    "ETotDischarge_runtime": {
        "device_class": "energy",
        "state_class": "total_increasing",
        "unit": "Wh",
    },
    "ETotDischarge_7D": {
        "device_class": "energy",
        "state_class": "total",
        "unit": "Wh",
    },
    "ETotDischarge_30D": {
        "device_class": "energy",
        "state_class": "total",
        "unit": "Wh",
    },
    # Power measurements (W, measurement)
    "W": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "WphA": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "WphB": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "WphC": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "DCW": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "DCW_1": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "DCW_2": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "PCh": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "PDh": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "WRtg": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "PacStandAlone": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "PacTogrid": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "Pin": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "Pin1": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "Pin2": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "PbaT": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    "MeterPgrid_Tot": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "HousePgrid_Tot": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "HousePgrid_L1": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "HousePgrid_L2": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "HousePgrid_L3": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "HousePInverter": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    # Note: "pn" in datalogger context is product number (text), not power
    # Removed from here to avoid confusion
    # Reactive Power (var, measurement)
    "VAr": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "VArphA": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "VArphB": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "VArphC": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "MeterQgrid_L1": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "MeterQgrid_L2": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "MeterQgrid_L3": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    "MeterQgrid_Tot": {
        "device_class": "reactive_power",
        "state_class": "measurement",
        "unit": "var",
    },
    # Apparent Power (VA, measurement)
    "VA": {
        "device_class": "apparent_power",
        "state_class": "measurement",
        "unit": "VA",
    },
    "VAphA": {
        "device_class": "apparent_power",
        "state_class": "measurement",
        "unit": "VA",
    },
    "VAphB": {
        "device_class": "apparent_power",
        "state_class": "measurement",
        "unit": "VA",
    },
    "VAphC": {
        "device_class": "apparent_power",
        "state_class": "measurement",
        "unit": "VA",
    },
    "VARtg": {
        "device_class": "apparent_power",
        "state_class": "measurement",
        "unit": "VA",
    },
    # Note: "sn" in datalogger context is serial number (text), not apparent power
    # Removed from here to avoid confusion
    # Current (A, measurement)
    "A": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "AphA": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "AphB": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "AphC": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "DCA": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "DCA_1": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "DCA_2": {"device_class": "current", "state_class": "measurement", "unit": "A"},
    "MeterIgrid_L1": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "A",
    },
    "MeterIgrid_L2": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "A",
    },
    "MeterIgrid_L3": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "A",
    },
    "HouseIgrid_L1": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "A",
    },
    "HouseIgrid_L2": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "A",
    },
    "HouseIgrid_L3": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "A",
    },
    "ILeakDcAc": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "mA",
    },
    "ILeakDcDc": {
        "device_class": "current",
        "state_class": "measurement",
        "unit": "mA",
    },
    # Voltage (V, measurement)
    "PhVphA": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PhVphB": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PhVphC": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PPVphAB": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PPVphBC": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PPVphCA": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PhVphAB": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PhVphBC": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "PhVphCA": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "DCV": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "DCV_1": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "DCV_2": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VGnd": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "Vgnd": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VBulk": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VBulkMid": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VgridL1_N": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VgridL2_N": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VgridL3_N": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "MeterVgrid_L1": {
        "device_class": "voltage",
        "state_class": "measurement",
        "unit": "V",
    },
    "MeterVgrid_L2": {
        "device_class": "voltage",
        "state_class": "measurement",
        "unit": "V",
    },
    "MeterVgrid_L3": {
        "device_class": "voltage",
        "state_class": "measurement",
        "unit": "V",
    },
    "VcMax": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    "VcMin": {"device_class": "voltage", "state_class": "measurement", "unit": "V"},
    # Frequency (Hz, measurement)
    "Hz": {"device_class": "frequency", "state_class": "measurement", "unit": "Hz"},
    "HzA": {"device_class": "frequency", "state_class": "measurement", "unit": "Hz"},
    "HzB": {"device_class": "frequency", "state_class": "measurement", "unit": "Hz"},
    "HzC": {"device_class": "frequency", "state_class": "measurement", "unit": "Hz"},
    "MeterFgrid": {
        "device_class": "frequency",
        "state_class": "measurement",
        "unit": "Hz",
    },
    # Temperature (°C, measurement)
    "TmpCab": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TmpSnk": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TmpTrns": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TmpOt": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TmpInv": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TmpBst": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TempInv": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "TempBst": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "Temp1": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    "Booster_Tmp": {
        "device_class": "temperature",
        "state_class": "measurement",
        "unit": "°C",
    },
    # Battery (%, measurement)
    "Soc": {"device_class": "battery", "state_class": "measurement", "unit": "%"},
    "Soh": {"device_class": "battery", "state_class": "measurement", "unit": "%"},
    # Power Factor (no unit, measurement)
    "PF": {"device_class": "power_factor", "state_class": "measurement", "unit": ""},
    "cosPhi": {
        "device_class": "power_factor",
        "state_class": "measurement",
        "unit": "",
    },
    # Fan Speed (RPM, measurement)
    "Fan1rpm": {"device_class": None, "state_class": "measurement", "unit": "RPM"},
    "Fan2rpm": {"device_class": None, "state_class": "measurement", "unit": "RPM"},
    # Network Monitoring (%, measurement)
    # Note: WiFi link quality is a percentage (0-100%), not signal strength in dB/dBm
    "wlan0_link_quality": {
        "device_class": None,
        "state_class": "measurement",
        "unit": "%",
        "icon": "mdi:wifi-cog",
    },
    "wlan0_ipaddr_local": {"icon": "mdi:wifi-cog"},
    "wlan0_mode": {"icon": "mdi:wifi-cog"},
    "wlan0_essid": {"icon": "mdi:wifi-cog"},
    # System Monitoring
    "sys_load": {"icon": "mdi:gauge"},
    "type": {"icon": "mdi:information-box-outline"},
    # Data Size (MB, measurement) - Changed from B to MB for better readability
    # Values are converted from bytes to MB in normalizer.py
    "store_size": {
        "device_class": "data_size",
        "state_class": "measurement",
        "unit": "MB",
    },
    "flash_free": {
        "device_class": "data_size",
        "state_class": "measurement",
        "unit": "MB",
    },
    "free_ram": {
        "device_class": "data_size",
        "state_class": "measurement",
        "unit": "MB",
    },
    # Duration (s, total_increasing)
    "uptime": {
        "device_class": "duration",
        "state_class": "total_increasing",
        "unit": "s",
    },
    # Apparent Energy (VAh) - no standard HA device_class, use energy icon
    "E2_runtime": {
        "device_class": None,
        "state_class": "total_increasing",
        "unit": "VAh",
        "icon": "mdi:lightning-bolt",
    },
    "E2_7D": {
        "device_class": None,
        "state_class": "total",
        "unit": "VAh",
        "icon": "mdi:lightning-bolt",
    },
    "E2_30D": {
        "device_class": None,
        "state_class": "total",
        "unit": "VAh",
        "icon": "mdi:lightning-bolt",
    },
    # Peak Power (W, measurement) - power device_class
    # Note: Uses W (not kW) per SunSpec specification and livedata format
    # Feeds endpoint shows kW for display, but livedata provides W values
    "PowerPeakAbs": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "PowerPeakToday": {
        "device_class": "power",
        "state_class": "measurement",
        "unit": "W",
    },
    "Ppeak": {"device_class": "power", "state_class": "measurement", "unit": "W"},
    # ===========================================================================
    # CYCLE COUNTERS (v1.1.5+) - No unit, state_class=total_increasing, precision=0
    # ===========================================================================
    "Chc": {
        "device_class": None,
        "state_class": "total_increasing",
        "unit": "",
        "precision": 0,
    },
    "Dhc": {
        "device_class": None,
        "state_class": "total_increasing",
        "unit": "",
        "precision": 0,
    },
    "CycleNum": {
        "device_class": None,
        "state_class": "total_increasing",
        "unit": "",
        "precision": 0,
    },
    # ===========================================================================
    # STATE SENSORS (v1.1.6+) - Diagnostic, no unit, NO precision (display text, not numbers)
    # ===========================================================================
    "GlobalSt": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
    },
    "DcSt1": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
    },
    "DcSt2": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
    },
    "InverterSt": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
    },
    "AlarmSt": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
    },
    "AlarmState": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
    },
    # ===========================================================================
    # INTEGER COUNT SENSORS (v1.1.7+) - Precision=0
    # ===========================================================================
    "BattNum": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "precision": 0,
    },
    "NumOfMPPT": {
        "device_class": None,
        "state_class": "measurement",
        "unit": "channels",
        "precision": 0,
    },
    "CountryStd": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "icon": "mdi:information-box-outline",
        "precision": 0,
    },
    # ===========================================================================
    # STATE/FLAG SENSORS (v1.1.5+) - Diagnostic, precision=0
    # ===========================================================================
    "BatteryMode": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "precision": 0,
    },
    "BattExtCtrlState": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "precision": 0,
    },
    "BattExtCtrlEna": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "precision": 0,
    },
    "PacDeratingFlags": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "precision": 0,
    },
    "SacDeratingFlags": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "precision": 0,
    },
    "ClockState": {
        "device_class": None,
        "state_class": None,
        "unit": "",
        "entity_category": "diagnostic",
        "precision": 0,
    },
    # ===========================================================================
    # RESISTANCE SENSORS (v1.1.4+) - Omega icon, precision=2
    # ===========================================================================
    "Isolation_Ohm1": {
        "device_class": None,  # Exception - no HA device_class supports MOhm
        "state_class": "measurement",
        "unit": "MOhm",
        "icon": "mdi:omega",
        "precision": 2,
    },
    # ===========================================================================
    # CURRENT SENSORS (v1.1.3+) - Precision=2 for Ampere measurements
    # ===========================================================================
    "Iba": {
        "device_class": "current",
        "unit": "A",
        "state_class": "measurement",
        "precision": 2,
    },
    "Iin1": {
        "device_class": "current",
        "unit": "A",
        "state_class": "measurement",
        "precision": 2,
    },
    "Iin2": {
        "device_class": "current",
        "unit": "A",
        "state_class": "measurement",
        "precision": 2,
    },
    "IleakInv": {
        "device_class": "current",
        "unit": "mA",  # Standardized to mA (conversion from A in normalizer)
        "state_class": "measurement",
        "precision": 2,
    },
    "IleakDC": {
        "device_class": "current",
        "unit": "mA",  # Standardized to mA (conversion from A in normalizer)
        "state_class": "measurement",
        "precision": 2,
    },
    # ===========================================================================
    # VOLTAGE SENSORS (v1.1.3+) - Precision=1 for Volt measurements
    # ===========================================================================
    "Vba": {
        "device_class": "voltage",
        "unit": "V",
        "state_class": "measurement",
        "precision": 1,
    },
    "ShU": {
        "device_class": "voltage",
        "unit": "V",
        "state_class": "measurement",
        "precision": 1,
    },
    "Vin1": {
        "device_class": "voltage",
        "unit": "V",
        "state_class": "measurement",
        "precision": 1,
    },
    "Vin2": {
        "device_class": "voltage",
        "unit": "V",
        "state_class": "measurement",
        "precision": 1,
    },
    # ===========================================================================
    # POWER SENSORS (v1.1.3+) - Precision=0 for Watt measurements
    # ===========================================================================
    "Pba": {
        "device_class": "power",
        "unit": "W",
        "state_class": "measurement",
        "precision": 0,
    },
    "MeterPgrid_L1": {
        "device_class": "power",
        "unit": "W",
        "state_class": "measurement",
        "precision": 0,
    },
    "MeterPgrid_L2": {
        "device_class": "power",
        "unit": "W",
        "state_class": "measurement",
        "precision": 0,
    },
    "MeterPgrid_L3": {
        "device_class": "power",
        "unit": "W",
        "state_class": "measurement",
        "precision": 0,
    },
    # ===========================================================================
    # TEMPERATURE SENSORS (v1.1.3+) - Precision=1 for Celsius measurements
    # ===========================================================================
    "Tba": {
        "device_class": "temperature",
        "unit": "°C",
        "state_class": "measurement",
        "precision": 1,
    },
    "TcMax": {
        "device_class": "temperature",
        "unit": "°C",
        "state_class": "measurement",
        "precision": 1,
    },
    "TcMin": {
        "device_class": "temperature",
        "unit": "°C",
        "state_class": "measurement",
        "precision": 1,
    },
    # ===========================================================================
    # ENERGY SENSORS (v1.1.3+) - Precision=2 for Wh measurements with decimals
    # ===========================================================================
    "ECt": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "EDt": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "E4_runtime": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "E4_7D": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "E4_30D": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "E5_runtime": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "E5_7D": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "E5_30D": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "Ein1": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "Ein2": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    "EBackup": {
        "device_class": "energy",
        "unit": "Wh",
        "state_class": "total_increasing",
        "precision": 2,
    },
    # ===========================================================================
    # FREQUENCY SENSORS (v1.1.3+) - Precision=2 for Hz measurements
    # ===========================================================================
    "Fcc": {
        "device_class": "frequency",
        "unit": "Hz",
        "state_class": "measurement",
        "precision": 2,
    },
    # ===========================================================================
    # BATTERY SOC (v1.1.3+) - Precision=0 for percentage
    # ===========================================================================
    "TSoc": {
        "device_class": "battery",
        "unit": "%",
        "state_class": "measurement",
        "precision": 0,
    },
    # ===========================================================================
    # TIMESTAMP (v1.1.3+) - NO device_class (displays formatted string, not "2 hours ago")
    # ===========================================================================
    "SysTime": {
        "device_class": None,  # Not "timestamp" - we want formatted string, not relative time
        "unit": "",
        "state_class": "",
        # No precision - returns formatted date string, not numeric value (v1.1.9)
    },
}

# ==============================================================================
# DESCRIPTION IMPROVEMENTS (from apply_mapping_fixes.py)
# ==============================================================================

DESCRIPTION_IMPROVEMENTS = {
    # M1 Common Model - Fixed generic descriptions
    # Battery specific
    "Tmp": "Battery temperature",
    "BattNum": "Number of battery cells or modules",
    "Chc": "Battery charge cycles counter",
    "Dhc": "Battery discharge cycles counter",
    "ETotCharge": "Total battery charge energy - Lifetime",
    "ETotDischarge": "Total battery discharge energy - Lifetime",
    # Energy counters with periods
    # E0 series: Inverter AC energy produced (from VSN700 feed title)
    "E0_runtime": "Inverter AC energy produced - Since Restart",
    "E0_7D": "Inverter AC energy produced - Last 7 Days",
    "E0_30D": "Inverter AC energy produced - Last 30 Days",
    "E0_1Y": "Inverter AC energy produced - Last Year",
    # E1 series: Inverter absorbed AC energy (from VSN700 feed title)
    "E1_runtime": "Inverter absorbed AC energy - Since Restart",
    "E1_7D": "Inverter absorbed AC energy - Last 7 Days",
    "E1_30D": "Inverter absorbed AC energy - Last 30 Days",
    "E1_1Y": "Inverter absorbed AC energy - Last Year",
    # E2 series: Inverter apparent AC energy (from VSN700 feed title)
    "E2_runtime": "Inverter apparent AC energy - Since Restart",
    "E2_7D": "Inverter apparent AC energy - Last 7 Days",
    "E2_30D": "Inverter apparent AC energy - Last 30 Days",
    "E2_1Y": "Inverter apparent AC energy - Last Year",
    # E3 series: Energy exported to grid (from VSN700 feed title: "E3 - Energy to the Grid")
    "E3_runtime": "Energy exported to grid - Since Restart",
    "E3_7D": "Energy exported to grid - Last 7 Days",
    "E3_30D": "Energy exported to grid - Last 30 Days",
    "E3_1Y": "Energy exported to grid - Last Year",
    # E4 series: Self-consumed energy (keep current, not in VSN700 feeds)
    "E4_runtime": "Energy - Self-Consumed (Since Restart)",
    "E4_7D": "Energy - Self-Consumed (Current Week)",
    "E4_30D": "Energy - Self-Consumed (Current Month)",
    "E4_1Y": "Energy - Self-Consumed (Current Year)",
    # E5 series: Battery charge energy (keep current, not in VSN700 feeds)
    "E5_runtime": "Battery charge energy - Since Restart",
    "E5_7D": "Battery charge energy - Last 7 Days",
    "E5_30D": "Battery charge energy - Last 30 Days",
    "E5_1Y": "Battery charge energy - Last Year",
    # E6 series: House consumption from inverter (from VSN700 feed title: "E6 - Consumed Energy from Inverter")
    "E6_runtime": "House consumption from inverter - Since Restart",
    "E6_7D": "House consumption from inverter - Last 7 Days",
    "E6_30D": "House consumption from inverter - Last 30 Days",
    "E6_1Y": "House consumption from inverter - Last Year",
    # E7 series: Total house consumption (from VSN700 feed title: "E7 - Consumed Energy")
    "E7_runtime": "Total house consumption - Since Restart",
    "E7_7D": "Total house consumption - Last 7 Days",
    "E7_30D": "Total house consumption - Last 30 Days",
    "E7_1Y": "Total house consumption - Last Year",
    # E8 series: Energy imported from grid (from VSN700 feed title: "E8 - Energy from the Grid")
    "E8_runtime": "Energy imported from grid - Since Restart",
    "E8_7D": "Energy imported from grid - Last 7 Days",
    "E8_30D": "Energy imported from grid - Last 30 Days",
    "E8_1Y": "Energy imported from grid - Last Year",
    # House meter
    "HousePgrid_L1": "House Power Phase A",
    "HousePgrid_L2": "House Power Phase B",
    "HousePgrid_L3": "House Power Phase C",
    "HouseIgrid_L1": "House Current Phase A",
    "HouseIgrid_L2": "House Current Phase B",
    "HouseIgrid_L3": "House Current Phase C",
    # System info
    "SplitPhase": "Split-phase configuration flag",
    "CountryStd": "Country grid standard setting",
    # Temperature monitoring
    "TcMax": "Maximum cabinet temperature recorded",
    "TcMin": "Minimum cabinet temperature recorded",
    # Digital inputs
    "DI0_mode": "Operating mode configuration for digital input 0",
    "DI1_mode": "Operating mode configuration for digital input 1",
    # Energy counters - base totals (lifetime cumulative)
    "EBackup": "Inverter backup output energy - Lifetime",
    "ETotal": "Inverter AC energy produced - Lifetime",
    "Ein": "Inverter PV energy (DC input) - Lifetime",
    "Ein1": "DC energy (String 1) - Lifetime",
    "Ein2": "DC energy (String 2) - Lifetime",
    "ECharge": "Battery charge energy - Lifetime",
    "EDischarge": "Battery discharge energy - Lifetime",
    "EGridImport": "Energy imported from grid - Lifetime",
    "EGridExport": "Energy exported to grid - Lifetime",
    "ETotalAbsorbed": "Inverter absorbed AC energy - Lifetime",
    "ETotalApparent": "Inverter apparent AC energy - Lifetime",
    # Energy by time periods
    "DayWH": "Energy AC - Produced (Today)",
    "WeekWH": "Energy - Week",
    "MonthWH": "Energy - Month",
    "YearWH": "Energy - Year",
    "E6_total": "House consumption from inverter - Lifetime",
    "E7_total": "Total house consumption - Lifetime",
    # E15 series: Inverter backup output energy (from VSN700 feed title)
    "E15_runtime": "Inverter backup output energy - Since Restart",
    "E15_7D": "Inverter backup output energy - Last 7 Days",
    "E15_30D": "Inverter backup output energy - Last 30 Days",
    # Ein series: Inverter PV energy DC input (from VSN700 feed title)
    "Ein_runtime": "Inverter PV energy (DC input) - Since Restart",
    "Ein_7D": "Inverter PV energy (DC input) - Last 7 Days",
    "Ein_30D": "Inverter PV energy (DC input) - Last 30 Days",
    # ECharge/EDischarge series: Per-battery measurements (from VSN700 feed title)
    "ECharge_runtime": "Battery charge energy - Since Restart",
    "ECharge_7D": "Battery charge energy - Last 7 Days",
    "ECharge_30D": "Battery charge energy - Last 30 Days",
    "EDischarge_runtime": "Battery discharge energy - Since Restart",
    "EDischarge_7D": "Battery discharge energy - Last 7 Days",
    "EDischarge_30D": "Battery discharge energy - Last 30 Days",
    # ETotCharge/ETotDischarge series: Total for all batteries (from VSN700 feed title)
    "ETotCharge_runtime": "Total battery charge energy - Since Restart",
    "ETotCharge_7D": "Total battery charge energy - Last 7 Days",
    "ETotCharge_30D": "Total battery charge energy - Last 30 Days",
    "ETotDischarge_runtime": "Total battery discharge energy - Since Restart",
    "ETotDischarge_7D": "Total battery discharge energy - Last 7 Days",
    "ETotDischarge_30D": "Total battery discharge energy - Last 30 Days",
    # String current/power measurements
    "Iin1": "DC Current (String 1)",
    "Iin2": "DC Current (String 2)",
    "Pin": "Total DC Input Power",
    "Pin1": "DC Input Power (String 1)",
    "Pin2": "DC Input Power (String 2)",
    # Grid monitoring and support
    "HzA": "AC Frequency Phase A",
    "HzB": "AC Frequency Phase B",
    "HzC": "AC Frequency Phase C",
    "Fcc": "Central controller frequency",
    # Leakage and safety monitoring
    "IleakDC": "DC side ground fault leakage current",
    "IleakInv": "Inverter ground fault leakage current",
    "Iba": "Battery DC current flow",
    # Inverter status and power
    "InvState": "Current operational state of inverter",
    "Ppeak": "Maximum instantaneous power",
    "WRtg": "Inverter rated power capacity",
    # System measurements
    "SysTime": "Inverter system timestamp",
    "Vgnd": "Voltage to ground reference",
    # Alarm and status points
    "Alarm St": "Current alarm status code",
    "Global St": "Overall inverter operating state",
    "D I Status": "Digital input status register",
    "Warning Flags": "Active warning condition flags",
    "P A C Derating Flags": "AC power derating active condition flags",
    "Dc St1": "DC string 1 connection status",
    "Dc St2": "DC string 2 connection status",
    "D C1 State": "DC string 1 operating state",
    "D C2 State": "DC string 2 operating state",
    # Temperature points (additional)
    "Temp Inv": "Inverter internal temperature",
    "Temp Bst": "DC-DC boost converter temperature",
    "Tba": "Battery pack temperature",
    # Voltage measurements (additional)
    "Vbulk": "DC bus bulk capacitor voltage",
    "Vbulk_mid": "DC bus bulk capacitor midpoint voltage",
    "VoutRS": "AC Voltage Phase A-B",
    "VoutST": "AC Voltage Phase B-C",
    "VoutTR": "AC Voltage Phase C-A",
    "Vgrid L1 N": "AC Voltage Phase A-N",
    "Vgrid L2 N": "AC Voltage Phase B-N",
    "Vin1": "DC string 1 input voltage",
    "Vin2": "DC string 2 input voltage",
    "Vba": "Battery pack voltage",
    "Vc Max": "Maximum voltage across battery cells",
    "Vc Min": "Minimum voltage across battery cells",
    # Power measurements (additional)
    "Pba": "Battery charge/discharge power",
    "P Ch": "Battery charging power",
    "P Dh": "Battery discharging power",
    "Pac Stand Alone": "AC power in stand-alone (off-grid) mode",
    "Pac Togrid": "AC power exported to grid",
    "Pn": "Inverter nominal rated power",
    "Sn": "Inverter nominal apparent power rating",
    "Cos phi": "AC power factor (cosine phi)",
    # Energy counters (additional)
    "E Total Absorbed": "Cumulative energy absorbed from grid",
    "E Total Apparent": "Cumulative apparent energy (kVAh)",
    # Fan and cooling
    "Fan1Rpm": "Cooling fan 1 speed in RPM",
    "Fan2Rpm": "Cooling fan 2 speed in RPM",
    # Battery
    "T Soc": "Battery state of charge percentage",
    # Grid control
    "Grid Ext Ctrl Ena": "External grid control enabled status",
    "Grid Ext Ctrl State": "External grid control current state",
    # Status/Alarm Points - Fixed keys to match SunSpec Normalized Names
    "AlarmSt": "Current alarm status code",
    "DcSt1": "DC string 1 connection status",
    "DcSt2": "DC string 2 connection status",
    "DI_status": "Digital input status register",
    "GlobalSt": "Global inverter status",
    "WarningFlags": "Active warning condition flags",
    "PACDeratingFlags": "AC power derating active condition flags",
    "DC1State": "DC string 1 operating state",
    "DC2State": "DC string 2 operating state",
    # Meter/Grid Points - Using actual SunSpec names
    "W": "AC Power",
    "A": "AC Current",
    "VAr": "Reactive power at grid connection point",
    "PF": "Power factor",
    "DCW": "DC Power",
    "DCW_1": "DC Power (String 1)",
    "DCW_2": "DC Power (String 2)",
    "DCA_1": "DC Current (String 1)",
    "DCA_2": "DC Current (String 2)",
    "Hz": "Grid frequency",
    "PhVphA": "AC Voltage Phase A-N",
    "VgridL1_N": "AC Voltage Phase A-N",
    "VgridL2_N": "AC Voltage Phase B-N",
    # Voltage Points - Using actual SunSpec names
    "PhVphAB": "AC Voltage Phase A-B",
    "PhVphBC": "AC Voltage Phase B-C",
    "PhVphCA": "AC Voltage Phase C-A",
    "VGnd": "Voltage to ground reference",
    "DCV_1": "DC Voltage (String 1)",
    "DCV_2": "DC Voltage (String 2)",
    "VcMax": "Maximum voltage across battery cells",
    "VcMin": "Minimum voltage across battery cells",
    # Power Points - Using actual SunSpec names
    "PacStandAlone": "AC power in stand-alone (off-grid) mode",
    "PacTogrid": "AC power exported to grid",
    # Note: sn/pn in datalogger context are identification fields, not power
    "pn": "Product number",
    "sn": "Serial number",
    # Control Points - Using actual SunSpec names
    "gridExtCtrlEna": "External grid control enabled status",
    "gridExtCtrlState": "External grid control current state",
    # WiFi/Network - Using actual SunSpec names
    "wlan0_essid": "WiFi network name (SSID)",
    "wlan0_link_quality": "WiFi link quality percentage",
    # M1 Common Model - Using actual SunSpec names
    "Mn": "Manufacturer name",
    "Md": "Model identifier",
    "Vr": "Firmware version",
    "Opt": "Configuration options",
    "SN": "Device serial number",
    "DA": "Device Modbus address",
    # Temperature/Other Points - Using actual SunSpec names
    "TempInv": "Inverter internal temperature",
    "TempBst": "DC-DC boost converter temperature",
    "TSoc": "Battery state of charge percentage",
    "Temp1": "Additional temperature sensor 1",
    "ShU": "Current shunt voltage",
    "cosPhi": "Power factor (cosine phi)",
    # Other measurements
    "Isolation_Ohm1": "DC insulation resistance to ground",
    "Sh U": "Current shunt voltage",
    "Type": "Inverter device type identifier",
    # ==========================================
    # v2.0.7 MASSIVE Systematic Improvements (58 points)
    # Fixes: WiFi/Network, Meter, Device/Status, Battery/Power
    # ==========================================
    # WiFi/Network (11 points)
    "wlan_0_broadcast": "WiFi broadcast address",
    "wlan_0_dhcpState": "WiFi DHCP state (acquired/pending)",
    "wlan_0_dns": "WiFi DNS server address",
    "wlan_0_gateway": "WiFi gateway IP address",
    "wlan_0_ipaddr": "WiFi IPv4 address",
    "wlan_0_netmask": "WiFi network subnet mask",
    "wlan_0_status": "WiFi connection status",
    "wlan_ap_status": "WiFi access point mode status",
    "wlan_fsm_status": "WiFi finite state machine status",
    "wlan0_ipaddr_local": "WiFi local IP address",
    "wlan0_mode": "WiFi operating mode (client/AP)",
    # Meter (14 points)
    "MeterPgrid_L1": "Meter Power Phase A",
    "MeterPgrid_L2": "Meter Power Phase B",
    "MeterPgrid_L3": "Meter Power Phase C",
    "MeterPgrid_Tot": "Meter active power - Total",
    "MeterQgrid_L1": "Meter Reactive Power Phase A",
    "MeterQgrid_L2": "Meter Reactive Power Phase B",
    "MeterQgrid_L3": "Meter Reactive Power Phase C",
    "MeterQgrid_Tot": "Meter reactive power - Total",
    "HousePgrid_Tot": "House load active power - Total",
    "PbaT": "Battery power - Total",
    "MeterVgrid_L1": "Meter Voltage Phase A",
    "MeterVgrid_L2": "Meter Voltage Phase B",
    "MeterVgrid_L3": "Meter Voltage Phase C",
    "MeterIgrid_L1": "Meter Current Phase A",
    "MeterIgrid_L2": "Meter Current Phase B",
    "MeterIgrid_L3": "Meter Current Phase C",
    "MeterFgrid": "Meter AC frequency",
    # Device/Firmware/Manufacturing (16 points)
    "device_0_devName": "Device 0 component name",
    "device_1_devName": "Device 1 component name",
    "device_2_devName": "Device 2 component name",
    "device_0_devType": "Device 0 component type",
    "device_1_devType": "Device 1 component type",
    "device_2_devType": "Device 2 component type",
    "device_0_fwVer": "Firmware version",
    "device_1_fwVer": "Firmware version",
    "device_2_fwVer": "Firmware version",
    "fw_build_date": "Firmware build date and time",
    "fw_build_number": "Firmware build commit hash",
    "fw_release_number": "Firmware release version number",
    "fw_revision": "Firmware revision identifier",
    "mfg_part_number": "Manufacturing part number",
    "mfg_serial_number": "Manufacturing serial number",
    "mfg_week_year": "Manufacturing week and year (WWYY)",
    # Battery/Power/Control (17 points)
    "BattExtCtrlEna": "Battery external control enabled flag",
    "BattExtCtrlState": "Battery external control state",
    "CycleNum": "Battery charge/discharge cycle count",
    "PCh": "Battery charging power",
    "PDh": "Battery discharging power",
    "Fan1rpm": "Cooling fan 1 speed in RPM",
    "Fan2rpm": "Cooling fan 2 speed in RPM",
    "NumOfMPPT": "Number of MPPT DC input channels",
    "QACDeratingFlags": "Reactive power AC derating flags",
    "SACDeratingFlags": "Apparent power AC derating flags",
    "VBulkMid": "DC bus bulk capacitor midpoint voltage",
    "ILeakDcAc": "DC-to-AC ground leakage current",
    "ILeakDcDc": "DC-to-DC ground leakage current",
    "Isolation_Ohm1": "DC insulation resistance to ground",
    "m126Mod_Ena": "SunSpec Model 126 module enabled",
    "m132Mod_Ena": "SunSpec Model 132 module enabled",
    # VSN-specific points (user-requested fixes)
    "InverterSt": "Inverter status",
    "free_ram": "Available RAM",
    "flash_free": "Available flash storage",
    "logger_sn": "Logger serial number",
    "logger_loggerId": "Logger ID",
    # Additional improvements (v2.0.10)
    "GlobState": "Global operational state",
    "PowerPeakAbs": "Peak Output Power - Lifetime",
    "PowerPeakToday": "Peak Output Power - Today",
    "VBulk": "DC Voltage (Bulk Capacitor)",
    "sys_load": "System load average",
    "store_size": "Flash storage used",
    "ECt": "Total energy from current transformer (CT)",
    "EDt": "Total energy from direct transducer (DT)",
    "FRT_state": "Fault Ride Through (FRT) status",
}

# ==============================================================================
# CHANGELOG ENTRIES
# ==============================================================================

CHANGELOG_ENTRIES = [
    {
        "date": "2024-11-02",
        "version": "1.0.0",
        "type": "Initial Generation",
        "description": "Created comprehensive mapping with all customizations built-in",
        "details": [
            "Generated 210 deduplicated points (from 277 raw)",
            "Applied model flag detection from point names",
            "Implemented 4-tier description priority system",
            "Added comprehensive category assignment logic",
        ],
        "source": "Script automation",
    },
    {
        "date": "2024-11-02",
        "version": "1.0.0",
        "type": "Display Name Updates",
        "description": "Applied 20 display name corrections for better clarity",
        "details": [
            "Datalogger: Serial Number, IP address, WiFi SSID, Wlan 0 mode",
            "Inverter: Alarm status, Booster temperature, DC capacitor/mid-point voltage",
            "MPPT: DC current/voltage #1/#2 (instead of 'for string X')",
            "Device info: Model, Options, Serial Number, Firmware Version, Manufacturer",
            "Grid: Phase Voltage AN (instead of 'Phase A to neutral voltage measurement')",
        ],
        "source": "User request",
    },
    {
        "date": "2024-11-02",
        "version": "1.0.0",
        "type": "Data Fixes",
        "description": "Fixed incorrect device_class and unit assignments",
        "details": [
            "Device Address: Removed device_class=current and unit=A",
            "Firmware Version: Removed device_class=voltage and unit=V",
            "System timestamp: Set device_class=timestamp with entity_category=diagnostic",
        ],
        "source": "User request",
    },
    {
        "date": "2024-11-02",
        "version": "1.0.0",
        "type": "Description Improvements",
        "description": "Applied 108 description enhancements",
        "details": [
            "M1 Common Model: Enhanced descriptions for Mn, Md, Opt, Vr, SN, DA",
            "Phase measurements: Clarified voltage/current descriptions",
            "MPPT: Improved DC measurement descriptions",
            "Battery: Added comprehensive battery point descriptions",
            "Energy counters: Standardized period descriptions (lifetime, 7D, 30D, 1Y)",
            "House meter: Clarified consumption measurements",
            "System info: Added descriptive text for configuration points",
        ],
        "source": "Automated logic improvement",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.1",
        "type": "VSN300 Parser Fix",
        "description": "Fixed C_ prefix parsing for VSN300 Model 1 (Common) points",
        "details": [
            "Added special handler for C_ prefix points to map to Model 1",
            "Fixed 6 points: C_Mn, C_Md, C_SN, C_Vr, C_Opt, C_DA",
            "Updated main processing loop to include VSN300-only points in mapping",
            "Improved VSN300 name detection for points in VSN_TO_SUNSPEC_MAP",
            "These points now get proper SunSpec labels like 'Manufacturer', 'Serial Number', etc.",
        ],
        "source": "User-reported issue fix",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.2",
        "type": "SunSpec Parser Improvements",
        "description": "Major improvements to SunSpec workbook parser with updated data file",
        "details": [
            "Fixed column mapping to use correct columns: C=Name, M=Label, N=Description",
            "Expanded from reading 12 hardcoded tabs to all 112 numeric model sheets",
            "Updated SunSpec workbook file with complete M160 descriptions",
            "M160 MPPT points now have proper descriptions (DC Current, DC Voltage, DC Power)",
            "Improved coverage from ~10% to 100% of available SunSpec models",
            "Maintained ABB Excel usage only for M64061 proprietary model",
        ],
        "source": "User-reported parser issue",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.3",
        "type": "VSN Point Description Improvements",
        "description": "Improved labels and descriptions for 29 VSN-specific points",
        "details": [
            "Temperature points: TcMax/TcMin now show 'Cabinet Temperature Maximum/Minimum'",
            "Digital inputs: DI0_mode/DI1_mode clarified as 'Digital Input X Mode'",
            "Energy counters: EBackup/ECt/EDt/Ein1/Ein2 with clear energy source descriptions",
            "String measurements: Iin1/2, Pin1/2 now indicate 'String X Input Current/Power'",
            "Grid monitoring: FRT_state explained as 'Fault Ride Through Status'",
            "Phase frequencies: HzA/B/C labeled as 'Phase X Frequency'",
            "Leakage currents: IleakDC/IleakInv with ground fault context",
            "System measurements: Improved Vgnd, SysTime, cosPhi descriptions",
            "Added 29 new display name corrections for better UI presentation",
        ],
        "source": "User feedback on generic labels",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.4",
        "type": "Major Point Description Overhaul",
        "description": "Fixed M1 generic descriptions and improved 48 additional VSN points",
        "details": [
            "Fixed M1 Common Model: Replaced generic 'Manufacturer specific value' with meaningful descriptions",
            "Status/Alarms (9 points): Alarm St, Global St, DC states, Warning Flags with clear explanations",
            "Temperature (4 points): Temp Inv/Bst/1, Tba with proper component identification",
            "Voltage (16 points): Vbulk, VoutRS/ST/TR, Vgrid L1/2-N, Vin1/2, Vba, Vc Max/Min clarified",
            "Power (8 points): Pba, P Ch/Dh, Pac Stand Alone/Togrid, Pn/Sn with operational context",
            "Energy (2 points): E Total Absorbed/Apparent with measurement type specified",
            "Fans (2 points): Fan1Rpm/Fan2Rpm labeled as 'Fan X Speed'",
            "Battery (1 point): T Soc as 'Battery State of Charge'",
            "Grid Control (2 points): Grid Ext Ctrl Ena/State with clear status descriptions",
            "Other (5 points): Riso, Sh U, Type with technical explanations",
            "Added 48 display name corrections for improved Home Assistant UI",
        ],
        "source": "User feedback on remaining generic/unclear descriptions",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.5",
        "type": "Comprehensive Point Improvements - Final Round (FAILED)",
        "description": "Attempted to fix 65 points but used incorrect SunSpec names (32% success rate)",
        "details": [
            "CRITICAL ISSUE: Used REST API names instead of SunSpec Normalized Names as dictionary keys",
            "71% of improvements failed to apply (46 out of 65 points)",
            "Wrong keys: Pgrid→should be W, Igrid→A, Qgrid→VAr, Vgrid→PhVphA, E_Ch→ECharge, E_Dh→EDischarge",
            "Capitalization errors: DIStatus→DI_status, Cosphi→cosPhi, Pn→pn, Sn→sn",
            "Missing underscores: VgridL1N→VgridL1_N, VgridL2N→VgridL2_N",
            "Non-existent points: St, Hz1, VgridL1, V, Vac, VacL, VAC, VoutAC, VacN, BatCtrlDCh, etc.",
            "Only 19 points actually improved successfully",
            "Version 2.0.6 contains comprehensive corrections",
        ],
        "source": "Post-release audit revealed systematic mapping errors",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.6",
        "type": "CORRECTED Point Improvements - v2.0.5 Bug Fixes",
        "description": "Fixed all 57 points using correct SunSpec Normalized Names as dictionary keys",
        "details": [
            "Status/Alarm (9): DI_status, AlarmSt, GlobalSt, WarningFlags, PACDeratingFlags, DcSt1/2, DC1/2State",
            "Meter/Grid (14): W, A, VAr, PF, Hz, PhVphA, VgridL1_N, VgridL2_N, ECharge, EDischarge, ETotalAbsorbed/Apparent",
            "Voltage (14): PhVphAB/BC/CA (line-to-line), VGnd, VBulk, DCV_1/2, VcMax/Min",
            "Power (4): PacStandAlone, PacTogrid, pn, sn (corrected to lowercase)",
            "Control (2): gridExtCtrlEna, gridExtCtrlState (corrected capitalization)",
            "WiFi/Network (2): wlan0_essid, wlan0_link_quality (was wlan0_quality)",
            "M1 Common (6): Mn, Md, Vr, Opt, SN, DA with proper descriptions",
            "Other (6): TempInv, TempBst, TSoc, Temp1, ShU, cosPhi",
            "KEY FIX: All dictionary keys now use SunSpec Normalized Names, not REST API names",
            "Fixed PhVphAB/BC/CA line-to-line voltages that were showing VoutRS/ST/TR descriptions",
            "Fixed VGnd (capital G) that was showing just 'Vgnd' as description",
            "Corrected all display names to match improved descriptions",
        ],
        "source": "User-requested audit and comprehensive correction of v2.0.5 failures",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.7",
        "type": "MASSIVE Systematic Improvements - 58 Additional Points",
        "description": "Fixed 189 points (73.3% of all points) with systematic issues after comprehensive audit",
        "details": [
            "WiFi/Network (11): Fixed 'Wlan 0' prefixes → 'WiFi', expanded abbreviations (DHCP, DNS, IP, FSM)",
            "Meter (14): Expanded all meter point descriptions (Pgrid→Power, Qgrid→Reactive Power, etc.)",
            "Device/Status (16): Fixed device_X, fw_X, mfg_X points with clear descriptions",
            "Battery/Power (17): BattExtCtrl, PCh/PDh, Fans, NumOfMPPT, Derating, Leakage, etc.",
            "SYSTEMATIC FIXES: Removed tech prefixes, expanded abbreviations, made user-friendly",
            "Issues fixed: 172 repeats, 87 bad displays, 65 short descriptions, 26 tech prefixes, 14 spacing",
            "Post-audit success rate increased from 32% to expected ~90%+",
            "All 61 user-specified problematic points now have proper descriptions and display names",
            "Moved DESCRIPTION_IMPROVEMENTS to Priority 1 to override feeds data with curated descriptions",
        ],
        "source": "User-requested comprehensive audit revealing 73% of points had issues",
    },
    {
        "date": "2024-11-02",
        "version": "2.0.8",
        "type": "Label Spacing Fix - Acronyms",
        "description": "Fixed horrible single-letter spacing in labels for acronyms",
        "details": [
            "Fixed 'F R T State' → 'FRT State' (Fault Ride Through)",
            "Fixed 'Num Of M P P T' → 'Num Of MPPT'",
            "Fixed 'Q A C Derating Flags' → 'QAC Derating Flags'",
            "Fixed 'S A C Derating Flags' → 'SAC Derating Flags'",
            "Fixed 'P A C Derating Flags' → 'PAC Derating Flags'",
            "Fixed 'D I Status' → 'DI Status'",
            "Fixed 'Day W H' → 'Day WH', 'Month W H' → 'Month WH', etc.",
            "Added acronym preprocessing in generate_label_from_name() to prevent letter-by-letter splitting",
            "Acronyms now preserved: FRT, MPPT, QAC, SAC, PAC, DI, WH",
        ],
        "source": "User feedback on horrible single-letter spacing in labels",
    },
    {
        "date": "2024-11-03",
        "version": "2.0.9",
        "type": "Comprehensive Data Quality Improvements & HA Metadata",
        "description": "Fixed case inconsistencies, removed prefixes, standardized time periods, corrected units, and added comprehensive HA metadata",
        "details": [
            "Label Corrections (2): Fixed 'D C1 State' → 'DC1 State', 'D C2 State' → 'DC2 State'",
            "Prefix Removal (12): Removed 'E# -' and 'Ein -' prefixes from descriptions and display names",
            "Time Period Standardization (30): Changed 'last 30 days'/'- 30D' → 'last month', 'last 7 days'/'- 7D' → 'last week'",
            "Unit Corrections (4): Fixed DayWH, MonthWH, WeekWH, YearWH from 'kWh' → 'Wh' to match naming",
            "Comprehensive HA Metadata (150+ points): Added device_class, state_class, and unit for energy, power, voltage, current, temperature, frequency, battery, etc.",
            "Created SUNSPEC_TO_HA_METADATA dictionary with proper Home Assistant metadata for all measurement types",
            "All corrections applied centrally in create_row_with_model_flags() for consistency across all points",
            "Expected result: ~95% of points now have proper HA metadata, cleaner descriptions, consistent time periods",
        ],
        "source": "User request for systematic data quality improvements before manual review",
    },
]

# ==============================================================================
# PART 2: HELPER FUNCTIONS
# ==============================================================================


def load_feeds_titles(vsn300_feeds_path, vsn700_feeds_path):
    """Load title and unit data from VSN300 and VSN700 feeds.json files."""

    def is_title_a_description(point_name, title):
        """Determine if a title is a description or just a point name reference."""
        if not title:
            return False
        if title.lower() == point_name.lower():
            return False
        if title in ["NO_DESCR", "N/A", "", "-"]:
            return False
        return True

    feeds_titles = {}

    # Process VSN300 feeds
    if vsn300_feeds_path.exists():
        with open(vsn300_feeds_path) as f:
            vsn300_feeds = json.load(f)
            # feeds is a dictionary, not a list!
            feeds_dict = vsn300_feeds.get("feeds", {})
            for feed_data in feeds_dict.values():
                datastreams = feed_data.get("datastreams", {})
                for point_name, point_info in datastreams.items():
                    title = point_info.get("title", "")
                    units = point_info.get("units", "")
                    is_description = is_title_a_description(point_name, title)
                    feeds_titles[point_name] = {
                        "title": title,
                        "units": units,
                        "is_description": is_description,
                        "source": "VSN300 feeds",
                    }

    # Process VSN700 feeds
    if vsn700_feeds_path.exists():
        with open(vsn700_feeds_path) as f:
            vsn700_feeds = json.load(f)
            # feeds is a dictionary, not a list!
            feeds_dict = vsn700_feeds.get("feeds", {})
            for feed_data in feeds_dict.values():
                datastreams = feed_data.get("datastreams", {})
                for point_name, point_info in datastreams.items():
                    title = point_info.get("title", "")
                    units = point_info.get("units", "")
                    is_description = is_title_a_description(point_name, title)
                    if point_name not in feeds_titles:
                        feeds_titles[point_name] = {
                            "title": title,
                            "units": units,
                            "is_description": is_description,
                            "source": "VSN700 feeds",
                        }

    return feeds_titles


def load_vsn_status(vsn300_status_path, vsn700_status_path):
    """Load status.json data from VSN300 and VSN700."""
    status_points = {}

    # Process VSN300 status
    if vsn300_status_path.exists():
        with open(vsn300_status_path) as f:
            vsn300_status = json.load(f)
            if "keys" in vsn300_status:
                for key, info in vsn300_status["keys"].items():
                    # Extract clean point name (e.g., "fw.release_number" -> "fw_release_number")
                    point_name = key.replace(".", "_")
                    label = info.get("label", "")
                    value = info.get("value", "")

                    status_points[point_name] = {
                        "source": "VSN300 status",
                        "label": label
                        if label
                        else generate_label_from_name(point_name),
                        "value_example": value,
                        "vsn300": True,
                        "vsn700": False,
                    }

    # Process VSN700 status
    if vsn700_status_path.exists():
        with open(vsn700_status_path) as f:
            vsn700_status = json.load(f)
            if "keys" in vsn700_status:
                for key, info in vsn700_status["keys"].items():
                    # Extract clean point name
                    point_name = key.replace(".", "_")
                    label = info.get("label", "") if isinstance(info, dict) else ""
                    value = info.get("value", "") if isinstance(info, dict) else ""

                    if point_name in status_points:
                        status_points[point_name]["vsn700"] = True
                    else:
                        status_points[point_name] = {
                            "source": "VSN700 status",
                            "label": label
                            if label
                            else generate_label_from_name(point_name),
                            "value_example": value,
                            "vsn300": False,
                            "vsn700": True,
                        }

    return status_points


def load_m64061_from_abb_excel(excel_path):
    """Load M64061 model data from ABB_SunSpec_Modbus.xlsx."""
    import openpyxl

    m64061_points = {}

    if not excel_path.exists():
        print(f"  WARNING: ABB SunSpec Excel not found at {excel_path}")
        return m64061_points

    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # The M64061 data is in the "Inverter Modbus" sheet starting at row 96
    if "Inverter Modbus" in wb.sheetnames:
        ws = wb["Inverter Modbus"]

        # M64061 starts at row 95 with the header "ABB Inverter Vendor Model (64061)"
        # Actual data points start at row 96
        # Columns appear to be: A-D (addresses), E-F (unknown), G (Size), H (Access), I (Name)

        in_m64061_section = False
        for row in range(95, ws.max_row + 1):
            # Check if we're at the M64061 header
            cell_a = ws.cell(row, 1).value
            if cell_a and "64061" in str(cell_a):
                in_m64061_section = True
                continue

            # If we're in the M64061 section
            if in_m64061_section:
                # Check if we've reached another model section
                if cell_a and "Model" in str(cell_a) and "64061" not in str(cell_a):
                    break

                # Extract point data from columns
                name = ws.cell(row, 9).value  # Column I - Name
                if name and name != "Name" and not str(name).startswith("="):
                    # Get other data if available
                    size = ws.cell(row, 7).value  # Column G - Size
                    access = ws.cell(row, 8).value  # Column H - Access (R/RW)

                    # Some known M64061 points from the original script
                    if name in [
                        "Alm1",
                        "Alm2",
                        "Alm3",
                        "AlarmState",
                        "FaultStatus",
                        "BatteryMode",
                        "BatteryStatus",
                        "IsolResist",
                    ]:
                        m64061_points[name] = {
                            "size": size,
                            "access": access,
                            "label": generate_label_from_name(name),
                            "description": f"M64061 proprietary point: {name}",
                        }

    wb.close()
    return m64061_points


def load_sunspec_models_metadata(workbook_path):
    """Load label and description from SunSpec models workbook."""
    import openpyxl

    models_data = {}
    wb = openpyxl.load_workbook(workbook_path, read_only=True)

    # Get all numeric sheet names (model numbers)
    model_sheets = []
    for sheet_name in wb.sheetnames:
        # Include numeric sheets (model numbers) but exclude "Index" and other non-model sheets
        if sheet_name.isdigit():
            model_sheets.append(sheet_name)

    print(f"  Found {len(model_sheets)} model sheets to process")

    for sheet_name in model_sheets:
        ws = wb[sheet_name]
        model_key = f"M{sheet_name}"
        models_data[model_key] = {}

        # SunSpec workbook standard column mapping:
        # Column C (3) = Name
        # Column M (13) = Label
        # Column N (14) = Description
        name_col = 3  # Column C
        label_col = 13  # Column M
        desc_col = 14  # Column N

        # Row 1 contains headers, data starts at row 2
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row_idx, name_col).value
            if not name:
                continue

            label = ws.cell(row_idx, label_col).value
            description = ws.cell(row_idx, desc_col).value

            models_data[model_key][name] = {
                "label": label or "",
                "description": description or "",
            }

    wb.close()
    return models_data


def clean_energy_prefix(text):
    """Remove 'E# -' and 'Ein -' prefixes from descriptions and display names (v2.0.9)."""
    if not text:
        return text

    # Pattern: E0 -, E1 -, E2 -, ..., E99 -, Ein -
    import re

    text = re.sub(r"^E\d+ - ", "", text)
    return re.sub(r"^Ein - ", "", text)


def standardize_time_periods(text):
    """Standardize time period references (v2.0.9).

    Replace:
    - 'last 30 days' → 'last month'
    - '- 30D' → '- last month'
    - 'last 7 days' → 'last week'
    - '- 7D' → '- last week'
    """
    if not text:
        return text

    # Replace variations of 30 days
    text = text.replace("last 30 days", "last month")
    text = text.replace("- 30D", "- last month")
    text = text.replace(" 30D", " last month")

    # Replace variations of 7 days
    text = text.replace("last 7 days", "last week")
    text = text.replace("- 7D", "- last week")
    return text.replace(" 7D", " last week")


def apply_label_corrections(label):
    """Apply label corrections from LABEL_CORRECTIONS dictionary (v2.0.9)."""
    if label in LABEL_CORRECTIONS:
        return LABEL_CORRECTIONS[label]
    return label


def generate_label_from_name(point_name):
    """Generate a human-readable label from a point name."""
    # Handle VSN300 SunSpec patterns
    if "_" in point_name and point_name.startswith("m"):
        match = re.match(r"m(\d+)_(\d+)_(.+)", point_name)
        if match:
            model_num, instance, point = match.groups()
            base_label = generate_label_from_name(point)
            if "_" in point and point.split("_")[-1].isdigit():
                return base_label
            return base_label

    # Handle M64061 state codes
    state_codes = {
        "Alm1": "Alarm 1",
        "Alm2": "Alarm 2",
        "Alm3": "Alarm 3",
        "AlarmState": "Alarm State",
        "FaultStatus": "Fault Status",
        "BatteryMode": "Battery Mode",
        "BatteryStatus": "Battery Status",
        "IsolResist": "Isolation Resistance",
    }
    if point_name in state_codes:
        return state_codes[point_name]

    # Handle system monitoring points
    system_points = {
        "flash_free": "Flash Memory Free",
        "free_ram": "Free RAM",
        "fw_ver": "Firmware Version",
        "hw_ver": "Hardware Version",
        "store_size": "Storage Size",
        "sys_load": "System Load",
        "uptime": "System Uptime",
        "Sn": "Serial Number",
    }
    if point_name in system_points:
        return system_points[point_name]

    # Handle periodic energy counters
    if "_" in point_name:
        base, suffix = point_name.rsplit("_", 1)
        period_map = {
            "runtime": "Lifetime",
            "7D": "Last Week",  # v2.0.9: Changed from "7 Day"
            "30D": "Last Month",  # v2.0.9: Changed from "30 Day"
            "1Y": "1 Year",
        }
        if suffix in period_map:
            base_label = generate_label_from_name(base)
            return f"{base_label} {period_map[suffix]}"

    # Handle common abbreviations
    abbreviations = {
        "Soc": "State of Charge",
        "Soh": "State of Health",
        "Tmp": "Temperature",
        "Vbat": "Battery Voltage",
        "Ibat": "Battery Current",
        "Pbat": "Battery Power",
    }
    for abbr, full in abbreviations.items():
        if point_name.startswith(abbr):
            return full + point_name[len(abbr) :]

    # Handle multi-letter acronyms BEFORE splitting (v2.0.7 fix for spacing)
    # Replace acronyms with a placeholder to prevent letter-by-letter splitting
    acronym_map = {
        "FRT": "FaultRideThrough",
        "MPPT": "Mppt",
        "QAC": "Qac",
        "SAC": "Sac",
        "PAC": "Pac",
        "DI": "DigitalInput",
        "WH": "Wh",
    }
    temp_name = point_name
    for acronym, replacement in acronym_map.items():
        temp_name = temp_name.replace(acronym, replacement)

    # Split CamelCase and snake_case
    words = re.sub(r"([A-Z])", r" \1", temp_name).split()
    words = [w for word in words for w in word.split("_") if w]

    result = " ".join(words).title()

    # Post-process to fix acronym labels
    result = result.replace("Fault Ride Through", "FRT")
    result = result.replace("Mppt", "MPPT")
    result = result.replace("Qac", "QAC")
    result = result.replace("Sac", "SAC")
    result = result.replace("Pac", "PAC")
    return result.replace("Digital Input", "DI")


def generate_simplified_point_name(label, model):
    """Generate simplified HA entity name from label."""
    # Convert to lowercase and replace spaces with underscores
    name = label.lower().replace(" ", "_")

    # Remove special characters
    name = re.sub(r"[^a-z0-9_]", "", name)

    # Replace multiple underscores with single
    name = re.sub(r"_+", "_", name)

    # Strip leading/trailing underscores
    return name.strip("_")


def lookup_label_description(models_data, model, sunspec_point):
    """Lookup label and description from SunSpec metadata."""
    label = ""
    description = ""

    if model and f"M{model.replace('M', '')}" in models_data:
        model_data = models_data[f"M{model.replace('M', '')}"]

        # Handle repeating groups (e.g., DCA_1 → DCA)
        base_point = sunspec_point
        if "_" in sunspec_point and sunspec_point.split("_")[-1].isdigit():
            base_point = "_".join(sunspec_point.split("_")[:-1])
            suffix = sunspec_point.split("_")[-1]
        else:
            suffix = None

        if base_point in model_data:
            label = model_data[base_point].get("label", "")
            description = model_data[base_point].get("description", "")
            if suffix:
                label = f"{label} #{suffix}"
                description = f"{description} for string {suffix}"

    return label, description


def get_description_with_priority(
    vsn_name,
    sunspec_name,
    model,
    label,
    workbook_description,
    feeds_info,
    vsn700_lookup=None,
    vsn300_name=None,
):
    """Get description using 5-tier priority system with data source tracking.

    Priority order (highest to lowest):
    0. VSN700 cross-reference for VSN300-only points
    1. DESCRIPTION_IMPROVEMENTS (our curated improvements) - HIGHEST PRIORITY
    2. SunSpec workbook description (except M1 boilerplate)
    3. Feeds title if it's a description
    4. Label fallback
    """
    # Priority 0: Cross-reference VSN700 lookup for VSN300-only points
    if vsn700_lookup and vsn300_name and vsn300_name == vsn_name:
        vsn700_equivalent = vsn700_lookup.get(vsn_name, {})
        if vsn700_equivalent.get("description"):
            return (
                vsn700_equivalent["description"],
                "VSN700 Cross-Reference",
                vsn700_equivalent.get("label"),
            )

    # Priority 1: DESCRIPTION_IMPROVEMENTS - our curated improvements take precedence
    if sunspec_name in DESCRIPTION_IMPROVEMENTS:
        return DESCRIPTION_IMPROVEMENTS[sunspec_name], "Curated Improvement", None

    # Priority 2: SunSpec workbook description (except M1 boilerplate)
    if workbook_description and workbook_description not in ["", "N/A", None]:
        if model != "M1" or "common model" not in workbook_description.lower():
            return workbook_description, "SunSpec Description", None

    # Priority 3: Feeds title if it's a description
    if feeds_info and feeds_info.get("is_description"):
        return (
            feeds_info["title"],
            f"VSN Feeds ({feeds_info.get('source', 'Unknown')})",
            None,
        )

    # Priority 4: Label fallback
    if label:
        return label, "Generated from Label", None

    return "Unknown measurement", "Fallback", None


def get_comprehensive_category(
    label, description, model_flags, vsn300_name, vsn700_name
):
    """Determine category with comprehensive logic."""
    label_lower = (label or "").lower()
    desc_lower = (description or "").lower()

    # Energy Counter (highest priority)
    if any(
        pattern in label_lower
        for pattern in ["e0", "e1", "e2", "e3", "e4", "e5", "e6", "e7", "e8"]
    ):
        if not any(x in label_lower for x in ["charge", "discharge"]):
            return "Energy Counter"
    if "energy" in label_lower or "energy" in desc_lower:
        if "battery" not in label_lower and "battery" not in desc_lower:
            return "Energy Counter"

    # House Meter
    if "house" in label_lower or "house" in desc_lower:
        return "House Meter"
    if vsn700_name and vsn700_name.startswith("House"):
        return "House Meter"

    # System Monitoring
    system_keywords = [
        "flash",
        "ram",
        "uptime",
        "load",
        "sys_",
        "wlan",
        "ip",
        "essid",
        "mode",
    ]
    if any(kw in label_lower for kw in system_keywords):
        return "System Monitoring"

    # Battery (excluding energy counters)
    battery_keywords = ["battery", "batt", "cell", "soc", "soh", "charge", "discharge"]
    if any(kw in label_lower for kw in battery_keywords):
        if "energy" not in label_lower:
            return "Battery"

    # Status
    status_keywords = ["status", "state", "alarm", "fault", "event", "mode", "control"]
    if any(kw in label_lower for kw in status_keywords):
        return "Status"

    # Device Info
    if model_flags.get("M1") == "YES":
        return "Device Info"

    # Phase-based categorization
    if any(
        phase in label_lower
        for phase in ["phase", "l1", "l2", "l3", "phv", "aph", "wph"]
    ):
        if model_flags.get("M103") == "YES" or model_flags.get("M160") == "YES":
            return "Inverter"
        if model_flags.get("M203") == "YES":
            return "Meter"

    # DC measurements
    if label_lower.startswith("dc") or "mppt" in label_lower:
        return "Inverter"

    # Model-based fallback
    if model_flags.get("M103") == "YES" or model_flags.get("M160") == "YES":
        return "Inverter"
    if model_flags.get("M203") == "YES":
        return "Meter"
    if model_flags.get("M802") == "YES":
        return "Battery"
    if model_flags.get("M64061") == "YES":
        return "Inverter"
    if model_flags.get("ABB_Proprietary") == "YES":
        return "System"
    if (
        model_flags.get("VSN300_Only") == "YES"
        or model_flags.get("VSN700_Only") == "YES"
    ):
        return "Datalogger"

    return "Unknown"


# ==============================================================================
# PART 3: DEDUPLICATION AND PROCESSING LOGIC
# ==============================================================================


def detect_models_from_point(vsn300_name, vsn700_name, model_hint=None):
    """Detect which models a point belongs to based on its names and hints."""
    models = set()

    # From VSN300 name pattern
    if vsn300_name and vsn300_name.startswith("m"):
        match = re.match(r"m(\d+)_", vsn300_name)
        if match:
            model_num = match.group(1)
            models.add(f"M{model_num}")

    # From VSN700 mapping
    if vsn700_name in VSN_TO_SUNSPEC_MAP:
        model = VSN_TO_SUNSPEC_MAP[vsn700_name].get("model")
        if model:
            models.add(model)

    # From M64061 points
    if vsn700_name in M64061_POINTS or vsn300_name in M64061_POINTS:
        models.add("M64061")

    # From ABB Proprietary
    if vsn700_name in ABB_PROPRIETARY or vsn300_name in ABB_PROPRIETARY:
        models.add("ABB_Proprietary")

    # From model hint
    if model_hint:
        models.add(model_hint.replace("ABB Proprietary", "ABB_Proprietary"))

    # VSN-only detection
    if not models:
        if vsn300_name and not vsn700_name:
            models.add("VSN300_Only")
        elif vsn700_name and not vsn300_name:
            models.add("VSN700_Only")

    return models


def normalize_vsn700_name(vsn700_name):
    """Apply VSN700 to VSN300 name normalization for duplicate detection.

    Args:
        vsn700_name: VSN700 REST API point name

    Returns:
        Normalized name that will be used as sunspec_name for deduplication

    """
    if vsn700_name in VSN_NAME_NORMALIZATION:
        return VSN_NAME_NORMALIZATION[vsn700_name]
    return vsn700_name


def merge_duplicate_rows(rows_by_sunspec):
    """Merge duplicate rows with the same SunSpec name."""
    merged_rows = []

    for group in rows_by_sunspec.values():
        if len(group) == 1:
            merged_rows.append(group[0])
            continue

        # Start with first row as base
        merged = dict(group[0])

        # Collect all models from all duplicates
        all_models = set()
        for row in group:
            models = detect_models_from_point(
                row.get("vsn300_name"), row.get("vsn700_name"), row.get("model")
            )
            all_models.update(models)

        # Update model flags
        for model_flag in MODEL_FLAGS:
            merged[model_flag] = "YES" if model_flag in all_models else "NO"

        # Merge VSN names (keep both if different)
        vsn700_names = [
            r.get("vsn700_name")
            for r in group
            if r.get("vsn700_name") and r.get("vsn700_name") != "N/A"
        ]
        vsn300_names = [
            r.get("vsn300_name")
            for r in group
            if r.get("vsn300_name") and r.get("vsn300_name") != "N/A"
        ]

        merged["vsn700_name"] = vsn700_names[0] if vsn700_names else "N/A"
        merged["vsn300_name"] = vsn300_names[0] if vsn300_names else "N/A"

        # Choose best description
        descriptions = [(r.get("description"), r.get("data_source")) for r in group]
        best_desc = sorted(
            descriptions,
            key=lambda x: (
                0
                if "SunSpec" in x[1]
                else 1
                if "VSN" in x[1]
                else 2
                if "Enhanced" in x[1]
                else 3
                if "Cross-Reference" in x[1]
                else 4
            ),
        )[0]
        merged["description"] = best_desc[0]
        merged["data_source"] = best_desc[1]

        merged_rows.append(merged)

    return merged_rows


def apply_display_name_corrections(row):
    """Apply display name corrections and standardization to a row.

    Applies in order:
    1. DISPLAY_NAME_CORRECTIONS - fixes/corrections from previous versions
    2. DISPLAY_NAME_STANDARDIZATION - TYPE-first pattern for grouping
    """
    original_display = row.get("ha_display_name", "")

    # Step 1: Apply corrections (legacy fixes)
    if original_display in DISPLAY_NAME_CORRECTIONS:
        row["ha_display_name"] = DISPLAY_NAME_CORRECTIONS[original_display]

    # Step 2: Apply standardization (v1.0.0-beta.22 - TYPE-first pattern)
    current_display = row.get("ha_display_name", "")
    if current_display in DISPLAY_NAME_STANDARDIZATION:
        row["ha_display_name"] = DISPLAY_NAME_STANDARDIZATION[current_display]

    return row


def apply_device_class_fixes(row):
    """Apply device class and unit fixes to a row."""
    label = row.get("label", "")

    if label in DEVICE_CLASS_FIXES:
        fixes = DEVICE_CLASS_FIXES[label]
        if "device_class" in fixes:
            row["device_class"] = fixes["device_class"] or ""
        if "unit" in fixes:
            row["units"] = fixes["unit"] or ""
        if "entity_category" in fixes:
            row["entity_category"] = fixes["entity_category"]

    return row


def _get_suggested_precision(sunspec_name, device_class, units, state_class, entity_category):
    """Determine suggested display precision for a sensor.

    Args:
        sunspec_name: SunSpec normalized name of the sensor
        device_class: Home Assistant device class
        units: Unit of measurement
        state_class: Home Assistant state class
        entity_category: Entity category (diagnostic, config, etc.)

    Returns:
        int or str: Precision value (0-2) or empty string for state sensors

    """
    # State sensors with state mappings should NOT have precision at all (they display text, not numbers)
    # These sensors have integer state codes mapped to human-readable strings in const.py
    # Return None so they don't get a precision field in the mapping
    STATE_SENSORS = {"GlobalSt", "DcSt1", "DcSt2", "InverterSt", "AlarmState", "AlarmSt"}
    if sunspec_name in STATE_SENSORS:
        return None  # No precision field - these are text-based sensors

    # Timestamp sensors that return formatted strings should NOT have precision
    # Raw values from API are integers (Aurora epoch timestamps), but sensor.py converts
    # them to formatted datetime strings (e.g., "2025-11-17 14:30:13")
    # Having precision makes HA think it's numeric, causing ValueError
    TIMESTAMP_SENSORS = {"SysTime"}
    if sunspec_name in TIMESTAMP_SENSORS:
        return None  # No precision field - these return formatted date strings

    # Special case: system_load gets 2 decimals despite no unit (float value 0.00-100.00)
    if sunspec_name == "sys_load":
        return 2

    # Unit-based precision (highest priority)
    if units:
        # Power, energy counters, apparent energy, duration, data: 0 decimals
        if units in ("W", "kW", "Wh", "kWh", "var", "VAR", "VAh", "kVAh", "s", "B"):
            return 0
        # Voltage, low current (mA), temperature, battery %, capacity, storage: 1 decimal
        if units in ("V", "mA", "%", "Ah", "°C", "°F", "MB", "GB"):
            return 1
        # High current (A), frequency, large resistance: 2 decimals
        if units in ("A", "Hz", "MOhm", "MΩ", "kΩ"):
            return 2
        # Small resistance, cycles, channels: 0 decimals
        if units in ("Ω", "Ohm", "cycles", "channels"):
            return 0

    # Device class fallback for sensors with empty units
    if device_class:
        # Power-related, duration, timestamps, battery percentage: 0 decimals
        if device_class in ("power", "reactive_power", "apparent_power", "duration", "timestamp", "battery"):
            return 0
        # Voltage and temperature: 1 decimal
        if device_class in ("voltage", "temperature"):
            return 1
        # Current, frequency, energy: 2 decimals
        if device_class in ("current", "frequency", "energy"):
            return 2

    # Default: 0 for unitless numeric sensors
    return 0


def create_row_with_model_flags(
    vsn700_name,
    vsn300_name,
    sunspec_name,
    ha_name,
    in_livedata,
    in_feeds,
    label,
    description,
    ha_display_name,
    category,
    units,
    state_class,
    device_class,
    entity_category,
    available_in_modbus,
    data_source,
    model_notes,
    models,
):
    """Create a complete row with model flags."""
    # Initialize model flags
    model_flags = dict.fromkeys(MODEL_FLAGS, "NO")
    for model in models:
        if model in model_flags:
            model_flags[model] = "YES"

    # Build complete row
    row = {
        **model_flags,
        "vsn700_name": vsn700_name,
        "vsn300_name": vsn300_name,
        "sunspec_name": sunspec_name,
        "ha_name": ha_name,
        "in_livedata": in_livedata,
        "in_feeds": in_feeds,
        "label": label,
        "description": description,
        "ha_display_name": ha_display_name,
        "category": category,
        "units": units,
        "state_class": state_class,
        "device_class": device_class,
        "entity_category": entity_category,
        "available_in_modbus": available_in_modbus,
        "data_source": data_source,
        "model_notes": model_notes,
        "icon": "",
    }

    # Apply corrections
    row = apply_display_name_corrections(row)

    # ==========================================
    # v2.0.9 CORRECTIONS
    # ==========================================

    # 1. Apply label corrections FIRST (D C1 State → DC1 State, Sn → Serial Number, etc.)
    # IMPORTANT: Must run before apply_device_class_fixes since DEVICE_CLASS_FIXES uses corrected label names
    row["label"] = apply_label_corrections(row["label"])

    # 2. Apply device class fixes (now that labels are corrected)
    row = apply_device_class_fixes(row)

    # 3. Clean E# and Ein prefixes from description and ha_display_name
    row["description"] = clean_energy_prefix(row["description"])
    row["ha_display_name"] = clean_energy_prefix(row["ha_display_name"])

    # 4. Standardize time periods (30 days → month, 7 days → week)
    row["description"] = standardize_time_periods(row["description"])
    row["ha_display_name"] = standardize_time_periods(row["ha_display_name"])

    # 5. Apply unit corrections (Wh vs kWh fixes)
    if sunspec_name in UNIT_CORRECTIONS:
        row["units"] = UNIT_CORRECTIONS[sunspec_name]

    # 6. Normalize temperature units (degC → °C)
    if row["units"] in TEMPERATURE_UNIT_FIX:
        row["units"] = TEMPERATURE_UNIT_FIX[row["units"]]

    # 7. Apply HA metadata from SUNSPEC_TO_HA_METADATA dictionary
    if sunspec_name in SUNSPEC_TO_HA_METADATA:
        metadata = SUNSPEC_TO_HA_METADATA[sunspec_name]
        # Only override if current values are empty/None
        if not row["device_class"] or row["device_class"] == "":
            row["device_class"] = metadata.get("device_class") or ""
        if not row["state_class"] or row["state_class"] == "":
            row["state_class"] = metadata.get("state_class") or ""
        # ALWAYS override units from SUNSPEC_TO_HA_METADATA (takes precedence over feeds data)
        # This ensures SunSpec-compliant units (W, Wh) are used instead of display units (kW, kWh)
        if metadata.get("unit"):
            row["units"] = metadata.get("unit")
        if not row["icon"] or row["icon"] == "":
            row["icon"] = metadata.get("icon") or ""
        # Apply entity_category if specified in metadata
        if not row["entity_category"] and metadata.get("entity_category"):
            row["entity_category"] = metadata.get("entity_category")

    # Update category if needed
    if row["category"] in ["Unknown", "", None]:
        row["category"] = get_comprehensive_category(
            row["label"], row["description"], model_flags, vsn300_name, vsn700_name
        )

    # Set diagnostic icon for all diagnostic entities (unless already has an icon)
    if row["entity_category"] == "diagnostic" and not row.get("icon"):
        row["icon"] = "mdi:information-box-outline"

    # 8. Set Suggested Display Precision based on metadata or auto-calculated
    # Check if precision was specified in SUNSPEC_TO_HA_METADATA first (takes priority)
    if sunspec_name in SUNSPEC_TO_HA_METADATA and "precision" in SUNSPEC_TO_HA_METADATA[sunspec_name]:
        row["precision"] = SUNSPEC_TO_HA_METADATA[sunspec_name]["precision"]
    else:
        # Auto-calculate precision based on units, device_class, etc.
        row["precision"] = _get_suggested_precision(
            sunspec_name=sunspec_name,
            device_class=row["device_class"],
            units=row["units"],
            state_class=row["state_class"],
            entity_category=row["entity_category"]
        )

    return row


# ==============================================================================
# PART 4: MAIN GENERATION FUNCTIONS
# ==============================================================================


def generate_changelog_sheet(wb):
    """Generate the Changelog sheet with all updates."""
    ws = wb.create_sheet("Changelog")

    # Headers
    headers = ["Date", "Version", "Type", "Description", "Details", "Source"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add changelog entries
    for entry in CHANGELOG_ENTRIES:
        details_str = "\n".join(f"• {detail}" for detail in entry["details"])
        ws.append(
            [
                entry["date"],
                entry["version"],
                entry["type"],
                entry["description"],
                details_str,
                entry["source"],
            ]
        )

    # Adjust column widths
    ws.column_dimensions["A"].width = 12  # Date
    ws.column_dimensions["B"].width = 10  # Version
    ws.column_dimensions["C"].width = 20  # Type
    ws.column_dimensions["D"].width = 50  # Description
    ws.column_dimensions["E"].width = 80  # Details
    ws.column_dimensions["F"].width = 20  # Source

    # Enable text wrapping for details
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def generate_summary_sheet(wb, rows):
    """Generate the Summary sheet with statistics."""
    ws = wb.create_sheet("Summary")

    # Calculate statistics
    total_points = len(rows)
    model_counts = defaultdict(int)
    category_counts = defaultdict(int)

    for row in rows:
        # Count models
        for model in MODEL_FLAGS:
            if row.get(model) == "YES":
                model_counts[model] += 1

        # Count categories
        category = row.get("category", "Unknown")
        category_counts[category] += 1

    # Add summary data
    ws.append(["Metric", "Count"])
    ws.append(["Total Points", total_points])
    ws.append([])

    ws.append(["Model Distribution", ""])
    for model in MODEL_FLAGS:
        ws.append([f"  {model}", model_counts.get(model, 0)])
    ws.append([])

    ws.append(["Category Distribution", ""])
    for category, count in sorted(
        category_counts.items(), key=lambda x: x[1], reverse=True
    ):
        ws.append([f"  {category}", count])

    # Style headers
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for cell in [ws.cell(row=1, column=1), ws.cell(row=1, column=2)]:
        cell.fill = header_fill
        cell.font = header_font

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


# ==============================================================================
# HELPER FUNCTIONS FOR generate_mapping_excel_complete()
# ==============================================================================


def _load_data_sources(data_dir):
    """Load all data sources including SunSpec metadata, M64061, feeds, and status."""
    print("Loading data sources...")

    # Load SunSpec metadata
    sunspec_path = data_dir / "sunspec" / "models_workbook.xlsx"
    if sunspec_path.exists():
        print(f"  Loading SunSpec models from {sunspec_path}")
        sunspec_metadata = load_sunspec_models_metadata(sunspec_path)
    else:
        print(f"  WARNING: SunSpec workbook not found at {sunspec_path}")
        sunspec_metadata = {}

    # Load M64061 from ABB Excel
    abb_excel_path = data_dir / "sunspec" / "ABB_SunSpec_Modbus.xlsx"
    print(f"  Loading M64061 model from {abb_excel_path}")
    m64061_metadata = load_m64061_from_abb_excel(abb_excel_path)
    print(f"    Found {len(m64061_metadata)} M64061 points")

    # Load feeds titles
    vsn300_feeds_path = data_dir / "vsn300" / "feeds.json"
    vsn700_feeds_path = data_dir / "vsn700" / "feeds.json"
    print("  Loading VSN feeds data...")
    feeds_titles = load_feeds_titles(vsn300_feeds_path, vsn700_feeds_path)

    # Load status data
    vsn300_status_path = data_dir / "vsn300" / "status.json"
    vsn700_status_path = data_dir / "vsn700" / "status.json"
    print("  Loading VSN status data...")
    status_points = load_vsn_status(vsn300_status_path, vsn700_status_path)
    print(f"    Found {len(status_points)} status points")

    return sunspec_metadata, m64061_metadata, feeds_titles, status_points


def _extract_point_sets(data_dir):
    """Extract point sets from VSN300 and VSN700 livedata and feeds files."""
    vsn700_livedata_path = data_dir / "vsn700" / "livedata.json"
    vsn700_feeds_path = data_dir / "vsn700" / "feeds.json"
    vsn300_livedata_path = data_dir / "vsn300" / "livedata.json"

    vsn700_livedata_points = set()
    vsn700_feeds_points = set()
    vsn300_livedata_points = set()

    if vsn700_livedata_path.exists():
        with open(vsn700_livedata_path) as f:
            data = json.load(f)
            for device_data in data.values():
                if isinstance(device_data, dict) and "points" in device_data:
                    for point in device_data["points"]:
                        vsn700_livedata_points.add(point["name"])

    if vsn700_feeds_path.exists():
        with open(vsn700_feeds_path) as f:
            data = json.load(f)
            feeds_dict = data.get("feeds", {})
            for feed_data in feeds_dict.values():
                datastreams = feed_data.get("datastreams", {})
                vsn700_feeds_points.update(datastreams.keys())

    if vsn300_livedata_path.exists():
        with open(vsn300_livedata_path) as f:
            data = json.load(f)
            for device_data in data.values():
                if isinstance(device_data, dict) and "points" in device_data:
                    for point in device_data["points"]:
                        vsn300_livedata_points.add(point["name"])

    print(f"  Found {len(vsn700_livedata_points)} VSN700 livedata points")
    print(f"  Found {len(vsn700_feeds_points)} VSN700 feeds points")
    print(f"  Found {len(vsn300_livedata_points)} VSN300 livedata points")

    return vsn700_livedata_points, vsn700_feeds_points, vsn300_livedata_points


def _process_sunspec_points(
    vsn700_livedata_points,
    vsn700_feeds_points,
    vsn300_livedata_points,
    sunspec_metadata,
    feeds_titles,
    processed_points,
):
    """Process standard SunSpec points from VSN_TO_SUNSPEC_MAP."""
    all_rows = []

    for vsn_name, mapping in VSN_TO_SUNSPEC_MAP.items():
        if (
            vsn_name in vsn700_livedata_points
            or vsn_name in vsn700_feeds_points
            or vsn_name in vsn300_livedata_points
        ):
            sunspec_name = mapping["sunspec"] if mapping["sunspec"] else vsn_name
            model = mapping["model"]

            # Lookup from SunSpec metadata
            label, workbook_description = lookup_label_description(
                sunspec_metadata, model, sunspec_name
            )

            if not label:
                label = generate_label_from_name(vsn_name)

            ha_name = generate_simplified_point_name(label, model)

            # Get description with priority
            feeds_info = feeds_titles.get(vsn_name)
            description, data_source, _ = get_description_with_priority(
                vsn_name, sunspec_name, model, label, workbook_description, feeds_info
            )

            # Determine VSN300 name
            if vsn_name in vsn300_livedata_points:
                vsn300_name = vsn_name
            elif mapping.get("modbus", "N/A") in vsn300_livedata_points:
                vsn300_name = mapping.get("modbus", "N/A")
            else:
                vsn300_name = "N/A"

            # Detect models
            models = detect_models_from_point(vsn300_name, vsn_name, model)

            # Create row
            row = create_row_with_model_flags(
                vsn700_name=vsn_name,
                vsn300_name=vsn300_name,
                sunspec_name=sunspec_name,
                ha_name=ha_name,
                in_livedata="✓" if vsn_name in vsn700_livedata_points else "",
                in_feeds="✓" if vsn_name in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=description if description != label else label,
                category=mapping["category"],
                units=mapping["units"],
                state_class=mapping["state_class"] or "",
                device_class=mapping["device_class"] or "",
                entity_category="",
                available_in_modbus=mapping["in_modbus"],
                data_source=data_source,
                model_notes="",
                models=models,
            )

            all_rows.append(row)
            processed_points.add(vsn_name)

    return all_rows


def _process_m64061_points(
    vsn700_livedata_points, vsn700_feeds_points, feeds_titles, processed_points
):
    """Process M64061 proprietary points."""
    all_rows = []

    for point_name in M64061_POINTS:
        if point_name in vsn700_livedata_points and point_name not in processed_points:
            label = generate_label_from_name(point_name)
            ha_name = generate_simplified_point_name(label, "M64061")

            # Get description
            feeds_info = feeds_titles.get(point_name)
            description, data_source, _ = get_description_with_priority(
                point_name, point_name, "M64061", label, "", feeds_info
            )

            # Apply VSN700 name normalization for duplicate detection
            normalized_name = normalize_vsn700_name(point_name)

            row = create_row_with_model_flags(
                vsn700_name=point_name,
                vsn300_name="N/A",
                sunspec_name=normalized_name,
                ha_name=ha_name,
                in_livedata="✓" if point_name in vsn700_livedata_points else "",
                in_feeds="✓" if point_name in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=description if description != label else label,
                category="Status",
                units=feeds_info.get("units", "") if feeds_info else "",
                state_class="",
                device_class="",
                entity_category="",
                available_in_modbus="NO",
                data_source=data_source,
                model_notes="ABB/FIMER proprietary model 64061",
                models={"M64061"},
            )

            all_rows.append(row)
            processed_points.add(point_name)

    return all_rows


def _process_abb_proprietary_points(
    vsn700_livedata_points,
    vsn700_feeds_points,
    vsn300_livedata_points,
    feeds_titles,
    processed_points,
):
    """Process ABB proprietary points."""
    all_rows = []

    for point_name in ABB_PROPRIETARY:
        in_vsn700 = (
            point_name in vsn700_livedata_points or point_name in vsn700_feeds_points
        )
        in_vsn300 = point_name in vsn300_livedata_points

        if (in_vsn700 or in_vsn300) and point_name not in processed_points:
            label = generate_label_from_name(point_name)
            ha_name = generate_simplified_point_name(label, "ABB_Proprietary")

            # Get description
            feeds_info = feeds_titles.get(point_name)
            description, data_source, _ = get_description_with_priority(
                point_name, point_name, "ABB_Proprietary", label, "", feeds_info
            )

            # Determine category
            category = "System"
            if "energy" in point_name.lower():
                category = "Energy Counter"
            elif any(kw in point_name.lower() for kw in ["alarm", "fault", "status"]):
                category = "Status"

            row = create_row_with_model_flags(
                vsn700_name=point_name if in_vsn700 else "N/A",
                vsn300_name=point_name if in_vsn300 else "N/A",
                sunspec_name=point_name,
                ha_name=ha_name,
                in_livedata="✓"
                if (
                    point_name in vsn700_livedata_points
                    or point_name in vsn300_livedata_points
                )
                else "",
                in_feeds="✓" if point_name in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=description if description != label else label,
                category=category,
                units=feeds_info.get("units", "") if feeds_info else "",
                state_class="",
                device_class="",
                entity_category="",
                available_in_modbus="NO",
                data_source=data_source,
                model_notes="ABB/FIMER proprietary",
                models={"ABB_Proprietary"},
            )

            all_rows.append(row)
            processed_points.add(point_name)

    return all_rows


def _process_vsn300_only_points(
    vsn300_livedata_points, sunspec_metadata, feeds_titles, processed_points
):
    """Process VSN300-only points not yet processed."""
    all_rows = []

    for point_name in vsn300_livedata_points:
        if point_name not in processed_points:
            # Extract model from VSN300 naming pattern
            model = None
            sunspec_name = point_name

            # Special handling for C_ prefix (Common Model M1 points)
            if point_name.startswith("C_"):
                model = "M1"
                sunspec_name = point_name[2:]
            elif point_name.startswith("m") and "_" in point_name:
                # Standard pattern m{model}_{instance}_{point}
                match = re.match(r"m(\d+)_(\d+)_(.+)", point_name)
                if match:
                    model_num, instance, sunspec_point = match.groups()
                    model = f"M{model_num}"
                    sunspec_name = sunspec_point

            # Apply VSN300 sunspec name normalization for duplicate detection
            sunspec_name = normalize_vsn700_name(sunspec_name)

            # Lookup from SunSpec metadata if we have a model
            label = ""
            workbook_description = ""
            if model and sunspec_metadata:
                label, workbook_description = lookup_label_description(
                    sunspec_metadata, model, sunspec_name
                )

            # If no label from SunSpec, generate from name
            if not label:
                label = generate_label_from_name(point_name)

            ha_name = generate_simplified_point_name(label, model or "VSN300_Only")

            # Get description
            feeds_info = feeds_titles.get(point_name)
            description, data_source, _ = get_description_with_priority(
                point_name, sunspec_name, model, label, workbook_description, feeds_info
            )

            # Detect models
            models = detect_models_from_point(point_name, None, model)
            if not models:
                models = {"VSN300_Only"}

            # Determine category
            model_flags = {
                flag: "YES" if flag in models else "NO" for flag in MODEL_FLAGS
            }
            category = get_comprehensive_category(
                label, description, model_flags, point_name, None
            )

            row = create_row_with_model_flags(
                vsn700_name="N/A",
                vsn300_name=point_name,
                sunspec_name=sunspec_name,
                ha_name=ha_name,
                in_livedata="✓",
                in_feeds="✓" if feeds_info else "",
                label=label,
                description=description,
                ha_display_name=description if description != label else label,
                category=category,
                units=feeds_info.get("units", "") if feeds_info else "",
                state_class="",
                device_class="",
                entity_category="",
                available_in_modbus="YES" if model else "NO",
                data_source=data_source,
                model_notes="VSN300-specific point",
                models=models,
            )

            all_rows.append(row)
            processed_points.add(point_name)

    return all_rows


def _process_vsn700_only_points(
    vsn700_livedata_points, vsn700_feeds_points, feeds_titles, processed_points
):
    """Process VSN700-only points not yet processed."""
    all_rows = []
    all_vsn700_points = vsn700_livedata_points.union(vsn700_feeds_points)

    for point_name in all_vsn700_points:
        if point_name not in processed_points:
            label = generate_label_from_name(point_name)
            ha_name = generate_simplified_point_name(label, "VSN700_Only")

            # Get description
            feeds_info = feeds_titles.get(point_name)
            description, data_source, _ = get_description_with_priority(
                point_name, point_name, None, label, "", feeds_info
            )

            # Determine category
            model_flags = {"VSN700_Only": "YES"}
            for flag in MODEL_FLAGS:
                if flag != "VSN700_Only":
                    model_flags[flag] = "NO"
            category = get_comprehensive_category(
                label, description, model_flags, None, point_name
            )

            # Apply VSN700 name normalization for duplicate detection
            normalized_name = normalize_vsn700_name(point_name)

            row = create_row_with_model_flags(
                vsn700_name=point_name,
                vsn300_name="N/A",
                sunspec_name=normalized_name,
                ha_name=ha_name,
                in_livedata="✓" if point_name in vsn700_livedata_points else "",
                in_feeds="✓" if point_name in vsn700_feeds_points else "",
                label=label,
                description=description,
                ha_display_name=description if description != label else label,
                category=category,
                units=feeds_info.get("units", "") if feeds_info else "",
                state_class="",
                device_class="",
                entity_category="",
                available_in_modbus="NO",
                data_source=data_source,
                model_notes="VSN700-specific point",
                models={"VSN700_Only"},
            )

            all_rows.append(row)
            processed_points.add(point_name)

    return all_rows


def _process_status_points(status_points, processed_points):
    """Process status points from /status endpoint."""
    all_rows = []

    for point_name, status_info in status_points.items():
        if point_name not in processed_points:
            label = (
                status_info["label"]
                if status_info["label"]
                else generate_label_from_name(point_name)
            )
            ha_name = generate_simplified_point_name(label, "Status")

            # Get description with priority
            description, data_source, _ = get_description_with_priority(
                point_name, point_name, "Status", label, "", None
            )

            # Get display name from improvements
            ha_display_name = DISPLAY_NAME_CORRECTIONS.get(description, label)

            # Determine category based on point name
            category = "Device Info"
            if "fw" in point_name.lower() or "version" in point_name.lower():
                category = "Device Info"
            elif (
                "wlan" in point_name.lower()
                or "ip" in point_name.lower()
                or "network" in point_name.lower()
            ):
                category = "Network"
            elif "device" in point_name.lower():
                category = "Device Info"
            elif "logger" in point_name.lower():
                category = "Datalogger"

            # Determine which VSN models have this point
            vsn700_name = point_name if status_info["vsn700"] else "N/A"
            vsn300_name = point_name if status_info["vsn300"] else "N/A"

            # Set models based on availability
            models = set()
            if status_info["vsn300"] and not status_info["vsn700"]:
                models.add("VSN300_Only")
            elif status_info["vsn700"] and not status_info["vsn300"]:
                models.add("VSN700_Only")
            else:
                models.add("ABB_Proprietary")

            row = create_row_with_model_flags(
                vsn700_name=vsn700_name,
                vsn300_name=vsn300_name,
                sunspec_name=point_name,
                ha_name=ha_name,
                in_livedata="",
                in_feeds="",
                label=label,
                description=description,
                ha_display_name=ha_display_name,
                category=category,
                units="",
                state_class="",
                device_class="",
                entity_category="diagnostic",
                available_in_modbus="NO",
                data_source=status_info["source"],
                model_notes="From /status endpoint",
                models=models,
            )

            all_rows.append(row)
            processed_points.add(point_name)

    return all_rows


def _get_cell_value_for_header(header, row_data):
    """Get the cell value for a given header from row data."""
    if header in MODEL_FLAGS:
        return row_data.get(header, "NO")
    if header == "REST Name (VSN700)":
        return row_data.get("vsn700_name", "")
    if header == "REST Name (VSN300)":
        return row_data.get("vsn300_name", "")
    if header == "SunSpec Normalized Name":
        return row_data.get("sunspec_name", "")
    if header == "HA Name":
        return row_data.get("ha_name", "")
    if header == "In /livedata":
        return row_data.get("in_livedata", "")
    if header == "In /feeds":
        return row_data.get("in_feeds", "")
    if header == "Label":
        return row_data.get("label", "")
    if header == "Description":
        return row_data.get("description", "")
    if header == "HA Display Name":
        return row_data.get("ha_display_name", "")
    if header == "Category":
        return row_data.get("category", "")
    if header == "HA Unit of Measurement":
        return row_data.get("units", "")
    if header == "HA State Class":
        return row_data.get("state_class", "")
    if header == "HA Device Class":
        return row_data.get("device_class", "")
    if header == "HA Icon":
        return row_data.get("icon", "")
    if header == "Entity Category":
        return row_data.get("entity_category", "")
    if header == "Suggested Display Precision":
        return row_data.get("precision", "")
    if header == "Available in Modbus":
        return row_data.get("available_in_modbus", "")
    if header == "Data Source":
        return row_data.get("data_source", "")
    if header == "Model_Notes":
        return row_data.get("model_notes", "")
    return ""


def _create_excel_workbook(deduplicated_rows, output_dir):
    """Create Excel workbook with data, formatting, and additional sheets."""
    print("\nCreating Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Points"

    # Add headers
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Style headers
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Add data rows
    for row_idx, row_data in enumerate(deduplicated_rows, 2):
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            value = _get_cell_value_for_header(header, row_data)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        if header in MODEL_FLAGS:
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
        elif "REST Name" in header:
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        elif header in ["Description", "HA Display Name"]:
            ws.column_dimensions[get_column_letter(col_idx)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Generate Summary sheet
    generate_summary_sheet(wb, deduplicated_rows)

    # Generate Changelog sheet
    generate_changelog_sheet(wb)

    # Save workbook
    output_path = output_dir / "vsn-sunspec-point-mapping.xlsx"
    wb.save(output_path)

    return output_path


def generate_mapping_excel_complete():
    """Generate complete mapping Excel with all customizations built-in."""
    print(f"\n{'=' * 70}")
    print(f"VSN-SunSpec Point Mapping Generator v{SCRIPT_VERSION}")
    print(f"Generated: {SCRIPT_DATE}")
    print(f"{'=' * 70}\n")

    # Define paths relative to script location
    data_dir = SCRIPT_DIR / "data"
    output_dir = SCRIPT_DIR / "output"

    # Ensure output directory exists
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load all data sources
    sunspec_metadata, m64061_metadata, feeds_titles, status_points = _load_data_sources(
        data_dir
    )

    # Extract point sets from VSN data files
    vsn700_livedata_points, vsn700_feeds_points, vsn300_livedata_points = (
        _extract_point_sets(data_dir)
    )

    # Process all points
    print("\nProcessing points...")
    processed_points = set()
    all_rows = []

    # Process each point type
    all_rows.extend(
        _process_sunspec_points(
            vsn700_livedata_points,
            vsn700_feeds_points,
            vsn300_livedata_points,
            sunspec_metadata,
            feeds_titles,
            processed_points,
        )
    )

    all_rows.extend(
        _process_m64061_points(
            vsn700_livedata_points, vsn700_feeds_points, feeds_titles, processed_points
        )
    )

    all_rows.extend(
        _process_abb_proprietary_points(
            vsn700_livedata_points,
            vsn700_feeds_points,
            vsn300_livedata_points,
            feeds_titles,
            processed_points,
        )
    )

    all_rows.extend(
        _process_vsn300_only_points(
            vsn300_livedata_points, sunspec_metadata, feeds_titles, processed_points
        )
    )

    all_rows.extend(
        _process_vsn700_only_points(
            vsn700_livedata_points, vsn700_feeds_points, feeds_titles, processed_points
        )
    )

    all_rows.extend(_process_status_points(status_points, processed_points))

    print(f"  Processed {len(all_rows)} total points (including status endpoints)")

    # Deduplication
    print("\nDeduplicating points...")
    rows_by_sunspec = defaultdict(list)
    for row in all_rows:
        sunspec_name = row.get("sunspec_name", "UNKNOWN")
        rows_by_sunspec[sunspec_name].append(row)

    deduplicated_rows = merge_duplicate_rows(rows_by_sunspec)
    print(
        f"  Deduplicated from {len(all_rows)} to {len(deduplicated_rows)} unique points"
    )

    # Create Excel workbook and save
    output_path = _create_excel_workbook(deduplicated_rows, output_dir)

    print(f"\n{'=' * 70}")
    print(f"✓ Excel file created: {output_path}")
    print(f"  Total rows: {len(deduplicated_rows)}")
    print("  Sheets: All Points, Summary, Changelog")
    print(f"{'=' * 70}\n")

    return output_path


if __name__ == "__main__":
    generate_mapping_excel_complete()
