"""Convert VSN-SunSpec point mapping Excel (v2 with model flags) to JSON format.

This script reads the new Excel structure with model flag columns and generates
a JSON file where each point has an array of applicable models.
"""

import json
from pathlib import Path

import openpyxl


def convert_excel_to_json_v2():
    """Convert v2 Excel mapping file (with model flags) to JSON format."""
    excel_path = Path("docs/vsn-sunspec-point-mapping-v2.xlsx")
    json_path = Path("docs/vsn-sunspec-point-mapping.json")

    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["All Points"]

    # Get column headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"Found {len(headers)} columns")

    # Model flag columns
    model_flags = ["M1", "M103", "M160", "M203", "M802", "M64061", "VSN300_Only", "VSN700_Only", "ABB_Proprietary"]

    # Convert rows to list of dictionaries
    rows = []
    for _row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        row_dict = {}
        models = []

        for _col_idx, (header, value) in enumerate(zip(headers, row, strict=False)):
            # Handle model flags specially - convert to array
            if header in model_flags:
                if value == "YES":
                    # Convert flag name to model name
                    if header == "ABB_Proprietary":
                        models.append("ABB Proprietary")
                    elif header == "VSN300_Only":
                        models.append("VSN300-only")
                    elif header == "VSN700_Only":
                        models.append("VSN700-only")
                    else:
                        models.append(header)
                # Don't add individual flags to row_dict
                continue

            # Handle other values
            if value is None:
                value = ""
            elif value == "YES":
                value = True
            elif value == "NO":
                value = False

            row_dict[header] = value

        # Add models array to row
        row_dict["models"] = models

        rows.append(row_dict)

    print(f"Converted {len(rows)} rows")

    # Write JSON file
    print(f"Writing JSON file: {json_path}")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    print(f"✓ JSON file created: {json_path}")
    print(f"  Total entries: {len(rows)}")

    # Also copy to integration data folder
    data_path = Path("custom_components/abb_fimer_pvi_vsn_rest/data/vsn-sunspec-point-mapping.json")
    print(f"Copying to integration data folder: {data_path}")

    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    print(f"✓ Integration data file updated: {data_path}")

    # Print statistics
    print("\nModel Statistics:")
    for model in model_flags:
        model_name = model
        if model == "ABB_Proprietary":
            model_name = "ABB Proprietary"
        elif model == "VSN300_Only":
            model_name = "VSN300-only"
        elif model == "VSN700_Only":
            model_name = "VSN700-only"

        count = sum(1 for row in rows if model_name in row["models"])
        print(f"  {model_name}: {count} points")


if __name__ == "__main__":
    convert_excel_to_json_v2()
