"""Analyze and deduplicate VSN-SunSpec point mapping.

This script:
1. Reads the current mapping JSON (277 rows with duplicates)
2. Groups by SunSpec Normalized Name to find duplicates
3. Analyzes quality issues (Description = Label, Category = Unknown)
4. Fetches SunSpec model definitions for official descriptions
5. Generates new Excel file with:
   - 210 unique rows
   - Model flag columns (M1, M103, M160, M203, M802, M64061, VSN300_Only, VSN700_Only, ABB_Proprietary)
   - Quality analysis
6. Generates comprehensive analysis report
"""

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_current_mapping() -> list[dict[str, Any]]:
    """Load current mapping JSON file."""
    json_path = Path("docs/vsn-sunspec-point-mapping.json")
    print(f"Loading current mapping from: {json_path}")

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    print(f"Loaded {len(data)} total points")
    return data


def analyze_duplicates(data: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Group points by SunSpec Normalized Name to find duplicates."""
    groups = defaultdict(list)

    for item in data:
        sunspec_name = item.get("SunSpec Normalized Name", "")
        if sunspec_name:
            groups[sunspec_name].append(item)

    print(f"\nFound {len(groups)} unique SunSpec names")

    duplicates = {name: items for name, items in groups.items() if len(items) > 1}
    print(f"Found {len(duplicates)} SunSpec names with duplicates")

    total_duplicate_rows = sum(len(items) - 1 for items in duplicates.values())
    print(f"Total duplicate rows that can be eliminated: {total_duplicate_rows}")

    return groups


def analyze_quality_issues(data: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Analyze quality issues in the mapping."""
    issues = {
        "description_equals_label": [],
        "unknown_category": [],
        "empty_description": [],
        "empty_label": [],
    }

    for item in data:
        label = item.get("Label", "").strip()
        description = item.get("Description", "").strip()
        category = item.get("Category", "").strip()

        # Check if Description = Label
        if label and description and label == description:
            issues["description_equals_label"].append(item)

        # Check if Category = Unknown
        if category.lower() == "unknown":
            issues["unknown_category"].append(item)

        # Check for empty descriptions
        if not description:
            issues["empty_description"].append(item)

        # Check for empty labels
        if not label:
            issues["empty_label"].append(item)

    print("\nQuality Issues:")
    print(f"  Description = Label: {len(issues['description_equals_label'])} ({len(issues['description_equals_label'])/len(data)*100:.1f}%)")
    print(f"  Category = Unknown: {len(issues['unknown_category'])} ({len(issues['unknown_category'])/len(data)*100:.1f}%)")
    print(f"  Empty Description: {len(issues['empty_description'])}")
    print(f"  Empty Label: {len(issues['empty_label'])}")

    return issues


def extract_model_from_string(model_str: str) -> str:
    """Extract model identifier from SunSpec Model string."""
    if not model_str:
        return ""

    model_str = model_str.strip()

    # Check for specific model patterns
    if model_str.startswith("M") and model_str[1:].replace("M", "").isdigit():
        return model_str.upper()
    if model_str == "VSN300-only":
        return "VSN300_Only"
    if model_str == "VSN700-only":
        return "VSN700_Only"
    if model_str == "ABB Proprietary":
        return "ABB_Proprietary"

    return model_str


def merge_duplicate_group(group: list[dict[str, Any]]) -> dict[str, Any]:
    """Merge a group of duplicate points into a single deduplicated entry."""
    if not group:
        return {}

    # Start with the first item as base
    merged = dict(group[0])

    # Extract models from all items in group
    models = set()

    for item in group:
        model_str = item.get("SunSpec Model", "")
        model = extract_model_from_string(model_str)
        if model:
            models.add(model)

    # Initialize model flags
    model_flags = {
        "M1": "NO",
        "M103": "NO",
        "M160": "NO",
        "M203": "NO",
        "M802": "NO",
        "M64061": "NO",
        "VSN300_Only": "NO",
        "VSN700_Only": "NO",
        "ABB_Proprietary": "NO",
    }

    # Set flags based on models found
    for model in models:
        if model in model_flags:
            model_flags[model] = "YES"

    # Add model flags to merged entry
    merged.update(model_flags)

    # Remove old SunSpec Model column
    if "SunSpec Model" in merged:
        del merged["SunSpec Model"]

    # Add Model_Notes field
    merged["Model_Notes"] = ""

    # Determine quality flags
    label = merged.get("Label", "").strip()
    description = merged.get("Description", "").strip()
    category = merged.get("Category", "").strip()

    merged["Description_Quality"] = "NEEDS_FIX" if label == description else "GOOD"
    merged["Category_Status"] = "UNKNOWN" if category.lower() == "unknown" else "ASSIGNED"

    return merged


def suggest_description_improvements(merged_data: list[dict[str, Any]]) -> dict[str, str]:
    """Generate description improvement suggestions."""
    suggestions = {}

    # Manual fixes for known issues
    manual_fixes = {
        "CellTmpMax": "Maximum cell temperature in battery pack (°C)",
        "CellTmpMin": "Minimum cell temperature in battery pack (°C)",
        "ECt": "Total energy charged to battery since installation (Wh)",
        "EDt": "Total energy discharged from battery since installation (Wh)",
        "HouseIgrid_L1": "Household current consumption from grid on phase L1 (A)",
        "HouseIgrid_L2": "Household current consumption from grid on phase L2 (A)",
        "HouseIgrid_L3": "Household current consumption from grid on phase L3 (A)",
        "HousePgrid_L1": "Household active power consumption from grid on phase L1 (W)",
        "HousePgrid_L2": "Household active power consumption from grid on phase L2 (W)",
        "HousePgrid_L3": "Household active power consumption from grid on phase L3 (W)",
    }

    for item in merged_data:
        sunspec_name = item.get("SunSpec Normalized Name", "")
        ha_name = item.get("HA Name", "")
        label = item.get("Label", "")
        item.get("Description", "")

        # Check if needs improvement
        if item.get("Description_Quality") == "NEEDS_FIX":
            # Check manual fixes first
            if sunspec_name in manual_fixes:
                suggestions[sunspec_name] = manual_fixes[sunspec_name]
            elif ha_name in manual_fixes:
                suggestions[ha_name] = manual_fixes[ha_name]
            else:
                # Generate basic suggestion based on label
                suggestions[sunspec_name or ha_name] = f"{label} measurement point"

    return suggestions


def suggest_category_fixes(merged_data: list[dict[str, Any]]) -> dict[str, str]:
    """Suggest category fixes for Unknown categories."""
    suggestions = {}

    for item in merged_data:
        if item.get("Category_Status") == "UNKNOWN":
            ha_name = item.get("HA Name", "")
            sunspec_name = item.get("SunSpec Normalized Name", "")
            label = item.get("Label", "")

            # Category rules
            suggested_category = "Unknown"

            # Energy counter patterns
            if (ha_name.startswith(("e0_", "e1_", "e2_", "e3_", "e4_", "e5_", "e6_", "e7_", "e8_")) or "energy" in label.lower() or "charge" in label.lower() or "discharge" in label.lower()):
                suggested_category = "Energy Counter"

            # House meter patterns
            elif "house" in ha_name or "house" in label.lower():
                suggested_category = "House Meter"

            # System monitoring patterns
            elif ha_name in ("flash_free", "free_ram", "fw_ver", "store_size", "sys_load", "uptime"):
                suggested_category = "System Monitoring"

            # Phase measurements (check models)
            elif sunspec_name in ("AphA", "AphB", "AphC", "PhVphA", "PhVphB", "PhVphC"):
                # Check if M103/M160 (Inverter) or M201/M203 (Meter)
                if item.get("M103") == "YES" or item.get("M160") == "YES":
                    suggested_category = "Inverter"
                elif item.get("M203") == "YES":
                    suggested_category = "Meter"

            # Power measurements
            elif sunspec_name in ("W", "VAR", "VA", "Hz", "PF"):
                if item.get("M103") == "YES":
                    suggested_category = "Inverter"
                elif item.get("M203") == "YES":
                    suggested_category = "Meter"

            if suggested_category != "Unknown":
                suggestions[ha_name or sunspec_name] = suggested_category

    return suggestions


def create_excel_report(merged_data: list[dict[str, Any]], analysis: dict[str, Any], output_path: Path):
    """Create Excel file with deduplicated data and analysis."""
    print(f"\nCreating Excel file: {output_path}")

    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet 1: All Points (deduplicated)
    ws_points = wb.create_sheet("All Points", 0)
    create_all_points_sheet(ws_points, merged_data)

    # Sheet 2: Quality Report
    ws_quality = wb.create_sheet("Quality Report", 1)
    create_quality_report_sheet(ws_quality, analysis)

    # Sheet 3: Model Coverage
    ws_coverage = wb.create_sheet("Model Coverage", 2)
    create_model_coverage_sheet(ws_coverage, merged_data)

    # Save workbook
    wb.save(output_path)
    print("✓ Excel file created successfully")


def create_all_points_sheet(ws, merged_data: list[dict[str, Any]]):
    """Create the main points sheet with deduplicated data."""
    # Define column order (new structure)
    columns = [
        # Model Flags (first for visibility)
        "M1", "M103", "M160", "M203", "M802", "M64061",
        "VSN300_Only", "VSN700_Only", "ABB_Proprietary",

        # REST API Names
        "REST Name (VSN700)", "REST Name (VSN300)",

        # SunSpec Info
        "SunSpec Normalized Name", "HA Name",

        # Availability
        "In /livedata", "In /feeds",

        # Display Names
        "Label", "Description", "HA Display Name",

        # Classification
        "Category",

        # HA Configuration
        "HA Unit of Measurement",
        "HA State Class",
        "HA Device Class",
        "Entity Category",

        # Metadata
        "Available in Modbus", "Data Source",

        # Quality Flags
        "Description_Quality", "Category_Status",

        # Notes
        "Model_Notes",
    ]

    # Write headers
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, item in enumerate(merged_data, start=2):
        for col_idx, column in enumerate(columns, start=1):
            value = item.get(column, "")

            # Convert boolean True/False to YES/NO if needed
            if isinstance(value, bool):
                value = "YES" if value else "NO"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Highlight quality issues
            if (column == "Description_Quality" and value == "NEEDS_FIX") or (column == "Category_Status" and value == "UNKNOWN"):
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Freeze header row
    ws.freeze_panes = "A2"


def create_quality_report_sheet(ws, analysis: dict[str, Any]):
    """Create quality analysis report sheet."""
    # Title
    ws.cell(row=1, column=1, value="VSN-SunSpec Mapping Quality Report").font = Font(bold=True, size=14)

    row = 3

    # Statistics
    ws.cell(row=row, column=1, value="STATISTICS").font = Font(bold=True)
    row += 1

    stats = analysis["statistics"]
    for key, value in stats.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Description Quality Issues
    ws.cell(row=row, column=1, value="DESCRIPTION QUALITY ISSUES (Description = Label)").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="SunSpec Name").font = Font(bold=True)
    ws.cell(row=row, column=2, value="HA Name").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Current Description").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Suggested Description").font = Font(bold=True)
    row += 1

    for sunspec_name, suggestion in analysis["description_suggestions"].items():
        # Find the item
        for item in analysis["merged_data"]:
            if item.get("SunSpec Normalized Name") == sunspec_name or item.get("HA Name") == sunspec_name:
                ws.cell(row=row, column=1, value=item.get("SunSpec Normalized Name", ""))
                ws.cell(row=row, column=2, value=item.get("HA Name", ""))
                ws.cell(row=row, column=3, value=item.get("Description", ""))
                ws.cell(row=row, column=4, value=suggestion)
                row += 1
                break

    row += 2

    # Category Fixes
    ws.cell(row=row, column=1, value="CATEGORY FIXES (Unknown → Suggested)").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="HA Name").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Current Category").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Suggested Category").font = Font(bold=True)
    row += 1

    for ha_name, suggestion in analysis["category_suggestions"].items():
        ws.cell(row=row, column=1, value=ha_name)
        ws.cell(row=row, column=2, value="Unknown")
        ws.cell(row=row, column=3, value=suggestion)
        row += 1

    # Auto-size columns
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 50


def create_model_coverage_sheet(ws, merged_data: list[dict[str, Any]]):
    """Create model coverage matrix sheet."""
    # Title
    ws.cell(row=1, column=1, value="Model Coverage Matrix").font = Font(bold=True, size=14)

    row = 3

    # Count points per model
    models = ["M1", "M103", "M160", "M203", "M802", "M64061", "VSN300_Only", "VSN700_Only", "ABB_Proprietary"]

    ws.cell(row=row, column=1, value="Model").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Point Count").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Percentage").font = Font(bold=True)
    row += 1

    total_points = len(merged_data)

    for model in models:
        count = sum(1 for item in merged_data if item.get(model) == "YES")
        percentage = (count / total_points * 100) if total_points > 0 else 0

        ws.cell(row=row, column=1, value=model)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        row += 1

    # Auto-size columns
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


def main():
    """Main execution function."""
    print("=" * 80)
    print("VSN-SunSpec Mapping Analysis and Deduplication")
    print("=" * 80)

    # Load current mapping
    current_data = load_current_mapping()

    # Analyze duplicates
    groups = analyze_duplicates(current_data)

    # Analyze quality issues
    issues = analyze_quality_issues(current_data)

    # Merge duplicates
    print("\nMerging duplicates...")
    merged_data = []
    for _sunspec_name, group in sorted(groups.items()):
        merged = merge_duplicate_group(group)
        if merged:
            merged_data.append(merged)

    print(f"✓ Created {len(merged_data)} unique points (reduced from {len(current_data)})")

    # Generate improvement suggestions
    print("\nGenerating improvement suggestions...")
    description_suggestions = suggest_description_improvements(merged_data)
    category_suggestions = suggest_category_fixes(merged_data)

    print(f"  Description improvements: {len(description_suggestions)}")
    print(f"  Category fixes: {len(category_suggestions)}")

    # Prepare analysis data
    analysis = {
        "statistics": {
            "Total points (original)": len(current_data),
            "Unique SunSpec names": len(groups),
            "Duplicate rows eliminated": len(current_data) - len(merged_data),
            "Points needing description fix": len(issues["description_equals_label"]),
            "Points needing category fix": len(issues["unknown_category"]),
            "Final point count": len(merged_data),
            "Reduction percentage": f"{((len(current_data) - len(merged_data)) / len(current_data) * 100):.1f}%",
        },
        "merged_data": merged_data,
        "description_suggestions": description_suggestions,
        "category_suggestions": category_suggestions,
    }

    # Create Excel report
    output_path = Path("docs/vsn-sunspec-point-mapping.xlsx")
    create_excel_report(merged_data, analysis, output_path)

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Original points: {len(current_data)}")
    print(f"Deduplicated points: {len(merged_data)}")
    print(f"Rows eliminated: {len(current_data) - len(merged_data)} ({((len(current_data) - len(merged_data)) / len(current_data) * 100):.1f}%)")
    print(f"\nOutput file: {output_path}")
    print("\nNext steps:")
    print(f"  1. Review {output_path} - Sheet: 'All Points'")
    print("  2. Check 'Quality Report' sheet for suggested improvements")
    print("  3. Apply description and category fixes")
    print("  4. Verify model flags are correct")
    print("  5. Save and proceed with integration updates")
    print("=" * 80)


if __name__ == "__main__":
    main()
