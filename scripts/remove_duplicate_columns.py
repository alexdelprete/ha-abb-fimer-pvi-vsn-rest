"""Remove duplicate columns from mapping Excel file.

This script removes the duplicate non-HA-prefixed columns:
- "Units" (keeps "HA Unit of Measurement")
- "State Class" (keeps "HA State Class")
- "Device Class" (keeps "HA Device Class")
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def remove_duplicate_columns():
    """Remove duplicate columns from the mapping Excel file."""
    # File path
    excel_path = Path(__file__).parent.parent / "docs" / "vsn-sunspec-point-mapping-v2.xlsx"

    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find columns to remove
    columns_to_remove = ["Units", "State Class", "Device Class"]
    column_indices_to_remove = []

    # Get header row
    headers = [cell.value for cell in ws[1]]
    print(f"\nFound {len(headers)} columns in header row")

    # Find indices of columns to remove
    for col_name in columns_to_remove:
        try:
            col_idx = headers.index(col_name) + 1  # Excel is 1-indexed
            column_indices_to_remove.append(col_idx)
            print(f"  - Found '{col_name}' at column index {col_idx}")
        except ValueError:
            print(f"  - Warning: Column '{col_name}' not found in header")

    # Sort in reverse order to delete from right to left (avoid index shifting)
    column_indices_to_remove.sort(reverse=True)

    # Delete columns
    print(f"\nRemoving {len(column_indices_to_remove)} duplicate columns...")
    for col_idx in column_indices_to_remove:
        col_name = headers[col_idx - 1]
        print(f"  - Deleting column {col_idx}: '{col_name}'")
        ws.delete_cols(col_idx)

    # Save the modified workbook
    print("\nSaving modified Excel file...")
    wb.save(excel_path)
    print(f"✓ Successfully removed duplicate columns from {excel_path}")

    # Show remaining columns
    headers_after = [cell.value for cell in ws[1]]
    print(f"\nRemaining columns ({len(headers_after)}):")
    for i, header in enumerate(headers_after, 1):
        if header in ["HA Unit of Measurement", "HA State Class", "HA Device Class"]:
            print(f"  {i:2d}. {header} ✓")
        else:
            print(f"  {i:2d}. {header}")


if __name__ == "__main__":
    remove_duplicate_columns()
