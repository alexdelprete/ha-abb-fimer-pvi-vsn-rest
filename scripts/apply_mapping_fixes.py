"""Apply description and category fixes to deduplicated mapping Excel.

This script:
1. Loads the v2 Excel file (210 deduplicated rows)
2. Applies description improvements for all NEEDS_FIX points
3. Categorizes all Unknown points using comprehensive logic
4. Removes quality flag columns
5. Saves the final clean Excel file
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


# Comprehensive description improvements
DESCRIPTION_FIXES = {
    # User-specified priority fixes
    "CellTmpMax": "Maximum cell temperature in battery pack",
    "CellTmpMin": "Minimum cell temperature in battery pack",
    "ECt": "Total energy charged to battery since installation",
    "EDt": "Total energy discharged from battery since installation",

    # Phase-specific current fixes
    "HouseIgrid_L1": "Household current consumption from grid on phase L1",
    "HouseIgrid_L2": "Household current consumption from grid on phase L2",
    "HouseIgrid_L3": "Household current consumption from grid on phase L3",

    # Phase-specific power fixes
    "HousePgrid_L1": "Household active power consumption from grid on phase L1",
    "HousePgrid_L2": "Household active power consumption from grid on phase L2",
    "HousePgrid_L3": "Household active power consumption from grid on phase L3",

    # SunSpec M1 (Common) descriptions
    "Mn": "Manufacturer name from device common model",
    "Md": "Device model identifier from common model",
    "Opt": "Device options and features from common model",
    "Vr": "Firmware version from device common model",
    "SN": "Device serial number from common model",
    "DA": "Device Modbus address",

    # Phase voltage descriptions
    "PhVphA": "Phase A to neutral voltage measurement",
    "PhVphB": "Phase B to neutral voltage measurement",
    "PhVphC": "Phase C to neutral voltage measurement",

    # Phase current descriptions
    "AphA": "Phase A current measurement",
    "AphB": "Phase B current measurement",
    "AphC": "Phase C current measurement",

    # DC measurements
    "DCA_1": "DC current measurement for string 1",
    "DCV_1": "DC voltage measurement for string 1",
    "DCW_1": "DC power output for string 1",
    "DCA_2": "DC current measurement for string 2",
    "DCV_2": "DC voltage measurement for string 2",
    "DCW_2": "DC power output for string 2",

    # Battery specific
    "Tmp": "Battery pack temperature",
    "BattNum": "Number of battery modules in the system",
    "Chc": "Battery charge cycle count",
    "DetectedBatteryNumber": "Number of detected battery modules",
    "Dhc": "Battery discharge cycle count",

    # Energy counters - 30 day
    "E0_30D": "Total energy produced in last 30 days",
    "E1_30D": "Energy produced by string 1 in last 30 days",
    "E2_30D": "Energy produced by string 2 in last 30 days",
    "E3_30D": "Energy from auxiliary source in last 30 days",
    "E4_30D": "Energy counter 4 accumulated over 30 days",
    "E5_30D": "Energy counter 5 accumulated over 30 days",
    "E6_30D": "Total consumption energy in last 30 days",
    "E7_30D": "Grid export energy in last 30 days",
    "E8_30D": "Grid import energy in last 30 days",

    # Energy counters - 7 day
    "E0_7D": "Total energy produced in last 7 days",
    "E1_7D": "Energy produced by string 1 in last 7 days",
    "E2_7D": "Energy produced by string 2 in last 7 days",
    "E3_7D": "Energy from auxiliary source in last 7 days",
    "E4_7D": "Energy counter 4 accumulated over 7 days",
    "E5_7D": "Energy counter 5 accumulated over 7 days",
    "E6_7D": "Total consumption energy in last 7 days",
    "E7_7D": "Grid export energy in last 7 days",
    "E8_7D": "Grid import energy in last 7 days",

    # Energy counters - runtime
    "E0_runtime": "Total energy produced since commissioning",
    "E1_runtime": "Energy produced by string 1 since commissioning",
    "E2_runtime": "Energy produced by string 2 since commissioning",
    "E3_runtime": "Energy from auxiliary source since commissioning",
    "E4_runtime": "Energy counter 4 lifetime total",
    "E5_runtime": "Energy counter 5 lifetime total",
    "E6_runtime": "Total consumption energy since commissioning",
    "E7_runtime": "Grid export energy since commissioning",
    "E8_runtime": "Grid import energy since commissioning",

    # Battery energy
    "ECharge": "Energy charged to battery in current period",
    "EDischarge": "Energy discharged from battery in current period",
    "ETotCharge": "Total energy charged to battery since commissioning",
    "ETotDischarge": "Total energy discharged from battery since commissioning",

    # System info
    "SplitPhase": "Split phase configuration indicator",
    "CountryStd": "Country-specific electrical standard code",
}


def get_comprehensive_category(row_data: dict) -> str:
    """Determine category using comprehensive logic."""
    ha_name = row_data.get("HA Name", "").lower()
    sunspec_name = row_data.get("SunSpec Normalized Name", "")
    label = row_data.get("Label", "").lower()

    # Energy counter patterns (most specific first)
    energy_patterns = [
        "e0_", "e1_", "e2_", "e3_", "e4_", "e5_", "e6_", "e7_", "e8_",
        "echarge", "edischarge", "etotcharge", "etotdischarge",
        "ein", "eout"
    ]
    if any(ha_name.startswith(p) for p in energy_patterns) or "energy" in label:
        return "Energy Counter"

    # House meter patterns
    if "house" in ha_name or "house" in label or "HousePgrid" in sunspec_name or "HouseIgrid" in sunspec_name:
        return "House Meter"

    # System monitoring patterns
    system_patterns = ["flash_free", "free_ram", "fw_ver", "store_size", "sys_load", "uptime"]
    if ha_name in system_patterns or any(p in ha_name for p in ["flash", "ram", "uptime", "load"]):
        return "System Monitoring"

    # Battery patterns
    battery_patterns = ["battery", "batt", "cell", "soc", "soh", "charge", "discharge"]
    if any(p in ha_name for p in battery_patterns) or any(p in label for p in battery_patterns):
        # Check if it's actually an energy counter
        if not any(ha_name.startswith(p) for p in energy_patterns):
            return "Battery"

    # Phase measurements - check models to determine if Inverter or Meter
    phase_points = ["apha", "aphb", "aphc", "phvpha", "phvphb", "phvphc"]
    if sunspec_name.lower() in phase_points or ha_name in phase_points:
        # Check model flags
        if row_data.get("M103") == "YES" or row_data.get("M160") == "YES":
            return "Inverter"
        elif row_data.get("M203") == "YES" or row_data.get("M212") == "YES":
            return "Meter"
        else:
            # Default to Inverter if unclear
            return "Inverter"

    # Power/frequency measurements
    power_points = ["w", "var", "va", "hz", "pf"]
    if sunspec_name.lower() in power_points:
        if row_data.get("M103") == "YES":
            return "Inverter"
        elif row_data.get("M203") == "YES":
            return "Meter"

    # DC measurements (definitely Inverter)
    if ha_name.startswith("dc") or "string" in label:
        return "Inverter"

    # Device info (M1 common model)
    device_info_points = ["mn", "md", "opt", "vr", "sn", "da"]
    if sunspec_name.lower() in device_info_points or ha_name in device_info_points:
        return "Device Info"

    # Control/Status
    if "status" in ha_name or "state" in ha_name or "alarm" in ha_name or "event" in ha_name:
        return "Status"

    # VSN-specific system points
    if row_data.get("VSN300_Only") == "YES" or row_data.get("VSN700_Only") == "YES":
        if any(p in ha_name for p in ["device", "manufacturer", "model", "serial", "address"]):
            return "Device Info"
        else:
            return "System Monitoring"

    # Default based on model
    if row_data.get("M103") == "YES" or row_data.get("M160") == "YES":
        return "Inverter"
    elif row_data.get("M203") == "YES":
        return "Meter"
    elif row_data.get("M802") == "YES":
        return "Battery"

    # Last resort: keep Unknown (will need manual review)
    return "Unknown"


def apply_fixes_to_excel(input_path: Path, output_path: Path):
    """Load Excel, apply all fixes, and save."""
    print(f"Loading Excel file: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["All Points"]

    # Find column indices
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        headers[cell.value] = col_idx

    print(f"Found {len(headers)} columns")

    # Track statistics
    stats = {
        "descriptions_fixed": 0,
        "categories_fixed": 0,
        "total_rows": ws.max_row - 1,  # Exclude header
    }

    # Process each row
    for row_idx in range(2, ws.max_row + 1):
        # Read row data
        row_data = {}
        for header, col_idx in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[header] = cell_value if cell_value is not None else ""

        sunspec_name = row_data.get("SunSpec Normalized Name", "")
        ha_name = row_data.get("HA Name", "")
        description = row_data.get("Description", "")
        label = row_data.get("Label", "")
        category = row_data.get("Category", "")
        description_quality = row_data.get("Description_Quality", "")
        category_status = row_data.get("Category_Status", "")

        # Fix description if needed
        if description_quality == "NEEDS_FIX":
            # Try sunspec_name first, then ha_name
            new_description = DESCRIPTION_FIXES.get(sunspec_name) or DESCRIPTION_FIXES.get(ha_name)

            if new_description:
                ws.cell(row=row_idx, column=headers["Description"], value=new_description)
                ws.cell(row=row_idx, column=headers["HA Display Name"], value=new_description)
                stats["descriptions_fixed"] += 1
                print(f"  Fixed description for {sunspec_name or ha_name}: {label} → {new_description}")

        # Fix category if unknown
        if category_status == "UNKNOWN" or category.lower() == "unknown":
            new_category = get_comprehensive_category(row_data)
            ws.cell(row=row_idx, column=headers["Category"], value=new_category)
            if new_category != "Unknown":
                stats["categories_fixed"] += 1
                print(f"  Fixed category for {ha_name}: Unknown → {new_category}")

    # Remove quality flag columns
    print("\nRemoving quality flag columns...")
    cols_to_remove = ["Description_Quality", "Category_Status"]

    for col_name in cols_to_remove:
        if col_name in headers:
            col_idx = headers[col_name]
            ws.delete_cols(col_idx)
            print(f"  Removed column: {col_name}")
            # Update headers dict (shift indices after deleted column)
            headers = {k: (v - 1 if v > col_idx else v) for k, v in headers.items() if k != col_name}

    # Save updated workbook
    print(f"\nSaving updated Excel: {output_path}")
    wb.save(output_path)

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total rows processed: {stats['total_rows']}")
    print(f"Descriptions fixed: {stats['descriptions_fixed']}")
    print(f"Categories fixed: {stats['categories_fixed']}")
    print(f"Output file: {output_path}")
    print("=" * 80)

    return stats


def main():
    """Main execution function."""
    print("=" * 80)
    print("Applying Description and Category Fixes to Mapping")
    print("=" * 80)

    input_path = Path("docs/vsn-sunspec-point-mapping-v2.xlsx")
    output_path = Path("docs/vsn-sunspec-point-mapping-v2-fixed.xlsx")

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        return

    stats = apply_fixes_to_excel(input_path, output_path)

    print("\nNext steps:")
    print("  1. Review output file for any remaining 'Unknown' categories")
    print("  2. Verify descriptions are appropriate")
    print("  3. Replace original v2 file if satisfied")
    print("  4. Proceed with JSON regeneration")


if __name__ == "__main__":
    main()
