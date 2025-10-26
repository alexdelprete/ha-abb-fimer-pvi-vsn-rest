#!/usr/bin/env python3
"""Convert VSN-SunSpec mapping from Excel to JSON format.

This script converts the vsn-sunspec-point-mapping.xlsx file to a JSON format
that can be used by the normalizer without requiring openpyxl dependency.

Usage:
    python generate_mapping_json.py
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def convert_mapping_to_json():
    """Convert Excel mapping to JSON format."""
    script_dir = Path(__file__).parent
    excel_file = script_dir.parent / "docs" / "vsn-sunspec-point-mapping.xlsx"
    json_file = script_dir.parent / "docs" / "vsn-sunspec-point-mapping.json"

    print(f"Loading Excel file: {excel_file}")
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb["All Points"]

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)

    print(f"Found {len(headers)} columns")

    # Convert rows to list of dicts
    mappings = []
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue

        # Create dict from row
        mapping = {}
        for i, header in enumerate(headers):
            if i < len(row):
                value = row[i]
                # Convert to appropriate type
                if value is None:
                    mapping[header] = None
                elif isinstance(value, (int, float)):
                    mapping[header] = value
                else:
                    mapping[header] = str(value).strip() if value else None

        mappings.append(mapping)
        row_count += 1

    wb.close()

    print(f"Converted {row_count} mapping rows")

    # Write to JSON
    print(f"Writing JSON file: {json_file}")
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)

    print(f"✓ Successfully created {json_file}")
    print(f"  - {row_count} mappings")
    print(f"  - File size: {json_file.stat().st_size / 1024:.1f} KB")

    # Show statistics
    vsn300_count = sum(1 for m in mappings if m.get("VSN300 Name"))
    vsn700_count = sum(1 for m in mappings if m.get("VSN700 Name"))
    both_count = sum(1 for m in mappings if m.get("VSN300 Name") and m.get("VSN700 Name"))

    print(f"\nStatistics:")
    print(f"  - VSN300 points: {vsn300_count}")
    print(f"  - VSN700 points: {vsn700_count}")
    print(f"  - Shared points: {both_count}")
    print(f"  - VSN300-only: {vsn300_count - both_count}")
    print(f"  - VSN700-only: {vsn700_count - both_count}")


if __name__ == "__main__":
    convert_mapping_to_json()
