"""Convert VSN-SunSpec point mapping Excel file to JSON format."""

import json
from pathlib import Path

import openpyxl


def convert_excel_to_json():
    """Convert Excel mapping file to JSON format."""
    # Input and output paths
    excel_path = Path("docs/vsn-sunspec-point-mapping.xlsx")
    json_path = Path("docs/vsn-sunspec-point-mapping.json")

    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["All Points"]

    # Get column headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"Found {len(headers)} columns: {', '.join(headers)}")

    # Convert rows to list of dictionaries
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not any(row):
            continue

        # Create dictionary from row
        row_dict = {}
        for col_idx, (header, value) in enumerate(zip(headers, row)):
            # Convert None to empty string for consistency
            if value is None:
                value = ""
            # Convert booleans
            elif value == "YES":
                value = True
            elif value == "NO":
                value = False

            row_dict[header] = value

        rows.append(row_dict)

    print(f"Converted {len(rows)} rows")

    # Write JSON file
    print(f"Writing JSON file: {json_path}")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    print(f"✓ JSON file created: {json_path}")
    print(f"  Total entries: {len(rows)}")

    # Also copy to integration data folder
    data_path = Path("custom_components/abb_fimer_pvi_vsn_rest/data/vsn-sunspec-point-mapping.json")
    print(f"Copying to integration data folder: {data_path}")
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)

    print(f"✓ Integration data file updated: {data_path}")


if __name__ == "__main__":
    convert_excel_to_json()
